from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Depends, Form, Body
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient, AsyncIOMotorGridFSBucket
import copy
import pypdf
import os
import io
import base64
import logging
import bcrypt
import secrets
import shutil
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
from datetime import datetime, timezone, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
try:
    pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
except Exception as e:
    print(f"Font registration failed: {e}")
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, Frame, KeepInFrame, Table, TableStyle
from auth import get_current_user_safe

ROOT_DIR = Path(__file__).parent

pdfmetrics.registerFont(TTFont('Roboto', str(Path(__file__).parent / 'fonts' / 'Roboto-Regular.ttf')))
pdfmetrics.registerFont(TTFont('Roboto-Bold', str(Path(__file__).parent / 'fonts' / 'Roboto-Bold.ttf')))
pdfmetrics.registerFont(TTFont('Roboto-Medium', str(Path(__file__).parent / 'fonts' / 'Roboto-Medium.ttf')))
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Creator Consultant API")

# Auth: /api/auth/* — public (login) + JWT-protected (others)
import auth as auth_module  # noqa: E402
auth_module.init(db)
auth_public_router = APIRouter(prefix="/api")
auth_public_router.include_router(auth_module.router)

# Main API router — every endpoint here requires a valid JWT
api_router = APIRouter(prefix="/api", dependencies=[Depends(auth_module.get_current_user)])


# ---------------------- HELPERS ----------------------
GST_STATE_CODES = {
  "01": "Jammu and Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh",
  "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh",
  "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur",
  "15": "Mizoram", "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal",
  "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat",
  "25": "Daman and Diu", "26": "Dadra and Nagar Haveli", "27": "Maharashtra", "28": "Andhra Pradesh (Old)",
  "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala", "33": "Tamil Nadu",
  "34": "Puducherry", "35": "Andaman and Nicobar Islands", "36": "Telangana", "37": "Andhra Pradesh",
  "38": "Ladakh"
}

def format_place_of_supply(pos: str, gstin: str = "") -> str:
    code = (pos or "").strip()
    if not code and gstin and len(gstin) >= 2:
        code = gstin[:2]
    if not code:
        return ""
    if code.isdigit():
        code = code.zfill(2)
    return GST_STATE_CODES.get(code, code)

# ---------------------- MODELS ----------------------
class ClientIn(BaseModel):
    name: str
    phone: Optional[str] = ""
    email: Optional[str] = ""
    company: Optional[str] = ""
    address: Optional[str] = ""
    gstin: Optional[str] = ""
    pan: Optional[str] = ""
    place_of_supply: Optional[str] = ""
    gst_type: Optional[str] = ""
    principal_address: Optional[str] = ""


class Client(ClientIn):
    model_config = ConfigDict(extra="ignore")
    id: str
    last_edited_by_user_id: Optional[str] = None
    last_edited_by_username: Optional[str] = ""
    last_edited_at: Optional[str] = ""
    created_at: str


class PaginatedClients(BaseModel):
    data: List[Client]
    total: int


class ArchitectIn(BaseModel):
    name: str
    phone: Optional[str] = ""
    email: Optional[str] = ""
    firm: Optional[str] = ""


class Architect(ArchitectIn):
    model_config = ConfigDict(extra="ignore")
    id: str
    last_edited_by_user_id: Optional[str] = None
    last_edited_by_username: Optional[str] = ""
    last_edited_at: Optional[str] = ""
    created_at: str


class ProjectIn(BaseModel):
    name: str
    job_no: Optional[str] = ""
    client_id: Optional[str] = None
    architect_id: Optional[str] = None
    site_location: Optional[str] = ""
    quoted_amount: float = 0.0
    status: Optional[str] = "Outstanding"  # Outstanding / Settled
    notes: Optional[str] = ""
    assigned_engineer_ids: List[str] = []


class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    project_code: str  # e.g. CC-0001
    job_no: str = ""
    name: str
    client_id: Optional[str] = None
    client_name: Optional[str] = ""
    client_phone: Optional[str] = ""
    client_email: Optional[str] = ""
    architect_id: Optional[str] = None
    architect_name: Optional[str] = ""
    architect_phone: Optional[str] = ""
    architect_email: Optional[str] = ""
    site_location: str = ""
    quoted_amount: Optional[float] = 0.0
    received_amount: Optional[float] = 0.0
    outstanding_amount: Optional[float] = 0.0
    status: str = "Outstanding"
    notes: str = ""
    archived: bool = False
    assigned_engineer_ids: List[str] = []
    # Offer linkage (optional — filled when a project is created from an offer)
    offer_id: Optional[str] = None
    offer_code: Optional[str] = ""
    offer_type: Optional[str] = ""
    offer_file_path: Optional[str] = ""
    # Edit tracking
    last_edited_by_user_id: Optional[str] = None
    last_edited_by_username: Optional[str] = ""
    created_at: str


class PaginatedProjects(BaseModel):
    data: List[Project]
    total: int


class OfferIn(BaseModel):
    offer_type: str  # RCC / Steel / Audit / Other
    custom_type: Optional[str] = ""  # when offer_type == "Other"
    client_id: Optional[str] = None
    description: Optional[str] = ""
    site_location: Optional[str] = ""
    base_amount: float = 0.0  # pre-GST
    gst_percent: float = 18.0
    file_path: Optional[str] = ""  # path on user's PC, e.g. D:\Offers\2026\audit.pdf
    status: Optional[str] = "Pending"  # Pending / Accepted / Rejected
    offer_date: Optional[str] = None  # ISO
    notes: Optional[str] = ""
    reference_no: Optional[str] = ""  # like STR/AUDIT/2026/023
    # Editable PDF content — all optional; defaults used when blank
    subject: Optional[str] = ""  # overrides auto-generated SUBJECT line
    scope_of_work: Optional[str] = ""  # multiline scope block (overrides description for PDF)
    payment_schedule: Optional[List[dict]] = None  # [{label, percent}] — if empty, 50/50 default
    terms_conditions: Optional[List[str]] = None  # list of T&C bullets — if empty, defaults
    bank_details: Optional[str] = ""  # overrides default bank line
    signature_name: Optional[str] = ""  # overrides "Mr. Rutvij Patel..."
    company_header: Optional[str] = ""  # overrides "CREATOR RCC CONSULTANT LLP"
    company_tagline: Optional[str] = ""
    company_address: Optional[str] = ""
    intro_paragraph: Optional[str] = ""  # overrides default intro


class Offer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    offer_code: str  # OFR-0001
    offer_type: str
    custom_type: str = ""
    effective_type: str = ""  # convenience: the type to display (custom_type if type == Other else offer_type)
    client_id: Optional[str] = None
    client_name: Optional[str] = ""
    client_phone: Optional[str] = ""
    client_email: Optional[str] = ""
    description: str = ""
    site_location: str = ""
    base_amount: float = 0.0
    gst_percent: float = 18.0
    gst_amount: float = 0.0
    total_amount: float = 0.0
    file_path: str = ""
    status: str = "Pending"
    offer_date: str = ""
    reference_no: str = ""
    notes: str = ""
    # Editable PDF content
    subject: str = ""
    scope_of_work: str = ""
    payment_schedule: List[dict] = Field(default_factory=list)
    terms_conditions: List[str] = Field(default_factory=list)
    bank_details: str = ""
    signature_name: str = ""
    company_header: str = ""
    company_tagline: str = ""
    company_address: str = ""
    intro_paragraph: str = ""
    linked_project_id: Optional[str] = None
    linked_project_code: Optional[str] = ""
    created_at: str


class PaymentIn(BaseModel):
    project_id: str
    amount: float
    payment_date: Optional[str] = None  # ISO string
    notes: Optional[str] = ""
    invoice_no: Optional[str] = ""
    taxable_amount: Optional[float] = None


class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    project_id: str
    project_code: str
    amount: float
    taxable_amount: Optional[float] = None
    payment_date: str
    notes: str = ""
    invoice_no: Optional[str] = ""
    last_edited_by_user_id: Optional[str] = None
    last_edited_by_username: Optional[str] = ""
    last_edited_at: Optional[str] = ""
    created_at: str



# ---------------------- SITE VISIT (Engineer) MODELS ----------------------

class SiteVisitTemplateIn(BaseModel):
    name: str
    description: Optional[str] = ""
    checklist: List[str] = []  # e.g. ["Size of members as per drawing", "Reinforcement - Dia, No of bars", ...]


class SiteVisitTemplate(SiteVisitTemplateIn):
    model_config = ConfigDict(extra="ignore")
    id: str
    created_at: str


class ChecklistItem(BaseModel):
    label: str
    compliance: str = "yes"  # yes | no | na
    remark: Optional[str] = ""


class SiteVisitPhoto(BaseModel):
    data_url: Optional[str] = ""  # base64 image (legacy / inline use)
    url: Optional[str] = ""  # /api/uploads/site-visits/<filename> (preferred)
    caption: Optional[str] = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    geo_accuracy: Optional[float] = None
    captured_at: Optional[str] = ""  # ISO timestamp from the engineer's device


class SiteVisitIn(BaseModel):
    template_id: Optional[str] = None
    template_name: Optional[str] = ""
    job_no: Optional[str] = ""
    project_id: Optional[str] = None
    inspection_title: str
    visit_date: Optional[str] = None  # ISO date
    customer: Optional[str] = ""
    plot_no: Optional[str] = ""           # kept for backward-compat with older visits
    site_location: Optional[str] = ""     # NEW — auto-filled from linked project
    drg_no: Optional[str] = ""
    revision: Optional[str] = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    geo_accuracy: Optional[float] = None
    checklist: List[ChecklistItem] = []
    observations: List[str] = []
    photos: List[SiteVisitPhoto] = []
    engineer_name: Optional[str] = ""
    engineer_signature: Optional[str] = ""  # base64 data URL of signature pad
    site_person_name: Optional[str] = ""
    site_person_phone: Optional[str] = ""
    site_person_signature: Optional[str] = ""
    status: str = "submitted"  # draft | submitted


class SiteVisit(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    visit_code: str  # SV-0001
    template_id: Optional[str] = None
    template_name: str = ""
    job_no: str = ""
    project_id: Optional[str] = None
    project_code: Optional[str] = ""
    project_name: Optional[str] = ""
    inspection_title: str = ""
    visit_date: Optional[str] = None
    customer: str = ""
    plot_no: str = ""
    site_location: str = ""
    drg_no: str = ""
    revision: str = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    geo_accuracy: Optional[float] = None
    checklist: List[ChecklistItem] = []
    observations: List[str] = []
    photos: List[SiteVisitPhoto] = []
    engineer_name: str = ""
    engineer_signature: str = ""
    site_person_name: str = ""
    site_person_phone: str = ""
    site_person_signature: str = ""
    status: str = "submitted"
    is_pinned: bool = False
    public_token: Optional[str] = ""
    created_by_user_id: Optional[str] = None
    created_by_username: Optional[str] = ""
    last_edited_by_user_id: Optional[str] = None
    last_edited_by_username: Optional[str] = ""
    last_edited_at: Optional[str] = ""
    created_at: str

class CompanyDetailsIn(BaseModel):
    name: str
    address: str
    gstin: str
    mobile: str
    pan: str
    email: str
    bank_name: str
    bank_account_name: str
    bank_ifsc: str
    bank_account_no: str
    bank_branch: str
    upi_id: str
    qr_code_url: Optional[str] = ""
    company_logo_url: Optional[str] = ""


class InvoiceItem(BaseModel):
    service_description: str
    qty: float = 1.0
    rate: float

class InvoiceIn(BaseModel):
    type: str # "proforma" or "tax"
    invoice_date: str # "YYYY-MM-DD"
    expiry_date: Optional[str] = ""
    hsn_code: str = "998332"
    client_id: str
    client_name: str
    client_address: Optional[str] = ""
    client_gstin: Optional[str] = ""
    client_mobile: Optional[str] = ""
    client_pan: Optional[str] = ""
    place_of_supply: Optional[str] = ""
    project_id: Optional[str] = ""
    
    # Legacy fields (kept for backward compatibility)
    service_description: Optional[str] = ""
    qty: Optional[float] = 1.0
    rate: Optional[float] = 0.0
    
    # New items array
    items: Optional[List[InvoiceItem]] = []
    
    gst_percent: float = 18.0
    tds_percent: float = 10.0
    tds_section: str = "194J"
    received_amount: float = 0.0


class Invoice(InvoiceIn):
    id: str
    invoice_no: str
    created_by_user_id: Optional[str] = None
    created_by_username: Optional[str] = ""
    created_at: str


class PaginatedInvoices(BaseModel):
    data: List[Invoice]
    total: int


class DocumentTypeIn(BaseModel):
    name: str
    prefix: str  # e.g. "QT", "STAB" — used inside STR/{prefix}/{YYYY}/{counter:03}
    description: Optional[str] = ""
    year_reset: bool = True


class DocumentType(DocumentTypeIn):
    model_config = ConfigDict(extra="ignore")
    id: str
    counter: int = 0
    last_year: int = 0
    created_at: str


class DocumentIn(BaseModel):
    doc_type_id: str
    doc_number: Optional[str] = ""  # auto-generated if blank
    document_date: Optional[str] = None  # ISO date string
    client_id: Optional[str] = None
    architect_id: Optional[str] = None
    plot_place: Optional[str] = ""
    phase: Optional[str] = ""
    number_field: Optional[str] = ""  # the "Number" field on the form (free-text)
    remark: Optional[str] = ""
    contact_person: Optional[str] = ""
    mobile: Optional[str] = ""
    other_comments: Optional[str] = ""
    update_date: Optional[str] = None  # ISO date string


class Document(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    doc_type_id: str
    doc_type_name: str
    doc_number: str
    document_date: Optional[str] = None
    client_id: Optional[str] = None
    client_name: Optional[str] = ""
    architect_id: Optional[str] = None
    architect_name: Optional[str] = ""
    plot_place: str = ""
    phase: str = ""
    number_field: str = ""
    remark: str = ""
    contact_person: str = ""
    mobile: str = ""
    other_comments: str = ""
    update_date: Optional[str] = None
    archived: bool = False
    status: str = "pending"  # pending | confirmed | on_hold | cancelled
    confirmed: bool = False
    linked_project_id: Optional[str] = None
    linked_project_code: Optional[str] = ""
    linked_project_name: Optional[str] = ""
    linked_audit_id: Optional[str] = None
    linked_audit_code: Optional[str] = ""
    linked_audit_offer: Optional[str] = ""
    confirmed_at: Optional[str] = None
    last_edited_by_user_id: Optional[str] = None
    last_edited_by_username: Optional[str] = ""
    last_edited_at: Optional[str] = ""
    created_at: str




# ---------------------- HELPERS ----------------------
import uuid


def _new_id() -> str:
    return str(uuid.uuid4())


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


async def _next_project_code() -> str:
    """Get next sequential project code like CC-0001."""
    doc = await db.counters.find_one_and_update(
        {"_id": "project_code"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    # If just created, seq will be 1
    seq = doc.get("seq", 1) if doc else 1
    return f"CC-{seq:04d}"


async def _next_offer_code() -> str:
    """Get next sequential offer code like OFR-0001."""
    doc = await db.counters.find_one_and_update(
        {"_id": "offer_code"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = doc.get("seq", 1) if doc else 1
    return f"OFR-{seq:04d}"


async def _enrich_project(p: dict, client_map: Optional[dict] = None, architect_map: Optional[dict] = None) -> dict:
    """Attach client and architect names, compute outstanding.
    If client_map/architect_map are provided, use them (batch mode, no DB queries).
    Otherwise fall back to per-doc DB fetch (single-project mode).
    """
    cid = p.get("client_id")
    if cid:
        if client_map is not None:
            c = client_map.get(cid)
        else:
            c = await db.clients.find_one({"id": cid}, {"_id": 0})
        p["client_name"] = (c.get("name") if c else "") or ""
        p["client_phone"] = (c.get("phone") if c else "") or ""
        p["client_email"] = (c.get("email") if c else "") or ""
    else:
        p["client_name"] = ""
        p["client_phone"] = ""
        p["client_email"] = ""
    aid = p.get("architect_id")
    if aid:
        if architect_map is not None:
            a = architect_map.get(aid)
        else:
            a = await db.architects.find_one({"id": aid}, {"_id": 0})
        p["architect_name"] = (a.get("name") if a else "") or ""
        p["architect_phone"] = (a.get("phone") if a else "") or ""
        p["architect_email"] = (a.get("email") if a else "") or ""
    else:
        p["architect_name"] = ""
        p["architect_phone"] = ""
        p["architect_email"] = ""
    p["quoted_amount"] = float(p.get("quoted_amount", 0) or 0)
    p["received_amount"] = float(p.get("received_amount", 0) or 0)
    p["outstanding_amount"] = round(p["quoted_amount"] - p["received_amount"], 2)
    p["archived"] = bool(p.get("archived", False))
    # auto-update status
    if p["outstanding_amount"] <= 0 and p["quoted_amount"] > 0:
        p["status"] = "Settled"
    elif p.get("status") not in ("Settled", "Outstanding"):
        p["status"] = "Outstanding"
    return p


async def _recalculate_project_received(project_id: str):
    payments = await db.payments.find({"project_id": project_id}, {"_id": 0}).to_list(10000)
    received_amount = 0.0
    for p in payments:
        tax_amt = p.get("taxable_amount")
        if tax_amt is not None:
            received_amount += float(tax_amt)
        else:
            received_amount += float(p.get("amount", 0))
            
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if project:
        quoted = float(project.get("quoted_amount", 0) or 0)
        outstanding = round(quoted - received_amount, 2)
        status = "Settled" if (outstanding <= 0 and quoted > 0) else "Outstanding"
        await db.projects.update_one(
            {"id": project_id},
            {"$set": {
                "received_amount": received_amount,
                "outstanding_amount": outstanding,
                "status": status
            }}
        )


async def _enrich_projects_batch(projects: List[dict]) -> List[dict]:
    """Efficiently enrich a list of projects by batch-loading clients and architects.
    Avoids N+1 DB queries.
    """
    if not projects:
        return projects
    client_ids = {p["client_id"] for p in projects if p.get("client_id")}
    architect_ids = {p["architect_id"] for p in projects if p.get("architect_id")}
    client_map = {}
    architect_map = {}
    if client_ids:
        clients = await db.clients.find({"id": {"$in": list(client_ids)}}, {"_id": 0}).to_list(len(client_ids))
        client_map = {c["id"]: c for c in clients}
    if architect_ids:
        architects = await db.architects.find({"id": {"$in": list(architect_ids)}}, {"_id": 0}).to_list(len(architect_ids))
        architect_map = {a["id"]: a for a in architects}
    for p in projects:
        await _enrich_project(p, client_map=client_map, architect_map=architect_map)
    return projects


# ---------------------- CLIENTS ----------------------
@api_router.get("/clients/verify-gstin/{gstin}")
async def verify_gstin(gstin: str):
    api_key = os.environ.get("GSTZEN_API_KEY")
    if not api_key:
        raise HTTPException(500, "GSTZEN_API_KEY is not configured on the server.")
    if len(gstin) != 15:
        raise HTTPException(400, "GSTIN must be exactly 15 characters.")
    
    import httpx
    url = "https://my.gstzen.in/api/gstin-validator/"
    headers = {
        "Content-Type": "application/json",
        "Token": api_key
    }
    payload = {"gstin": gstin}
    
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(url, headers=headers, json=payload, timeout=10.0)
            if resp.status_code != 200:
                raise HTTPException(resp.status_code, "Failed to reach GSTZen API")
            
            data = resp.json()
            api_message = data.get("message", "")
            
            if "Validator Subscription has exhausted" in api_message:
                raise HTTPException(400, "Subscription Limit Exhausted")
            if "Validator Subscription has expired" in api_message:
                raise HTTPException(400, "Subscription Period Expired")

            if data.get("status") == 0 or not data.get("valid"):
                raise HTTPException(400, "Invalid GSTIN.")
            
            company = data.get("company_details", {})
            
            # Format principal address
            pradr = company.get("pradr", {})
            adadr = company.get("adadr", [])
            
            addr_obj = {}
            if pradr and isinstance(pradr, dict):
                addr_obj = pradr.get("addr", pradr)
            elif adadr and isinstance(adadr, list) and len(adadr) > 0:
                addr_obj = adadr[0].get("addr", adadr[0]) if isinstance(adadr[0], dict) else {}

            principal_address = ""
            if isinstance(addr_obj, str):
                principal_address = addr_obj
            elif isinstance(addr_obj, dict):
                principal_address = ", ".join(filter(None, [
                    addr_obj.get("building_name") or addr_obj.get("bno") or addr_obj.get("bnm"),
                    addr_obj.get("building_number"),
                    addr_obj.get("floor_number") or addr_obj.get("flno"),
                    addr_obj.get("street") or addr_obj.get("st"),
                    addr_obj.get("loc"),
                    addr_obj.get("city") or addr_obj.get("dst"),
                    addr_obj.get("state_in_address") or addr_obj.get("stcd"),
                    addr_obj.get("pincode") or addr_obj.get("pncd")
                ]))
                if not principal_address:
                    principal_address = addr_obj.get("addr", "")
            
            return {
                "name": company.get("trade_name") or company.get("legal_name", ""),
                "pan": company.get("pan", ""),
                "gst_type": company.get("gst_type", ""),
                "place_of_supply": format_place_of_supply(company.get("state_info", {}).get("code", ""), gstin),
                "principal_address": principal_address
            }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"GSTZen verification error: {str(e)}")
        raise HTTPException(500, f"Error verifying GSTIN: {str(e)}")

@api_router.get("/clients", response_model=List[Client])
async def list_clients():
    items = await db.clients.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.get("/clients/paginated", response_model=PaginatedClients)
async def list_clients_paginated(page: int = 1, limit: int = 25, q: Optional[str] = None):
    skip = (page - 1) * limit
    query = {}
    if q:
        regex = {"$regex": q, "$options": "i"}
        query = {
            "$or": [
                {"name": regex},
                {"company": regex},
                {"phone": regex},
                {"email": regex},
                {"gstin": regex},
                {"pan": regex},
                {"place_of_supply": regex},
                {"address": regex}
            ]
        }
    
    total = await db.clients.count_documents(query)
    items = await db.clients.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {"data": items, "total": total}


def _norm(s: Optional[str]) -> str:
    """Normalise a string for duplicate comparison (trim, lowercase, collapse spaces)."""
    return " ".join((s or "").strip().lower().split())


def _digits(s: Optional[str]) -> str:
    """Keep only digits — used so '+91 98765-43210' matches '9876543210'."""
    return "".join(ch for ch in (s or "") if ch.isdigit())


async def _find_duplicate_contact(
    collection,
    name: Optional[str],
    phone: Optional[str],
    email: Optional[str],
    exclude_id: Optional[str] = None,
) -> Optional[dict]:
    """Return an existing row that matches by name, phone digits, or email."""
    name_n = _norm(name)
    phone_n = _digits(phone)
    email_n = _norm(email)
    if not (name_n or phone_n or email_n):
        return None
    # Fetch and compare in Python so we can do digits-only / case-insensitive match
    # without needing extra indexes. Collections are small (typically <1000 rows).
    rows = await collection.find({}, {"_id": 0}).to_list(5000)
    for r in rows:
        if exclude_id and r.get("id") == exclude_id:
            continue
        if name_n and _norm(r.get("name")) == name_n:
            return r
        if phone_n and len(phone_n) >= 6 and _digits(r.get("phone")) == phone_n:
            return r
        if email_n and _norm(r.get("email")) == email_n:
            return r
    return None


@api_router.post("/clients/bulk-import")
async def bulk_import_clients(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only Excel files (.xlsx, .xls) are supported.")
    
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Failed to read Excel file: {str(e)}")

    header_row_idx = None
    gstin_col = None
    name_col = None
    pos_col = None
    phone_col = None
    email_col = None
    target_ws = None
    
    for sheet in wb.worksheets:
        # Reset column trackers for each sheet
        gstin_col = None
        name_col = None
        pos_col = None
        phone_col = None
        email_col = None
        
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            temp_gstin_col = None
            temp_name_col = None
            temp_pos_col = None
            temp_phone_col = None
            temp_email_col = None
            
            for c_idx, cell in enumerate(row):
                val = str(cell).strip().lower() if cell else ""
                if not val: continue
                
                if "gstin" in val:
                    temp_gstin_col = c_idx
                elif "customer name" in val or "receiver name" in val:
                    temp_name_col = c_idx
                elif "place of supply" in val:
                    temp_pos_col = c_idx
                elif "phone" in val or "mobile" in val:
                    temp_phone_col = c_idx
                elif "email" in val:
                    temp_email_col = c_idx
            
            if temp_name_col is not None and (temp_gstin_col is not None or temp_pos_col is not None):
                header_row_idx = r_idx
                gstin_col = temp_gstin_col
                name_col = temp_name_col
                pos_col = temp_pos_col
                phone_col = temp_phone_col
                email_col = temp_email_col
                target_ws = sheet
                break
                
        if target_ws is not None:
            break
            
    if target_ws is None:
        raise HTTPException(400, "Could not find required columns ('Customer Name', 'GSTIN', 'Place of supply') in any sheet of the Excel file.")
        
    ws = target_ws
        
    imported = 0
    skipped = 0
    duplicates_in_file = []
    duplicates_in_db = []
    
    seen_in_file = set()
    
    existing_clients = await db.clients.find({}).to_list(None)
    db_keys = set()
    for c in existing_clients:
        key = (
            str(c.get("name") or "").strip().lower(),
            str(c.get("phone") or "").strip().lower(),
            str(c.get("email") or "").strip().lower(),
            str(c.get("gstin") or "").strip().lower(),
            str(c.get("place_of_supply") or "").strip().lower()
        )
        db_keys.add(key)
    
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = str(row[name_col]).strip() if name_col is not None and len(row) > name_col and row[name_col] else ""
        if not name or name.lower() == "none":
            continue
            
        gstin = str(row[gstin_col]).strip() if gstin_col is not None and len(row) > gstin_col and row[gstin_col] else ""
        if gstin.lower() == "none": gstin = ""
        
        pos = str(row[pos_col]).strip() if pos_col is not None and len(row) > pos_col and row[pos_col] else ""
        if pos.lower() == "none": pos = ""

        phone = str(row[phone_col]).strip() if phone_col is not None and len(row) > phone_col and row[phone_col] else ""
        if phone.lower() == "none": phone = ""
        
        email = str(row[email_col]).strip() if email_col is not None and len(row) > email_col and row[email_col] else ""
        if email.lower() == "none": email = ""
        
        row_key = (name.lower(), phone.lower(), email.lower(), gstin.lower(), pos.lower())
        
        if row_key in seen_in_file:
            skipped += 1
            duplicates_in_file.append({"name": name, "gstin": gstin})
            continue
            
        seen_in_file.add(row_key)
        
        if row_key in db_keys:
            skipped += 1
            duplicates_in_db.append({"name": name, "gstin": gstin})
            continue
            
        doc = {
            "name": name,
            "gstin": gstin,
            "place_of_supply": format_place_of_supply(pos, gstin),
            "pan": gstin[2:12] if len(gstin) >= 12 else "",
            "phone": phone,
            "email": email,
            "company": name,
            "address": "",
            "id": _new_id(),
            "created_at": _now()
        }
        _stamp_edit(doc)
        await db.clients.insert_one(doc)
        imported += 1
        
    return {
        "total_scanned": imported + skipped,
        "imported": imported,
        "skipped": skipped,
        "duplicates_in_file": duplicates_in_file,
        "duplicates_in_db": duplicates_in_db,
        "message": f"Scanned {imported + skipped}. Imported {imported}. Skipped {skipped} duplicates."
    }


@api_router.post("/clients", response_model=Client)
async def create_client(data: ClientIn):
    dup = await _find_duplicate_contact(db.clients, data.name, data.phone, data.email)
    if dup:
        reason = (
            f"name '{dup.get('name')}'" if _norm(dup.get("name")) == _norm(data.name)
            else f"phone {dup.get('phone')}" if _digits(dup.get("phone")) == _digits(data.phone)
            else f"email {dup.get('email')}"
        )
        raise HTTPException(409, f"A client with this {reason} already exists. Open '{dup.get('name')}' instead, or change the details.")
    doc = data.model_dump()
    doc["id"] = _new_id()
    doc["created_at"] = _now()
    _stamp_edit(doc)
    await db.clients.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api_router.put("/clients/{client_id}", response_model=Client)
async def update_client(client_id: str, data: ClientIn):
    dup = await _find_duplicate_contact(db.clients, data.name, data.phone, data.email, exclude_id=client_id)
    if dup:
        reason = (
            f"name '{dup.get('name')}'" if _norm(dup.get("name")) == _norm(data.name)
            else f"phone {dup.get('phone')}" if _digits(dup.get("phone")) == _digits(data.phone)
            else f"email {dup.get('email')}"
        )
        raise HTTPException(409, f"Another client with this {reason} already exists.")
    update = data.model_dump()
    update["place_of_supply"] = format_place_of_supply(update.get("place_of_supply"), update.get("gstin"))
    _stamp_edit(update)
    result = await db.clients.find_one_and_update(
        {"id": client_id},
        {"$set": update},
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Client not found")
    # Keep cached client_name in projects in sync
    await db.projects.update_many(
        {"client_id": client_id}, {"$set": {"client_name": data.name}}
    )
    return result


@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str):
    res = await db.clients.delete_one({"id": client_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Client not found")
    await db.projects.update_many(
        {"client_id": client_id},
        {"$set": {"client_id": None, "client_name": ""}},
    )
    return {"ok": True}


# ---------------------- ARCHITECTS ----------------------
@api_router.get("/architects", response_model=List[Architect])
async def list_architects():
    items = await db.architects.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.get("/architects/{architect_id}")
async def get_architect_detail(architect_id: str):
    architect = await db.architects.find_one({"id": architect_id}, {"_id": 0})
    if not architect:
        raise HTTPException(404, "Architect not found")
    projects = await db.projects.find({"architect_id": architect_id}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    await _enrich_projects_batch(projects)
    # Audits linked to any client that this architect worked with. Audits don't
    # store an architect_id directly, so we bridge via projects.client_id.
    client_ids = [p.get("client_id") for p in projects if p.get("client_id")]
    audits = []
    if client_ids:
        audits = await db.audits.find(
            {"client_id": {"$in": list(set(client_ids))}, "archived": {"$ne": True}}, {"_id": 0}
        ).sort("created_at", -1).to_list(2000)
        for a in audits:
            a["outstanding_amount"] = round((a.get("total_amount") or 0) - (a.get("received_amount") or 0), 2)
    documents = await db.documents.find(
        {"architect_id": architect_id, "archived": {"$ne": True}}, {"_id": 0}
    ).sort("created_at", -1).to_list(2000)
    for d in documents:
        # client_name already enriched on write; architect_name fill so card is self-contained
        d.setdefault("architect_name", architect.get("name", ""))
    total_quoted = sum(p.get("quoted_amount", 0) for p in projects) + sum(a.get("total_amount", 0) for a in audits)
    total_received = sum(p.get("received_amount", 0) for p in projects) + sum(a.get("received_amount", 0) for a in audits)
    total_outstanding = round(total_quoted - total_received, 2)
    outstanding_count = (
        sum(1 for p in projects if p.get("status") != "Settled")
        + sum(1 for a in audits if a.get("status") != "Settled")
    )
    settled_count = (
        sum(1 for p in projects if p.get("status") == "Settled")
        + sum(1 for a in audits if a.get("status") == "Settled")
    )
    return {
        "architect": architect,
        "projects": projects,
        "audits": audits,
        "documents": documents,
        "stats": {
            "total_projects": len(projects),
            "total_audits": len(audits),
            "total_documents": len(documents),
            "total_quoted": round(total_quoted, 2),
            "total_received": round(total_received, 2),
            "total_outstanding": total_outstanding,
            "outstanding_count": outstanding_count,
            "settled_count": settled_count,
        },
    }

@api_router.get("/clients/{client_id}/ledger/export")
async def export_client_ledger(client_id: str):
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(404, "Client not found")

    projects = await db.projects.find({"client_id": client_id}, {"_id": 0}).to_list(None)
    project_map = {p["id"]: p for p in projects}
    project_ids = list(project_map.keys())

    invoices = await db.invoices.find({"client_id": client_id, "type": "tax"}, {"_id": 0}).to_list(None)
    payments = await db.payments.find({"project_id": {"$in": project_ids}}, {"_id": 0}).to_list(None)

    transactions = []
    
    for inv in invoices:
        items = inv.get("items", [])
        base_value = sum(float(it.get("qty", 1.0)) * float(it.get("rate", 0.0)) for it in items)
        gst_percent = float(inv.get("gst_percent", 18))
        gst_amount = base_value * (gst_percent / 100)
        total_amount_with_gst = base_value + gst_amount
        tds_percent = float(inv.get("tds_percent", 0))
        tds_amount = base_value * (tds_percent / 100) if tds_percent > 0 else 0
        payable_amount = round(total_amount_with_gst - tds_amount, 2)
        
        proj_name = ""
        if inv.get("project_id") and inv["project_id"] in project_map:
            p = project_map[inv["project_id"]]
            proj_name = p.get("project_code") or p.get("name")
        
        transactions.append({
            "date": inv.get("invoice_date") or inv.get("created_at", "")[:10],
            "type": "Invoice",
            "ref_no": inv.get("invoice_no", ""),
            "project": proj_name,
            "particulars": f"Tax Invoice Generated {inv.get('invoice_no', '')}",
            "debit": payable_amount,
            "credit": 0.0
        })

    for pay in payments:
        proj_name = ""
        if pay.get("project_id") and pay["project_id"] in project_map:
            p = project_map[pay["project_id"]]
            proj_name = p.get("project_code") or p.get("name")
            
        transactions.append({
            "date": pay.get("payment_date") or pay.get("created_at", "")[:10],
            "type": "Payment",
            "ref_no": pay.get("invoice_no", ""),
            "project": proj_name,
            "particulars": pay.get("notes", "Payment Received"),
            "debit": 0.0,
            "credit": float(pay.get("amount", 0))
        })

    transactions.sort(key=lambda x: x["date"])

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Client Ledger"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="061A11")
    title_font = Font(bold=True, size=14)

    ws.append([f"Ledger Account: {client.get('name', '')}"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Contact: {client.get('phone', '')} | {client.get('email', '')}"])
    ws.append([f"GSTIN: {client.get('gstin', 'N/A')}"])
    ws.append([])

    headers = ["Date", "Type", "Reference No", "Project", "Particulars", "Debit (INR)", "Credit (INR)", "Balance (INR)"]
    ws.append(headers)
    
    header_row = 5
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    running_balance = 0.0
    total_debit = 0.0
    total_credit = 0.0

    for t in transactions:
        running_balance += t["debit"]
        running_balance -= t["credit"]
        total_debit += t["debit"]
        total_credit += t["credit"]

        ws.append([
            t["date"],
            t["type"],
            t["ref_no"],
            t["project"],
            t["particulars"],
            round(t["debit"], 2) if t["debit"] else "",
            round(t["credit"], 2) if t["credit"] else "",
            round(running_balance, 2)
        ])

    ws.append([])
    ws.append(["", "", "", "", "TOTALS:", round(total_debit, 2), round(total_credit, 2), round(running_balance, 2)])
    summary_row = ws.max_row
    for col in range(5, 9):
        ws.cell(row=summary_row, column=col).font = Font(bold=True)

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    import re
    safe_name = re.sub(r'[^A-Za-z0-9_]', '_', client.get('name', 'Client'))
    filename = f"Ledger_{safe_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@api_router.get("/clients/{client_id}")
async def get_client_detail(client_id: str):
    client_doc = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client_doc:
        raise HTTPException(404, "Client not found")
    projects = await db.projects.find({"client_id": client_id}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    await _enrich_projects_batch(projects)
    # Audits linked to this client (native client_id link on the audit)
    audits = await db.audits.find(
        {"client_id": client_id, "archived": {"$ne": True}}, {"_id": 0}
    ).sort("created_at", -1).to_list(2000)
    for a in audits:
        a["outstanding_amount"] = round((a.get("total_amount") or 0) - (a.get("received_amount") or 0), 2)
    documents = await db.documents.find(
        {"client_id": client_id, "archived": {"$ne": True}}, {"_id": 0}
    ).sort("created_at", -1).to_list(2000)
    for d in documents:
        d.setdefault("client_name", client_doc.get("name", ""))
    total_quoted = sum(p.get("quoted_amount", 0) for p in projects) + sum(a.get("total_amount", 0) for a in audits)
    total_received = sum(p.get("received_amount", 0) for p in projects) + sum(a.get("received_amount", 0) for a in audits)
    total_outstanding = round(total_quoted - total_received, 2)
    outstanding_count = (
        sum(1 for p in projects if p.get("status") != "Settled")
        + sum(1 for a in audits if a.get("status") != "Settled")
    )
    settled_count = (
        sum(1 for p in projects if p.get("status") == "Settled")
        + sum(1 for a in audits if a.get("status") == "Settled")
    )
    return {
        "client": client_doc,
        "projects": projects,
        "audits": audits,
        "documents": documents,
        "stats": {
            "total_projects": len(projects),
            "total_audits": len(audits),
            "total_documents": len(documents),
            "total_quoted": round(total_quoted, 2),
            "total_received": round(total_received, 2),
            "total_outstanding": total_outstanding,
            "outstanding_count": outstanding_count,
            "settled_count": settled_count,
        },
    }


@api_router.post("/architects", response_model=Architect)
async def create_architect(data: ArchitectIn):
    dup = await _find_duplicate_contact(db.architects, data.name, data.phone, data.email)
    if dup:
        reason = (
            f"name '{dup.get('name')}'" if _norm(dup.get("name")) == _norm(data.name)
            else f"phone {dup.get('phone')}" if _digits(dup.get("phone")) == _digits(data.phone)
            else f"email {dup.get('email')}"
        )
        raise HTTPException(409, f"An architect with this {reason} already exists. Open '{dup.get('name')}' instead, or change the details.")
    doc = data.model_dump()
    doc["id"] = _new_id()
    doc["created_at"] = _now()
    _stamp_edit(doc)
    await db.architects.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api_router.put("/architects/{architect_id}", response_model=Architect)
async def update_architect(architect_id: str, data: ArchitectIn):
    dup = await _find_duplicate_contact(db.architects, data.name, data.phone, data.email, exclude_id=architect_id)
    if dup:
        reason = (
            f"name '{dup.get('name')}'" if _norm(dup.get("name")) == _norm(data.name)
            else f"phone {dup.get('phone')}" if _digits(dup.get("phone")) == _digits(data.phone)
            else f"email {dup.get('email')}"
        )
        raise HTTPException(409, f"Another architect with this {reason} already exists.")
    update = data.model_dump()
    _stamp_edit(update)
    result = await db.architects.find_one_and_update(
        {"id": architect_id},
        {"$set": update},
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Architect not found")
    await db.projects.update_many(
        {"architect_id": architect_id}, {"$set": {"architect_name": data.name}}
    )
    return result


@api_router.delete("/architects/{architect_id}")
async def delete_architect(architect_id: str):
    res = await db.architects.delete_one({"id": architect_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Architect not found")
    await db.projects.update_many(
        {"architect_id": architect_id},
        {"$set": {"architect_id": None, "architect_name": ""}},
    )
    return {"ok": True}


def _deny_engineer():
    """Raise 403 if the calling user is an engineer or draftsman. Used to lock financial /
    accounting endpoints (payments, audits, monthly revenue, etc.) so these roles
    can never read amounts."""
    user = get_current_user_safe()
    if user and user.get("role") in ("engineer", "draftsman"):
        raise HTTPException(status_code=403, detail="Engineers/Draftsmen are not allowed to view financial data")


def _require_admin():
    """Raise 403 unless the calling user is an admin. Used to gate destructive
    or counter-resetting endpoints (e.g. editing the Audit Offer numbering series)."""
    user = get_current_user_safe()
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")


# ---------------------- PROJECTS ----------------------
def _strip_financials_for_engineer(items):
    """If the caller is an engineer or draftsman, blank out money fields on project payload(s).
    These roles should never see quoted/received/outstanding amounts or payments.
    Accepts a single dict or a list and mutates in place."""
    user = get_current_user_safe()
    if not user or user.get("role") not in ("engineer", "draftsman"):
        return items
    blanked = ["quoted_amount", "received_amount", "outstanding_amount"]
    targets = items if isinstance(items, list) else [items]
    for t in targets:
        if not isinstance(t, dict):
            continue
        for k in blanked:
            if k in t:
                t[k] = None
        # Status (Outstanding / Settled) is derived from money, so hide it too.
        if "status" in t:
            t["status"] = ""
    return items


@api_router.get("/projects", response_model=List[Project])
async def list_projects(search: Optional[str] = None, include_archived: bool = False, archived_only: bool = False):
    query = {}
    if archived_only:
        query["archived"] = True
    elif not include_archived:
        query["archived"] = {"$ne": True}
    # NOTE: Engineers can browse all projects (read-only). This makes the
    # "Linked Project" picker on the New Site Visit form work even when no
    # explicit per-engineer assignment exists yet. RBAC is still enforced on
    # write endpoints (admins manage projects; engineers only create visits).
    if search:
        s = search.strip()
        client_ids = [c["id"] for c in await db.clients.find({"name": {"$regex": s, "$options": "i"}}, {"id": 1}).to_list(None)]
        arch_ids = [a["id"] for a in await db.architects.find({"name": {"$regex": s, "$options": "i"}}, {"id": 1}).to_list(None)]
        
        or_conds = [
            {"project_code": {"$regex": s, "$options": "i"}},
            {"job_no": {"$regex": s, "$options": "i"}},
            {"name": {"$regex": s, "$options": "i"}},
            {"site_location": {"$regex": s, "$options": "i"}},
        ]
        if client_ids:
            or_conds.append({"client_id": {"$in": client_ids}})
        if arch_ids:
            or_conds.append({"architect_id": {"$in": arch_ids}})
            
        query["$or"] = or_conds
    items = await db.projects.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    await _enrich_projects_batch(items)
    return _strip_financials_for_engineer(items)


@api_router.get("/projects/paginated", response_model=PaginatedProjects)
async def list_projects_paginated(
    page: int = 1, 
    limit: int = 25, 
    q: Optional[str] = None, 
    include_archived: bool = False, 
    archived_only: bool = False
):
    skip = (page - 1) * limit
    query = {}
    if archived_only:
        query["archived"] = True
    elif not include_archived:
        query["archived"] = {"$ne": True}
        
    if q:
        s = q.strip()
        client_ids = [c["id"] for c in await db.clients.find({"name": {"$regex": s, "$options": "i"}}, {"id": 1}).to_list(None)]
        arch_ids = [a["id"] for a in await db.architects.find({"name": {"$regex": s, "$options": "i"}}, {"id": 1}).to_list(None)]
        
        or_conds = [
            {"project_code": {"$regex": s, "$options": "i"}},
            {"job_no": {"$regex": s, "$options": "i"}},
            {"name": {"$regex": s, "$options": "i"}},
            {"site_location": {"$regex": s, "$options": "i"}},
        ]
        if client_ids:
            or_conds.append({"client_id": {"$in": client_ids}})
        if arch_ids:
            or_conds.append({"architect_id": {"$in": arch_ids}})
            
        query["$or"] = or_conds
        
    total = await db.projects.count_documents(query)
    items = await db.projects.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    await _enrich_projects_batch(items)
    items = _strip_financials_for_engineer(items)
    
    return {"data": items, "total": total}

@api_router.get("/projects/{project_id}", response_model=Project)
async def get_project(project_id: str):
    p = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Project not found")
    await _enrich_project(p)
    return _strip_financials_for_engineer(p)


@api_router.post("/projects", response_model=Project)
async def create_project(data: ProjectIn):
    doc = data.model_dump()
    doc["id"] = _new_id()
    doc["project_code"] = await _next_project_code()
    doc["received_amount"] = 0.0
    doc["created_at"] = _now()
    _stamp_edit(doc)
    await _enrich_project(doc)
    await db.projects.insert_one(doc.copy())
    await _log_activity(
        doc["id"], doc["project_code"], "PROJECT CREATED",
        f"Name: {doc.get('name', '')} | Client: {doc.get('client_name', '-')} | Architect: {doc.get('architect_name', '-')} | Quoted: {doc.get('quoted_amount', 0)}",
    )
    doc.pop("_id", None)
    return doc


@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, data: ProjectIn):
    existing = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Project not found")
    old_name = existing.get("name", "")
    old_quoted = existing.get("quoted_amount", 0)
    update = data.model_dump()
    existing.update(update)
    _stamp_edit(existing)
    await _enrich_project(existing)
    await db.projects.update_one({"id": project_id}, {"$set": existing})
    await _log_activity(
        project_id, existing.get("project_code", ""), "PROJECT UPDATED",
        f"Name: '{old_name}' -> '{existing.get('name', '')}' | Quoted: {old_quoted} -> {existing.get('quoted_amount', 0)}",
    )
    existing.pop("_id", None)
    return existing


@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str):
    res = await db.projects.delete_one({"id": project_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Project not found")
    await db.payments.delete_many({"project_id": project_id})
    await db.quote_revisions.delete_many({"project_id": project_id})
    await db.activity_log.delete_many({"project_id": project_id})
    return {"ok": True}


@api_router.post("/projects/{project_id}/archive")
async def archive_project(project_id: str):
    res = await db.projects.update_one({"id": project_id}, {"$set": {"archived": True}})
    if res.matched_count == 0:
        raise HTTPException(404, "Project not found")
    proj = await db.projects.find_one({"id": project_id}, {"_id": 0, "project_code": 1})
    await _log_activity(project_id, (proj or {}).get("project_code", ""), "PROJECT ARCHIVED", "")
    return {"ok": True, "archived": True}


@api_router.post("/projects/{project_id}/unarchive")
async def unarchive_project(project_id: str):
    res = await db.projects.update_one({"id": project_id}, {"$set": {"archived": False}})
    if res.matched_count == 0:
        raise HTTPException(404, "Project not found")
    proj = await db.projects.find_one({"id": project_id}, {"_id": 0, "project_code": 1})
    await _log_activity(project_id, (proj or {}).get("project_code", ""), "PROJECT RESTORED", "")
    return {"ok": True, "archived": False}


# ---------------------- PAYMENTS ----------------------
@api_router.get("/payments", response_model=List[Payment])
async def list_payments(project_id: Optional[str] = None):
    _deny_engineer()
    q = {"project_id": project_id} if project_id else {}
    items = await db.payments.find(q, {"_id": 0}).sort("payment_date", -1).to_list(5000)
    for p in items:
        if "taxable_amount" not in p or p["taxable_amount"] is None:
            p["taxable_amount"] = p["amount"]
    return items


@api_router.post("/payments", response_model=Payment)
async def create_payment(data: PaymentIn):
    project = await db.projects.find_one({"id": data.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(404, "Project not found")
    if data.amount <= 0:
        raise HTTPException(400, "Amount must be > 0")
    tax_amt = float(data.taxable_amount) if data.taxable_amount is not None else float(data.amount)
    doc = {
        "id": _new_id(),
        "project_id": data.project_id,
        "project_code": project.get("project_code", ""),
        "amount": float(data.amount),
        "taxable_amount": tax_amt,
        "payment_date": data.payment_date or _now(),
        "notes": data.notes or "",
        "created_at": _now(),
    }
    _stamp_edit(doc)
    await db.payments.insert_one(doc.copy())
    
    # Recalculate project totals
    await _recalculate_project_received(data.project_id)
    
    await _log_activity(
        data.project_id, project.get("project_code", ""),
        "PAYMENT ADDED",
        f"Amount: ₹ {tax_amt:,.2f} | Note: {data.notes or '-'}",
    )
    doc.pop("_id", None)
    return doc


@api_router.delete("/payments/{payment_id}")
async def delete_payment(payment_id: str):
    pay = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not pay:
        raise HTTPException(404, "Payment not found")
    await db.payments.delete_one({"id": payment_id})
    # Recompute project totals
    project = await db.projects.find_one({"id": pay["project_id"]}, {"_id": 0})
    if project:
        await _recalculate_project_received(pay["project_id"])
        tax_amt = pay.get("taxable_amount") if pay.get("taxable_amount") is not None else pay.get("amount", 0)
        await _log_activity(
            pay["project_id"], project.get("project_code", ""),
            "PAYMENT DELETED",
            f"Amount: ₹ {float(tax_amt):,.2f} | Note: {pay.get('notes', '-')}",
        )
        
    if pay.get("invoice_no"):
        invoice = await db.invoices.find_one({"invoice_no": pay["invoice_no"]})
        if invoice:
            await db.invoices.delete_one({"id": invoice["id"]})
            import re
            invoice_type = invoice.get("type", "tax_invoice")
            match = re.search(r'(\d+)$', pay["invoice_no"])
            if match:
                seq = int(match.group(1))
                counter_id = "proforma" if invoice_type == "proforma" else "tax_invoice"
                counter = await db.counters.find_one({"_id": counter_id})
                if counter and counter.get("seq") == seq:
                    await db.counters.update_one({"_id": counter_id}, {"$inc": {"seq": -1}})

    return {"ok": True}


# ---------------------- QUOTE REVISIONS ----------------------
class QuoteRevisionIn(BaseModel):
    new_amount: float
    reason: Optional[str] = ""


@api_router.get("/projects/{project_id}/revisions")
async def list_revisions(project_id: str):
    _deny_engineer()
    items = await db.quote_revisions.find({"project_id": project_id}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.post("/projects/{project_id}/revise-quote")
async def revise_quote(project_id: str, data: QuoteRevisionIn):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(404, "Project not found")
    old_amount = float(project.get("quoted_amount", 0) or 0)
    new_amount = float(data.new_amount)
    if new_amount < 0:
        raise HTTPException(400, "Amount must be >= 0")
    rev = {
        "id": _new_id(),
        "project_id": project_id,
        "project_code": project.get("project_code", ""),
        "old_amount": old_amount,
        "new_amount": new_amount,
        "reason": data.reason or "",
        "created_at": _now(),
    }
    await db.quote_revisions.insert_one(rev.copy())
    # Update project quoted amount, re-enrich
    project["quoted_amount"] = new_amount
    await _enrich_project(project)
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {
            "quoted_amount": new_amount,
            "outstanding_amount": project["outstanding_amount"],
            "status": project["status"],
        }},
    )
    await _log_activity(
        project_id, project.get("project_code", ""),
        "QUOTE REVISED",
        f"Old: ₹ {old_amount:,.2f} -> New: ₹ {new_amount:,.2f} | Reason: {data.reason or '-'}",
    )
    rev.pop("_id", None)
    return rev


# ---------------------- ACTIVITY LOG ----------------------
def _current_user_stamp() -> dict:
    """Returns user_id + username from the current request, or empty placeholders for system."""
    try:
        u = auth_module.get_current_user_safe()
    except Exception:
        u = None
    if u and u.get("id") and u["id"] != "anonymous":
        return {
            "user_id": u["id"],
            "username": u.get("username", ""),
        }
    return {"user_id": None, "username": "system"}


def _stamp_edit(doc: dict) -> dict:
    """Add last_edited_by + last_edited_at to a document. Mutates and returns it."""
    s = _current_user_stamp()
    doc["last_edited_by_user_id"] = s["user_id"]
    doc["last_edited_by_username"] = s["username"]
    doc["last_edited_at"] = _now().isoformat() if hasattr(_now(), "isoformat") else _now()
    return doc


async def _log_activity(project_id: str, project_code: str, action: str, detail: str = ""):
    try:
        s = _current_user_stamp()
        await db.activity_log.insert_one({
            "id": _new_id(),
            "project_id": project_id,
            "project_code": project_code,
            "action": action,
            "detail": detail,
            "user_id": s["user_id"],
            "username": s["username"],
            "created_at": _now(),
        })
    except Exception as e:
        logger.error(f"activity log error: {e}")


@api_router.get("/projects/{project_id}/activity")
async def list_activity(project_id: str):
    items = await db.activity_log.find({"project_id": project_id}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


# ---------------------- PER-PROJECT EXCEL EXPORT ----------------------
@api_router.get("/projects/{project_id}/export")
async def export_project_excel(project_id: str):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(404, "Project not found")
    await _enrich_project(project)
    payments = await db.payments.find({"project_id": project_id}, {"_id": 0}).sort("payment_date", 1).to_list(5000)
    revisions = await db.quote_revisions.find({"project_id": project_id}, {"_id": 0}).sort("created_at", 1).to_list(5000)
    activity = await db.activity_log.find({"project_id": project_id}, {"_id": 0}).sort("created_at", 1).to_list(5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Info"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="061A11")

    info_rows = [
        ("Project ID", project.get("project_code", "")),
        ("Project Name", project.get("name", "")),
        ("Client", project.get("client_name", "")),
        ("Architect", project.get("architect_name", "")),
        ("Site Location", project.get("site_location", "")),
        ("Current Quoted (INR)", project.get("quoted_amount", 0)),
        ("Received (INR)", project.get("received_amount", 0)),
        ("Outstanding (INR)", project.get("outstanding_amount", 0)),
        ("Status", project.get("status", "")),
        ("Notes", project.get("notes", "")),
        ("Created", project.get("created_at", "")),
    ]
    for label, value in info_rows:
        ws.append([label, value])
    for r in range(1, len(info_rows) + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50

    ws2 = wb.create_sheet("Payments")
    ws2.append(["#", "Date", "Amount (INR)", "Notes"])
    for col in range(1, 5):
        c = ws2.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
    for i, p in enumerate(payments, 1):
        tax_amt = float(p.get("taxable_amount") if p.get("taxable_amount") is not None else p.get("amount", 0))
        ws2.append([i, p.get("payment_date", ""), tax_amt, p.get("notes", "")])

    ws3 = wb.create_sheet("Quote Revisions")
    ws3.append(["#", "Old Amount (INR)", "New Amount (INR)", "Reason", "Date"])
    for col in range(1, 6):
        c = ws3.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
    for i, r in enumerate(revisions, 1):
        ws3.append([i, float(r.get("old_amount", 0)), float(r.get("new_amount", 0)), r.get("reason", ""), r.get("created_at", "")])

    ws4 = wb.create_sheet("Activity")
    ws4.append(["#", "Action", "Detail", "Date"])
    for col in range(1, 5):
        c = ws4.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
    for i, a in enumerate(activity, 1):
        ws4.append([i, a.get("action", ""), a.get("detail", ""), a.get("created_at", "")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"project_{project.get('project_code', 'export')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------- PDF (Invoice / Receipt) ----------------------
BRAND_GREEN = colors.HexColor("#0A2E1F")
BRAND_ACCENT = colors.HexColor("#10B981")
BRAND_MUTED = colors.HexColor("#526B60")


def _format_inr(n: float, show_decimals: bool = True) -> str:
    """Format number in Indian numbering system: 1,23,45,678.00"""
    try:
        if not show_decimals:
            n = round(n)
        neg = n < 0
        n = abs(float(n))
        whole = int(n)
        frac = round(n - whole, 2)
        s = str(whole)
        if len(s) > 3:
            last3 = s[-3:]
            rest = s[:-3]
            # Group rest by 2s
            parts = []
            while len(rest) > 2:
                parts.insert(0, rest[-2:])
                rest = rest[:-2]
            if rest:
                parts.insert(0, rest)
            s = ",".join(parts) + "," + last3
        if not show_decimals:
            return ("-" if neg else "") + s
        result = f"{s}.{int(round(frac * 100)):02d}"
        return ("-" if neg else "") + result
    except Exception:
        return f"{n:.2f}" if show_decimals else f"{int(round(n))}"



def _draw_rupee_right(c, x, y, symbol_str, amount_str, font_name, font_size):
    c.setFont(font_name, font_size)
    w_amt = c.stringWidth(amount_str, font_name, font_size)
    c.setFont("Roboto", font_size)
    w_sym = c.stringWidth(symbol_str, "Roboto", font_size)
    start_x = x - w_amt - w_sym
    c.drawString(start_x, y, symbol_str)
    c.setFont(font_name, font_size)
    c.drawString(start_x + w_sym, y, amount_str)

def _draw_pdf_header(c: canvas.Canvas, title: str, sub_id: str):
    width, height = A4
    # Top green band
    c.setFillColor(BRAND_GREEN)
    c.rect(0, height - 28 * mm, width, 28 * mm, fill=1, stroke=0)
    # Brand
    c.setFillColor(colors.white)
    c.setFont("Roboto-Bold", 18)
    c.drawString(18 * mm, height - 14 * mm, "CREATOR CONSULTANT")
    c.setFillColor(BRAND_ACCENT)
    c.setFont("Roboto", 8.5)
    c.drawString(18 * mm, height - 20 * mm, "Architecture • Engineering • Project Consultancy")
    # Title on right
    c.setFillColor(colors.white)
    c.setFont("Roboto-Bold", 22)
    c.drawRightString(width - 18 * mm, height - 14 * mm, title)
    c.setFont("Roboto", 8.5)
    c.drawRightString(width - 18 * mm, height - 20 * mm, sub_id)
    # Reset
    c.setFillColor(colors.black)


def _draw_kv(c, x, y, key, value, key_w=40 * mm, bold_value=False):
    c.setFont("Roboto", 8.5)
    c.setFillColor(BRAND_MUTED)
    c.drawString(x, y, key.upper())
    c.setFillColor(colors.black)
    c.setFont("Roboto-Bold" if bold_value else "Roboto", 11)
    c.drawString(x + key_w, y, value or "—")


def _draw_footer(c: canvas.Canvas):
    width, _ = A4
    c.setFillColor(BRAND_MUTED)
    c.setFont("Roboto", 8)
    c.drawString(18 * mm, 12 * mm, "This is a computer generated document from Creator Consultant.")
    c.drawRightString(width - 18 * mm, 12 * mm, datetime.now().strftime("Generated %d %b %Y, %H:%M"))


async def _build_receipt_pdf(payment: dict, project: dict, client_doc: Optional[dict]) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    pay_date = payment.get("payment_date", "")
    try:
        pay_date_str = datetime.fromisoformat(pay_date.replace("Z", "+00:00")).strftime("%d %b %Y")
    except Exception:
        pay_date_str = pay_date[:10]

    _draw_pdf_header(c, "RECEIPT", f"Receipt ID: {payment['id'][:8].upper()}")

    y = height - 42 * mm
    # Meta
    _draw_kv(c, 18 * mm, y, "Project", f"{project.get('project_code', '')} — {project.get('name', '')}", bold_value=True)
    y -= 7 * mm
    _draw_kv(c, 18 * mm, y, "Payment Date", pay_date_str, bold_value=True)
    y -= 7 * mm
    if client_doc:
        _draw_kv(c, 18 * mm, y, "Received From", client_doc.get("name", ""), bold_value=True)
        y -= 7 * mm
        if client_doc.get("address"):
            _draw_kv(c, 18 * mm, y, "Address", client_doc.get("address", ""))
            y -= 7 * mm
        contact_bits = []
        if client_doc.get("phone"):
            contact_bits.append(client_doc["phone"])
        if client_doc.get("email"):
            contact_bits.append(client_doc["email"])
        if contact_bits:
            _draw_kv(c, 18 * mm, y, "Contact", " • ".join(contact_bits))
            y -= 7 * mm

    # Amount highlight box
    y -= 6 * mm
    box_h = 28 * mm
    c.setFillColor(BRAND_GREEN)
    c.roundRect(18 * mm, y - box_h, width - 36 * mm, box_h, 6, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Roboto", 10)
    c.drawString(24 * mm, y - 10 * mm, "AMOUNT RECEIVED")
    c.setFont("Roboto-Bold", 26)
    c.drawString(24 * mm, y - 20 * mm, f"₹ {_format_inr(payment.get('amount', 0))}")
    c.setFillColor(BRAND_ACCENT)
    c.setFont("Roboto", 8.5)
    c.drawRightString(width - 24 * mm, y - 20 * mm, "Indian Rupees")

    y = y - box_h - 10 * mm
    c.setFillColor(colors.black)

    # Summary
    _draw_kv(c, 18 * mm, y, "Project Quoted", f"₹ {_format_inr(project.get('quoted_amount', 0))}")
    y -= 6 * mm
    _draw_kv(c, 18 * mm, y, "Total Received (incl. this)", f"₹ {_format_inr(project.get('received_amount', 0))}")
    y -= 6 * mm
    _draw_kv(c, 18 * mm, y, "Outstanding Balance", f"₹ {_format_inr(project.get('outstanding_amount', 0))}", bold_value=True)
    y -= 10 * mm

    if payment.get("notes"):
        c.setFillColor(BRAND_MUTED)
        c.setFont("Roboto", 8.5)
        c.drawString(18 * mm, y, "NOTES")
        c.setFillColor(colors.black)
        c.setFont("Roboto", 10)
        c.drawString(18 * mm, y - 5 * mm, payment["notes"][:110])
        y -= 12 * mm

    # Signature
    y = max(y, 40 * mm)
    c.setStrokeColor(BRAND_MUTED)
    c.line(width - 70 * mm, 32 * mm, width - 20 * mm, 32 * mm)
    c.setFillColor(BRAND_MUTED)
    c.setFont("Roboto", 8.5)
    c.drawString(width - 70 * mm, 28 * mm, "Authorised Signatory")
    c.setFont("Roboto-Bold", 10)
    c.setFillColor(BRAND_GREEN)
    c.drawString(width - 70 * mm, 35 * mm, "For Creator Consultant")

    _draw_footer(c)
    c.showPage()
    c.save()
    buf.seek(0)
    return buf.read()


async def _build_invoice_pdf(project: dict, client_doc: Optional[dict], architect_doc: Optional[dict]) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    _draw_pdf_header(c, "INVOICE", f"Project: {project.get('project_code', '')}")

    y = height - 42 * mm
    # Bill to
    c.setFillColor(BRAND_MUTED)
    c.setFont("Roboto", 8.5)
    c.drawString(18 * mm, y, "BILL TO")
    c.setFillColor(colors.black)
    c.setFont("Roboto-Bold", 13)
    y -= 6 * mm
    c.drawString(18 * mm, y, (client_doc or {}).get("name", project.get("client_name", "")) or "—")
    c.setFont("Roboto", 10)
    if client_doc:
        if client_doc.get("company"):
            y -= 5 * mm
            c.drawString(18 * mm, y, client_doc["company"])
        if client_doc.get("address"):
            y -= 5 * mm
            c.drawString(18 * mm, y, client_doc["address"][:80])
        line = " • ".join([x for x in [client_doc.get("phone"), client_doc.get("email")] if x])
        if line:
            y -= 5 * mm
            c.setFillColor(BRAND_MUTED)
            c.drawString(18 * mm, y, line)
            c.setFillColor(colors.black)

    # Meta on right
    ry = height - 42 * mm
    c.setFillColor(BRAND_MUTED)
    c.setFont("Roboto", 8.5)
    c.drawRightString(width - 18 * mm, ry, "INVOICE DATE")
    c.setFillColor(colors.black)
    c.setFont("Roboto-Bold", 12)
    c.drawRightString(width - 18 * mm, ry - 5 * mm, datetime.now().strftime("%d %b %Y"))
    c.setFillColor(BRAND_MUTED)
    c.setFont("Roboto", 8.5)
    c.drawRightString(width - 18 * mm, ry - 12 * mm, "PROJECT STATUS")
    c.setFillColor(BRAND_ACCENT if project.get("status") == "Settled" else colors.HexColor("#DC2626"))
    c.setFont("Roboto-Bold", 12)
    c.drawRightString(width - 18 * mm, ry - 17 * mm, project.get("status", "Outstanding").upper())

    # Line items table
    y -= 18 * mm
    c.setFillColor(BRAND_GREEN)
    c.rect(18 * mm, y - 8 * mm, width - 36 * mm, 8 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Roboto-Bold", 9)
    c.drawString(22 * mm, y - 5.5 * mm, "DESCRIPTION")
    c.drawRightString(width - 22 * mm, y - 5.5 * mm, "AMOUNT (INR)")

    y -= 8 * mm
    c.setFillColor(colors.black)

    # Project line
    y -= 8 * mm
    c.setFont("Roboto-Bold", 12)
    c.drawString(22 * mm, y, project.get("name", ""))
    c.setFont("Roboto", 8.5)
    c.setFillColor(BRAND_MUTED)
    y -= 5 * mm
    desc = project.get("site_location") or ""
    if desc:
        c.drawString(22 * mm, y, f"Location: {desc[:75]}")
        y -= 5 * mm
    if architect_doc and architect_doc.get("name"):
        c.drawString(22 * mm, y, f"Architect: {architect_doc['name']}")
        y -= 5 * mm

    c.setFillColor(colors.black)
    c.setFont("Roboto-Bold", 12)
    c.drawRightString(width - 22 * mm, y + 10 * mm, f"₹ {_format_inr(project.get('quoted_amount', 0))}")

    # Totals box
    y -= 8 * mm
    c.setStrokeColor(BRAND_MUTED)
    c.line(18 * mm, y, width - 18 * mm, y)

    y -= 8 * mm
    c.setFont("Roboto", 10)
    c.setFillColor(BRAND_MUTED)
    c.drawRightString(width - 60 * mm, y, "Quoted Amount")
    c.setFillColor(colors.black)
    c.setFont("Roboto", 11)
    c.drawRightString(width - 22 * mm, y, f"₹ {_format_inr(project.get('quoted_amount', 0))}")

    y -= 7 * mm
    c.setFillColor(BRAND_MUTED)
    c.setFont("Roboto", 10)
    c.drawRightString(width - 60 * mm, y, "Received")
    c.setFillColor(BRAND_ACCENT)
    c.setFont("Roboto", 11)
    c.drawRightString(width - 22 * mm, y, f"₹ {_format_inr(project.get('received_amount', 0))}")

    y -= 10 * mm
    # Outstanding highlight
    c.setFillColor(BRAND_GREEN)
    c.roundRect(width - 90 * mm, y - 4 * mm, 72 * mm, 14 * mm, 4, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Roboto-Bold", 10)
    c.drawString(width - 86 * mm, y + 3 * mm, "AMOUNT DUE")
    c.setFont("Roboto-Bold", 14)
    c.drawRightString(width - 22 * mm, y + 3 * mm, f"₹ {_format_inr(project.get('outstanding_amount', 0))}")

    # Footer notes
    c.setFillColor(BRAND_MUTED)
    c.setFont("Roboto", 8.5)
    c.drawString(18 * mm, 40 * mm, "Payment Terms: Due on receipt. Kindly pay via bank transfer, cheque or UPI.")
    c.drawString(18 * mm, 35 * mm, "Thank you for your business.")

    _draw_footer(c)
    c.showPage()
    c.save()
    buf.seek(0)
    return buf.read()


@api_router.get("/payments/{payment_id}/receipt")
async def payment_receipt_pdf(payment_id: str):
    _deny_engineer()
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(404, "Payment not found")
    project = await db.projects.find_one({"id": payment["project_id"]}, {"_id": 0})
    if not project:
        raise HTTPException(404, "Project not found")
    await _enrich_project(project)
    client_doc = None
    if project.get("client_id"):
        client_doc = await db.clients.find_one({"id": project["client_id"]}, {"_id": 0})
    pdf_bytes = await _build_receipt_pdf(payment, project, client_doc)
    filename = f"receipt_{project['project_code']}_{payment_id[:8]}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/projects/{project_id}/invoice")
async def project_invoice_pdf(project_id: str):
    _deny_engineer()
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(404, "Project not found")
    await _enrich_project(project)
    client_doc = None
    architect_doc = None
    if project.get("client_id"):
        client_doc = await db.clients.find_one({"id": project["client_id"]}, {"_id": 0})
    if project.get("architect_id"):
        architect_doc = await db.architects.find_one({"id": project["architect_id"]}, {"_id": 0})
    pdf_bytes = await _build_invoice_pdf(project, client_doc, architect_doc)
    filename = f"invoice_{project['project_code']}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------- AUDITS ----------------------
class AuditIn(BaseModel):
    audit_code: Optional[str] = ""      # editable; auto-filled if blank
    audit_offer: str = ""                # e.g. "Structural Audit" (free text)
    report_id: Optional[str] = ""        # editable; auto-filled if blank
    client_id: Optional[str] = None
    # Free-text client contact fields. When set, override whatever would be
    # enriched from the linked Client record. Used for one-off audits where
    # the customer isn't a recurring client.
    client_name_override: Optional[str] = ""
    client_phone_override: Optional[str] = ""
    client_email_override: Optional[str] = ""
    total_amount: float = 0.0
    status: Optional[str] = "Outstanding"
    notes: Optional[str] = ""
    file_path: Optional[str] = ""        # path on user's PC, e.g. D:\Audits\2026\STR-AUDIT-006.pdf


class Audit(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    audit_code: str
    audit_offer: str = ""
    report_id: str = ""
    client_id: Optional[str] = None
    client_name: Optional[str] = ""
    client_phone: Optional[str] = ""
    client_email: Optional[str] = ""
    client_name_override: Optional[str] = ""
    client_phone_override: Optional[str] = ""
    client_email_override: Optional[str] = ""
    total_amount: float = 0.0
    received_amount: float = 0.0
    outstanding_amount: float = 0.0
    status: str = "Outstanding"
    notes: str = ""
    file_path: str = ""
    archived: bool = False
    last_edited_by_user_id: Optional[str] = None
    last_edited_by_username: Optional[str] = ""
    last_edited_at: Optional[str] = ""
    created_at: str


async def _next_audit_code() -> str:
    doc = await db.counters.find_one_and_update(
        {"_id": "audit_code"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = doc.get("seq", 1) if doc else 1
    return f"AUD-{seq:04d}"


async def _next_report_id() -> str:
    year = datetime.now(timezone.utc).year
    doc = await db.counters.find_one_and_update(
        {"_id": f"report_id_{year}"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = doc.get("seq", 1) if doc else 1
    return f"RPT-{year}-{seq:03d}"


async def _audit_offer_config() -> dict:
    """Read the configurable Audit Offer numbering setup from the `settings`
    collection. Falls back to the original hard-coded defaults so existing
    deployments don't break before the admin opens Settings."""
    doc = await db.settings.find_one({"_id": "audit_offer_series"}) or {}
    return {
        "prefix": (doc.get("prefix") or "STR/AUD-OFR").strip().strip("/"),
        "year_reset": doc.get("year_reset", True),
        "pad": int(doc.get("pad") or 3),
    }


def _format_audit_offer(prefix: str, year: int, seq: int, pad: int = 3, year_reset: bool = True) -> str:
    body = f"{prefix}/{year}/{str(seq).zfill(pad)}" if year_reset else f"{prefix}/{str(seq).zfill(pad)}"
    return body


async def _next_audit_offer_number() -> str:
    """Auto-generate the next Audit Offer Number using the admin-configured prefix."""
    cfg = await _audit_offer_config()
    year = datetime.now(timezone.utc).year
    counter_id = f"audit_offer_{year}" if cfg["year_reset"] else "audit_offer_all"
    doc = await db.counters.find_one_and_update(
        {"_id": counter_id},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = doc.get("seq", 1) if doc else 1
    return _format_audit_offer(cfg["prefix"], year, seq, cfg["pad"], cfg["year_reset"])


async def _peek_next_audit_offer_number() -> str:
    """Return what the NEXT auto-generated audit_offer would be, without
    consuming the counter. Used for the form preview hint."""
    cfg = await _audit_offer_config()
    year = datetime.now(timezone.utc).year
    counter_id = f"audit_offer_{year}" if cfg["year_reset"] else "audit_offer_all"
    doc = await db.counters.find_one({"_id": counter_id})
    seq = (doc or {}).get("seq", 0) + 1
    return _format_audit_offer(cfg["prefix"], year, seq, cfg["pad"], cfg["year_reset"])




async def _enrich_audit(a: dict) -> dict:
    if a.get("client_id"):
        c = await db.clients.find_one({"id": a["client_id"]}, {"_id": 0})
        a["client_name"] = (c.get("name") if c else "") or ""
        a["client_phone"] = (c.get("phone") if c else "") or ""
        a["client_email"] = (c.get("email") if c else "") or ""
    else:
        a["client_name"] = a.get("client_name") or ""
        a["client_phone"] = a.get("client_phone") or ""
        a["client_email"] = a.get("client_email") or ""
    # Per-audit overrides win over the enriched client fields. Useful for
    # one-off audits where the customer isn't a recurring client, or where
    # the engineer wants to record a different point-of-contact for THIS audit.
    if a.get("client_name_override"):
        a["client_name"] = a["client_name_override"]
    if a.get("client_phone_override"):
        a["client_phone"] = a["client_phone_override"]
    if a.get("client_email_override"):
        a["client_email"] = a["client_email_override"]
    total = float(a.get("total_amount", 0) or 0)
    received = float(a.get("received_amount", 0) or 0)
    outstanding = max(0.0, total - received)
    a["outstanding_amount"] = outstanding
    if outstanding <= 0.009 and total > 0:
        a["status"] = "Settled"
    else:
        a["status"] = a.get("status") or "Outstanding"
        if a["status"] == "Settled":
            a["status"] = "Outstanding"
    return a


async def _log_audit_activity(audit_id: str, audit_code: str, action: str, detail: str = ""):
    try:
        s = _current_user_stamp()
        await db.activity_log.insert_one({
            "id": _new_id(),
            "audit_id": audit_id,
            "audit_code": audit_code,
            "project_id": None,
            "project_code": "",
            "action": action,
            "detail": detail,
            "user_id": s["user_id"],
            "username": s["username"],
            "created_at": _now(),
        })
    except Exception as e:
        logger.error(f"audit activity log error: {e}")


@api_router.get("/audits/next-offer-preview")
async def audits_next_offer_preview():
    """Returns the next Audit Offer Number that will be assigned (for form hint)."""
    _deny_engineer()
    return {"number": await _peek_next_audit_offer_number()}


@api_router.get("/audits/offer-series")
async def get_audit_offer_series():
    """Inspect the current Audit Offer numbering config for the Settings UI."""
    _deny_engineer()
    cfg = await _audit_offer_config()
    year = datetime.now(timezone.utc).year
    counter_id = f"audit_offer_{year}" if cfg["year_reset"] else "audit_offer_all"
    doc = await db.counters.find_one({"_id": counter_id})
    seq = (doc or {}).get("seq", 0)
    next_seq = seq + 1
    return {
        "year": year,
        "prefix": cfg["prefix"],
        "year_reset": cfg["year_reset"],
        "pad": cfg["pad"],
        "current_seq": seq,
        "next_seq": next_seq,
        "next_code": _format_audit_offer(cfg["prefix"], year, next_seq, cfg["pad"], cfg["year_reset"]),
    }


@api_router.put("/audits/offer-series")
async def set_audit_offer_series(payload: dict = Body(...)):
    """Update the Audit Offer numbering config (admin only).

    Body may include any of: `next_seq`, `prefix`, `year_reset`, `pad`.
    Refuses changes that would collide with an existing audit_offer."""
    _require_admin()
    cfg = await _audit_offer_config()

    # --- collect new values (fall back to current config) ---
    new_prefix = str(payload.get("prefix", cfg["prefix"])).strip().strip("/")
    if not new_prefix:
        raise HTTPException(400, "Prefix cannot be empty")
    new_year_reset = bool(payload.get("year_reset", cfg["year_reset"]))
    try:
        new_pad = int(payload.get("pad", cfg["pad"]))
    except (TypeError, ValueError):
        raise HTTPException(400, "pad must be a positive integer")
    if new_pad < 1 or new_pad > 6:
        raise HTTPException(400, "pad must be between 1 and 6")

    year = datetime.now(timezone.utc).year
    counter_id = f"audit_offer_{year}" if new_year_reset else "audit_offer_all"
    # Default next_seq = current counter + 1 (i.e. don't change it if not provided)
    current_doc = await db.counters.find_one({"_id": counter_id})
    current_seq = (current_doc or {}).get("seq", 0)
    if "next_seq" in payload:
        try:
            next_seq = int(payload["next_seq"])
        except (TypeError, ValueError):
            raise HTTPException(400, "next_seq must be a positive integer")
        if next_seq < 1:
            raise HTTPException(400, "next_seq must be at least 1")
    else:
        next_seq = current_seq + 1

    candidate = _format_audit_offer(new_prefix, year, next_seq, new_pad, new_year_reset)
    clash = await db.audits.find_one({"audit_offer": candidate}, {"_id": 0, "id": 1})
    if clash:
        raise HTTPException(400, f"An audit with offer number {candidate} already exists. Pick a different starting number.")

    # --- persist config + counter ---
    await db.settings.update_one(
        {"_id": "audit_offer_series"},
        {"$set": {"prefix": new_prefix, "year_reset": new_year_reset, "pad": new_pad}},
        upsert=True,
    )
    await db.counters.update_one(
        {"_id": counter_id},
        {"$set": {"seq": next_seq - 1}},
        upsert=True,
    )
    return {
        "ok": True,
        "year": year,
        "prefix": new_prefix,
        "year_reset": new_year_reset,
        "pad": new_pad,
        "next_seq": next_seq,
        "next_code": candidate,
    }



@api_router.get("/audits", response_model=List[Audit])
async def list_audits(archived: Optional[bool] = False, search: Optional[str] = None):
    _deny_engineer()
    query = {"archived": bool(archived)}
    if search:
        s = search.strip()
        query["$or"] = [
            {"audit_code": {"$regex": s, "$options": "i"}},
            {"audit_offer": {"$regex": s, "$options": "i"}},
            {"report_id": {"$regex": s, "$options": "i"}},
            {"client_name": {"$regex": s, "$options": "i"}},
            {"notes": {"$regex": s, "$options": "i"}},
        ]
    items = await db.audits.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    for a in items:
        await _enrich_audit(a)
    return items


@api_router.get("/audits/{audit_id}", response_model=Audit)
async def get_audit(audit_id: str):
    _deny_engineer()
    a = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not a:
        raise HTTPException(404, "Audit not found")
    await _enrich_audit(a)
    return a


@api_router.post("/audits", response_model=Audit)
async def create_audit(data: AuditIn):
    doc = data.model_dump()
    doc["id"] = _new_id()
    doc["audit_code"] = (doc.get("audit_code") or "").strip() or await _next_audit_code()
    doc["report_id"] = (doc.get("report_id") or "").strip() or await _next_report_id()
    # Audit Offer Number — auto-generate STR/AUD-OFR/YYYY/NNN if the engineer
    # didn't override it on the form.
    doc["audit_offer"] = (doc.get("audit_offer") or "").strip() or await _next_audit_offer_number()
    doc["received_amount"] = 0.0
    doc["archived"] = False
    doc["created_at"] = _now()
    _stamp_edit(doc)
    await _enrich_audit(doc)
    await db.audits.insert_one(doc.copy())
    await _log_audit_activity(
        doc["id"], doc["audit_code"], "AUDIT CREATED",
        f"Offer: {doc.get('audit_offer', '-')} | Client: {doc.get('client_name', '-')} | Total: {doc.get('total_amount', 0)}",
    )
    doc.pop("_id", None)
    return doc


@api_router.put("/audits/{audit_id}", response_model=Audit)
async def update_audit(audit_id: str, data: AuditIn):
    existing = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Audit not found")
    old_total = existing.get("total_amount", 0)
    update = data.model_dump()
    # Keep the existing code/report_id/audit_offer if new value is blank
    if not (update.get("audit_code") or "").strip():
        update["audit_code"] = existing.get("audit_code", "")
    if not (update.get("report_id") or "").strip():
        update["report_id"] = existing.get("report_id", "")
    if not (update.get("audit_offer") or "").strip():
        update["audit_offer"] = existing.get("audit_offer", "")
    existing.update(update)
    _stamp_edit(existing)
    await _enrich_audit(existing)
    await db.audits.update_one({"id": audit_id}, {"$set": existing})
    await _log_audit_activity(
        audit_id, existing.get("audit_code", ""), "AUDIT UPDATED",
        f"Total: {old_total} -> {existing.get('total_amount', 0)}",
    )
    existing.pop("_id", None)
    return existing


@api_router.delete("/audits/{audit_id}")
async def delete_audit(audit_id: str):
    res = await db.audits.delete_one({"id": audit_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Audit not found")
    await db.audit_payments.delete_many({"audit_id": audit_id})
    await db.audit_quote_revisions.delete_many({"audit_id": audit_id})
    await db.activity_log.delete_many({"audit_id": audit_id})
    return {"ok": True}


@api_router.post("/audits/{audit_id}/archive")
async def archive_audit(audit_id: str):
    res = await db.audits.update_one({"id": audit_id}, {"$set": {"archived": True}})
    if res.matched_count == 0:
        raise HTTPException(404, "Audit not found")
    a = await db.audits.find_one({"id": audit_id}, {"_id": 0, "audit_code": 1})
    await _log_audit_activity(audit_id, (a or {}).get("audit_code", ""), "AUDIT ARCHIVED", "")
    return {"ok": True, "archived": True}


@api_router.post("/audits/{audit_id}/unarchive")
async def unarchive_audit(audit_id: str):
    res = await db.audits.update_one({"id": audit_id}, {"$set": {"archived": False}})
    if res.matched_count == 0:
        raise HTTPException(404, "Audit not found")
    a = await db.audits.find_one({"id": audit_id}, {"_id": 0, "audit_code": 1})
    await _log_audit_activity(audit_id, (a or {}).get("audit_code", ""), "AUDIT RESTORED", "")
    return {"ok": True, "archived": False}


# ---------- Audit Payments (separate collection from project payments) ----------
class AuditPaymentIn(BaseModel):
    audit_id: str
    amount: float
    payment_date: Optional[str] = None
    notes: Optional[str] = ""


class AuditPayment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    audit_id: str
    audit_code: str
    amount: float
    payment_date: str
    notes: str = ""
    last_edited_by_user_id: Optional[str] = None
    last_edited_by_username: Optional[str] = ""
    last_edited_at: Optional[str] = ""
    created_at: str


@api_router.get("/audit-payments", response_model=List[AuditPayment])
async def list_audit_payments(audit_id: Optional[str] = None):
    q = {"audit_id": audit_id} if audit_id else {}
    items = await db.audit_payments.find(q, {"_id": 0}).sort("payment_date", -1).to_list(5000)
    return items


@api_router.post("/audit-payments", response_model=AuditPayment)
async def create_audit_payment(data: AuditPaymentIn):
    audit = await db.audits.find_one({"id": data.audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    if data.amount <= 0:
        raise HTTPException(400, "Amount must be > 0")
    doc = {
        "id": _new_id(),
        "audit_id": data.audit_id,
        "audit_code": audit.get("audit_code", ""),
        "amount": float(data.amount),
        "payment_date": data.payment_date or _now(),
        "notes": data.notes or "",
        "created_at": _now(),
    }
    _stamp_edit(doc)
    await db.audit_payments.insert_one(doc.copy())
    new_received = float(audit.get("received_amount", 0)) + float(data.amount)
    audit["received_amount"] = new_received
    await _enrich_audit(audit)
    await db.audits.update_one(
        {"id": data.audit_id},
        {"$set": {
            "received_amount": new_received,
            "outstanding_amount": audit["outstanding_amount"],
            "status": audit["status"],
        }},
    )
    await _log_audit_activity(
        data.audit_id, audit.get("audit_code", ""), "PAYMENT ADDED",
        f"Amount: ₹ {float(data.amount):,.2f} | Note: {data.notes or '-'}",
    )
    doc.pop("_id", None)
    return doc


@api_router.delete("/audit-payments/{payment_id}")
async def delete_audit_payment(payment_id: str):
    pay = await db.audit_payments.find_one({"id": payment_id}, {"_id": 0})
    if not pay:
        raise HTTPException(404, "Audit payment not found")
    await db.audit_payments.delete_one({"id": payment_id})
    audit = await db.audits.find_one({"id": pay["audit_id"]}, {"_id": 0})
    if audit:
        new_received = max(0.0, float(audit.get("received_amount", 0)) - float(pay["amount"]))
        audit["received_amount"] = new_received
        await _enrich_audit(audit)
        await db.audits.update_one(
            {"id": pay["audit_id"]},
            {"$set": {
                "received_amount": new_received,
                "outstanding_amount": audit["outstanding_amount"],
                "status": audit["status"],
            }},
        )
        await _log_audit_activity(
            pay["audit_id"], audit.get("audit_code", ""), "PAYMENT DELETED",
            f"Amount: ₹ {float(pay['amount']):,.2f} | Note: {pay.get('notes', '-')}",
        )
    return {"ok": True}


@api_router.get("/audits/{audit_id}/activity")
async def list_audit_activity(audit_id: str):
    items = await db.activity_log.find({"audit_id": audit_id}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


def _audit_to_project_shape(a: dict) -> dict:
    """Adapt audit dict to the shape expected by _build_invoice_pdf / _build_receipt_pdf."""
    return {
        "id": a.get("id", ""),
        "project_code": a.get("audit_code", ""),
        "name": a.get("audit_offer", "") or "Structural Audit",
        "client_id": a.get("client_id"),
        "client_name": a.get("client_name", ""),
        "client_phone": a.get("client_phone", ""),
        "client_email": a.get("client_email", ""),
        "architect_id": None,
        "architect_name": "",
        "architect_phone": "",
        "architect_email": "",
        "site_location": "",
        "quoted_amount": float(a.get("total_amount", 0) or 0),
        "received_amount": float(a.get("received_amount", 0) or 0),
        "outstanding_amount": float(a.get("outstanding_amount", 0) or 0),
        "status": a.get("status", "Outstanding"),
        "notes": a.get("notes", "") or f"Report ID: {a.get('report_id', '')}",
        "created_at": a.get("created_at", _now()),
    }


@api_router.get("/audits/{audit_id}/invoice")
async def audit_invoice_pdf(audit_id: str):
    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    await _enrich_audit(audit)
    client_doc = None
    if audit.get("client_id"):
        client_doc = await db.clients.find_one({"id": audit["client_id"]}, {"_id": 0})
    proj_shape = _audit_to_project_shape(audit)
    pdf_bytes = await _build_invoice_pdf(proj_shape, client_doc, None)
    filename = f"invoice_{audit['audit_code']}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/audit-payments/{payment_id}/receipt")
async def audit_payment_receipt_pdf(payment_id: str):
    payment = await db.audit_payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(404, "Audit payment not found")
    audit = await db.audits.find_one({"id": payment["audit_id"]}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    await _enrich_audit(audit)
    client_doc = None
    if audit.get("client_id"):
        client_doc = await db.clients.find_one({"id": audit["client_id"]}, {"_id": 0})
    proj_shape = _audit_to_project_shape(audit)
    # The receipt builder reads payment["project_code"]; map audit_code -> project_code
    payment_shape = {**payment, "project_id": payment["audit_id"], "project_code": payment["audit_code"]}
    pdf_bytes = await _build_receipt_pdf(payment_shape, proj_shape, client_doc)
    filename = f"receipt_{audit['audit_code']}_{payment_id[:8]}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---- Audit quote revisions ----
class AuditQuoteRevisionIn(BaseModel):
    new_amount: float
    reason: Optional[str] = ""


@api_router.get("/audits/{audit_id}/revisions")
async def list_audit_revisions(audit_id: str):
    items = await db.audit_quote_revisions.find({"audit_id": audit_id}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.post("/audits/{audit_id}/revise-quote")
async def revise_audit_quote(audit_id: str, data: AuditQuoteRevisionIn):
    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    old_amount = float(audit.get("total_amount", 0) or 0)
    new_amount = float(data.new_amount)
    if new_amount < 0:
        raise HTTPException(400, "Amount must be >= 0")
    rev = {
        "id": _new_id(),
        "audit_id": audit_id,
        "audit_code": audit.get("audit_code", ""),
        "old_amount": old_amount,
        "new_amount": new_amount,
        "reason": data.reason or "",
        "created_at": _now(),
    }
    await db.audit_quote_revisions.insert_one(rev.copy())
    audit["total_amount"] = new_amount
    await _enrich_audit(audit)
    await db.audits.update_one(
        {"id": audit_id},
        {"$set": {
            "total_amount": new_amount,
            "outstanding_amount": audit["outstanding_amount"],
            "status": audit["status"],
        }},
    )
    await _log_audit_activity(
        audit_id, audit.get("audit_code", ""),
        "QUOTE REVISED",
        f"Old: ₹ {old_amount:,.2f} -> New: ₹ {new_amount:,.2f} | Reason: {data.reason or '-'}",
    )
    rev.pop("_id", None)
    return rev


# ---- Per-audit Excel export ----
@api_router.get("/audits/{audit_id}/export")
async def export_audit_excel(audit_id: str):
    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    await _enrich_audit(audit)
    payments = await db.audit_payments.find({"audit_id": audit_id}, {"_id": 0}).sort("payment_date", 1).to_list(5000)
    revisions = await db.audit_quote_revisions.find({"audit_id": audit_id}, {"_id": 0}).sort("created_at", 1).to_list(5000)
    activity = await db.activity_log.find({"audit_id": audit_id}, {"_id": 0}).sort("created_at", 1).to_list(5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Info"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="061A11")

    info_rows = [
        ("Audit ID", audit.get("audit_code", "")),
        ("Audit Offer", audit.get("audit_offer", "")),
        ("Report ID", audit.get("report_id", "")),
        ("Client", audit.get("client_name", "")),
        ("Current Total (INR)", audit.get("total_amount", 0)),
        ("Received (INR)", audit.get("received_amount", 0)),
        ("Outstanding (INR)", audit.get("outstanding_amount", 0)),
        ("Status", audit.get("status", "")),
        ("Notes", audit.get("notes", "")),
        ("Created", audit.get("created_at", "")),
    ]
    for label, value in info_rows:
        ws.append([label, value])
    for r in range(1, len(info_rows) + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50

    ws2 = wb.create_sheet("Payments")
    ws2.append(["#", "Date", "Amount (INR)", "Notes"])
    for col in range(1, 5):
        c = ws2.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
    for i, p in enumerate(payments, 1):
        ws2.append([i, p.get("payment_date", ""), float(p.get("amount", 0)), p.get("notes", "")])

    ws3 = wb.create_sheet("Quote Revisions")
    ws3.append(["#", "Old Amount (INR)", "New Amount (INR)", "Reason", "Date"])
    for col in range(1, 6):
        c = ws3.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
    for i, r in enumerate(revisions, 1):
        ws3.append([i, float(r.get("old_amount", 0)), float(r.get("new_amount", 0)), r.get("reason", ""), r.get("created_at", "")])

    ws4 = wb.create_sheet("Activity")
    ws4.append(["#", "Action", "Detail", "Date"])
    for col in range(1, 5):
        c = ws4.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
    for i, a in enumerate(activity, 1):
        ws4.append([i, a.get("action", ""), a.get("detail", ""), a.get("created_at", "")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"audit_{audit.get('audit_code', 'export')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------- OFFERS ----------------------
def _effective_offer_type(offer_type: str, custom_type: str) -> str:
    if offer_type and offer_type.lower() == "other":
        return (custom_type or "Other").strip()
    return (offer_type or "").strip()


async def _enrich_offer(o: dict) -> dict:
    if o.get("client_id"):
        c = await db.clients.find_one({"id": o["client_id"]}, {"_id": 0})
        o["client_name"] = (c.get("name") if c else "") or ""
        o["client_phone"] = (c.get("phone") if c else "") or ""
        o["client_email"] = (c.get("email") if c else "") or ""
    else:
        o["client_name"] = ""
        o["client_phone"] = ""
        o["client_email"] = ""
    o["base_amount"] = float(o.get("base_amount", 0) or 0)
    o["gst_percent"] = float(o.get("gst_percent", 18) or 0)
    o["gst_amount"] = round(o["base_amount"] * o["gst_percent"] / 100.0, 2)
    o["total_amount"] = round(o["base_amount"] + o["gst_amount"], 2)
    o["effective_type"] = _effective_offer_type(o.get("offer_type", ""), o.get("custom_type", ""))
    o["custom_type"] = o.get("custom_type", "") or ""
    o["file_path"] = o.get("file_path", "") or ""
    o["description"] = o.get("description", "") or ""
    o["site_location"] = o.get("site_location", "") or ""
    o["status"] = o.get("status") or "Pending"
    o["offer_date"] = o.get("offer_date") or ""
    o["reference_no"] = o.get("reference_no") or ""
    o["notes"] = o.get("notes") or ""
    # Editable PDF content fields — normalize types
    o["subject"] = o.get("subject", "") or ""
    o["scope_of_work"] = o.get("scope_of_work", "") or ""
    ps = o.get("payment_schedule")
    o["payment_schedule"] = ps if isinstance(ps, list) else []
    tcs = o.get("terms_conditions")
    o["terms_conditions"] = tcs if isinstance(tcs, list) else []
    o["bank_details"] = o.get("bank_details", "") or ""
    o["signature_name"] = o.get("signature_name", "") or ""
    o["company_header"] = o.get("company_header", "") or ""
    o["company_tagline"] = o.get("company_tagline", "") or ""
    o["company_address"] = o.get("company_address", "") or ""
    o["intro_paragraph"] = o.get("intro_paragraph", "") or ""
    return o


@api_router.get("/offers", response_model=List[Offer])
async def list_offers(status: Optional[str] = None, search: Optional[str] = None):
    query = {}
    if status:
        query["status"] = status
    if search:
        s = search.strip()
        query["$or"] = [
            {"offer_code": {"$regex": s, "$options": "i"}},
            {"description": {"$regex": s, "$options": "i"}},
            {"client_name": {"$regex": s, "$options": "i"}},
            {"site_location": {"$regex": s, "$options": "i"}},
            {"reference_no": {"$regex": s, "$options": "i"}},
        ]
    items = await db.offers.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    for o in items:
        await _enrich_offer(o)
    return items


@api_router.get("/offers/{offer_id}", response_model=Offer)
async def get_offer(offer_id: str):
    o = await db.offers.find_one({"id": offer_id}, {"_id": 0})
    if not o:
        raise HTTPException(404, "Offer not found")
    await _enrich_offer(o)
    return o


@api_router.post("/offers", response_model=Offer)
async def create_offer(data: OfferIn):
    doc = data.model_dump()
    doc["id"] = _new_id()
    doc["offer_code"] = await _next_offer_code()
    doc["created_at"] = _now()
    doc["linked_project_id"] = None
    doc["linked_project_code"] = ""
    await _enrich_offer(doc)
    await db.offers.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api_router.put("/offers/{offer_id}", response_model=Offer)
async def update_offer(offer_id: str, data: OfferIn):
    existing = await db.offers.find_one({"id": offer_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Offer not found")
    update = data.model_dump()
    existing.update(update)
    await _enrich_offer(existing)
    await db.offers.update_one({"id": offer_id}, {"$set": existing})
    # If this offer is already linked to a project, sync offer metadata on the project
    if existing.get("linked_project_id"):
        await db.projects.update_one(
            {"id": existing["linked_project_id"]},
            {"$set": {
                "offer_type": existing["effective_type"],
                "offer_file_path": existing.get("file_path", ""),
            }},
        )
    existing.pop("_id", None)
    return existing


@api_router.delete("/offers/{offer_id}")
async def delete_offer(offer_id: str):
    res = await db.offers.delete_one({"id": offer_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Offer not found")
    # Clear offer linkage on any project pointing to it
    await db.projects.update_many(
        {"offer_id": offer_id},
        {"$set": {"offer_id": None, "offer_code": "", "offer_type": "", "offer_file_path": ""}},
    )
    return {"ok": True}


@api_router.post("/offers/{offer_id}/convert-to-project", response_model=Project)
async def convert_offer_to_project(offer_id: str):
    offer = await db.offers.find_one({"id": offer_id}, {"_id": 0})
    if not offer:
        raise HTTPException(404, "Offer not found")
    if offer.get("linked_project_id"):
        raise HTTPException(400, f"Offer already converted to project {offer.get('linked_project_code', '')}")
    await _enrich_offer(offer)
    # Create a project from this offer
    project_doc = {
        "id": _new_id(),
        "project_code": await _next_project_code(),
        "name": offer.get("description") or f"{offer.get('effective_type', 'Project')} for {offer.get('client_name', '')}".strip(),
        "client_id": offer.get("client_id"),
        "architect_id": None,
        "site_location": offer.get("site_location", ""),
        "quoted_amount": float(offer.get("total_amount", 0) or 0),  # use GST-inclusive total
        "received_amount": 0.0,
        "status": "Outstanding",
        "notes": offer.get("notes", ""),
        "offer_id": offer["id"],
        "offer_code": offer["offer_code"],
        "offer_type": offer["effective_type"],
        "offer_file_path": offer.get("file_path", ""),
        "created_at": _now(),
    }
    await _enrich_project(project_doc)
    await db.projects.insert_one(project_doc.copy())
    # Mark offer as Accepted and linked
    await db.offers.update_one(
        {"id": offer_id},
        {"$set": {
            "status": "Accepted",
            "linked_project_id": project_doc["id"],
            "linked_project_code": project_doc["project_code"],
        }},
    )
    project_doc.pop("_id", None)
    return project_doc


# ---------------------- AUTH (single shared password) ----------------------
class PasswordVerifyIn(BaseModel):
    password: str


class PasswordSetIn(BaseModel):
    current_password: Optional[str] = None  # required if password already set
    new_password: str


def _hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def _verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


@api_router.get("/auth/status")
async def auth_status():
    doc = await db.auth_config.find_one({"_id": "auth_config"})
    return {"password_set": bool(doc and doc.get("password_hash"))}


@api_router.post("/auth/verify")
async def auth_verify(data: PasswordVerifyIn):
    doc = await db.auth_config.find_one({"_id": "auth_config"})
    if not doc or not doc.get("password_hash"):
        # No password set yet — open app
        return {"ok": True, "password_set": False}
    if not _verify_password(data.password, doc["password_hash"]):
        raise HTTPException(401, "Incorrect password")
    return {"ok": True, "password_set": True}


@api_router.post("/auth/set-password")
async def auth_set_password(data: PasswordSetIn):
    if not data.new_password or len(data.new_password) < 4:
        raise HTTPException(400, "New password must be at least 4 characters")
    existing = await db.auth_config.find_one({"_id": "auth_config"})
    if existing and existing.get("password_hash"):
        if not data.current_password or not _verify_password(data.current_password, existing["password_hash"]):
            raise HTTPException(401, "Current password is incorrect")
    new_hash = _hash_password(data.new_password)
    await db.auth_config.update_one(
        {"_id": "auth_config"},
        {"$set": {"password_hash": new_hash, "updated_at": _now()}},
        upsert=True,
    )
    return {"ok": True}


# ---------------------- OFFER PDF GENERATION ----------------------
async def _build_offer_pdf(offer: dict, client_doc: Optional[dict]) -> bytes:
    """Generate Creator RCC Consultant LLP branded offer PDF.
    All content is editable per-offer with sensible defaults."""
    # Defaults
    DEFAULT_HEADER = "CREATOR RCC CONSULTANT LLP"
    DEFAULT_TAGLINE = "Leading Project Management Consultant  |  Structural Engineer"
    DEFAULT_ADDRESS = "A-001, Siddhivinayak Park, Sector 8A, Plot No. 21, Airoli, Navi Mumbai - 400 708  |  Ph: 9987076241  |  project@creatorconsultant.net"
    DEFAULT_INTRO = (
        "We, Creator RCC Consultant LLP, are a leading structural engineering and project management "
        "consultancy authorized by BMC, NMMC and TMC. We thank you for the opportunity and are pleased "
        "to submit our offer for the captioned work as detailed below."
    )
    DEFAULT_BANK = "BANK DETAILS  |  Kotak Bank  |  A/C: Creator RCC Consultant LLP  |  A/C No: 9987076241  |  IFSC: KKBK0001360  |  Branch: Airoli, Sector 6"
    DEFAULT_SIG = "Mr. Rutvij Patel — Consulting Structural Engineer"
    DEFAULT_TCS = [
        "Taxes (GST and any other applicable levies) to be paid by the Client.",
        "Any drill-holes / chipping and their filling during testing are the responsibility of the Owner.",
        "Scope excludes any additional tests/phases not listed above; these will be charged extra by mutual agreement.",
        "Payments to be made in favour of 'CREATOR RCC CONSULTANT LLP'.",
    ]

    company_header = offer.get("company_header") or DEFAULT_HEADER
    company_tagline = offer.get("company_tagline") or DEFAULT_TAGLINE
    company_address = offer.get("company_address") or DEFAULT_ADDRESS
    intro_paragraph = offer.get("intro_paragraph") or DEFAULT_INTRO
    bank_details = offer.get("bank_details") or DEFAULT_BANK
    signature_name = offer.get("signature_name") or DEFAULT_SIG
    tcs = offer.get("terms_conditions") or DEFAULT_TCS

    base = float(offer.get("base_amount", 0) or 0)
    gst_pct = float(offer.get("gst_percent", 18) or 0)
    gst_amt = round(base * gst_pct / 100.0, 2)
    grand = round(base + gst_amt, 2)

    # Payment schedule (fallback: 50/50)
    schedule = offer.get("payment_schedule") or []
    if not schedule:
        schedule = [
            {"label": "Advance on confirmation / appointment letter", "percent": 50.0},
            {"label": "On completion of final work / submission of report", "percent": 50.0},
        ]

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, leading=14, fontName="Roboto")

    # Header band
    c.setFillColor(BRAND_GREEN)
    c.rect(0, height - 32 * mm, width, 32 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Roboto-Bold", 18)
    c.drawString(18 * mm, height - 15 * mm, company_header[:60])
    c.setFillColor(BRAND_ACCENT)
    c.setFont("Roboto", 8.5)
    c.drawString(18 * mm, height - 21 * mm, company_tagline[:110])
    c.setFillColor(colors.white)
    c.setFont("Roboto", 8)
    c.drawString(18 * mm, height - 27 * mm, company_address[:160])

    # Reference / date
    y = height - 40 * mm
    c.setFillColor(BRAND_MUTED)
    c.setFont("Roboto", 8.5)
    c.drawString(18 * mm, y, "REF. NO.")
    c.drawRightString(width - 18 * mm, y, "DATE")
    y -= 5 * mm
    c.setFillColor(colors.black)
    c.setFont("Roboto-Bold", 12)
    c.drawString(18 * mm, y, offer.get("reference_no") or offer.get("offer_code", ""))
    odate = offer.get("offer_date") or ""
    try:
        dfmt = datetime.fromisoformat(odate.replace("Z", "+00:00")).strftime("%d-%m-%Y") if odate else datetime.now().strftime("%d-%m-%Y")
    except Exception:
        dfmt = datetime.now().strftime("%d-%m-%Y")
    c.drawRightString(width - 18 * mm, y, dfmt)

    # To
    y -= 12 * mm
    c.setFillColor(BRAND_MUTED)
    c.setFont("Roboto", 8.5)
    c.drawString(18 * mm, y, "TO,")
    y -= 5 * mm
    c.setFillColor(colors.black)
    c.setFont("Roboto-Bold", 12)
    c.drawString(18 * mm, y, (client_doc or {}).get("name") or offer.get("client_name") or "Client Name")
    c.setFont("Roboto", 10)
    if client_doc and client_doc.get("company"):
        y -= 5 * mm
        c.drawString(18 * mm, y, client_doc["company"])
    if client_doc and client_doc.get("address"):
        y -= 5 * mm
        c.drawString(18 * mm, y, client_doc["address"][:90])
    if offer.get("site_location"):
        y -= 5 * mm
        c.setFillColor(BRAND_MUTED)
        c.drawString(18 * mm, y, f"Site: {offer['site_location'][:90]}")
        c.setFillColor(colors.black)

    # Subject
    y -= 10 * mm
    c.setFillColor(BRAND_GREEN)
    c.setFont("Roboto-Bold", 12)
    subject_override = (offer.get("subject") or "").strip()
    subject = subject_override or f"SUBJECT: Proposal for {offer.get('effective_type', '')} — {offer.get('description', '')[:80]}".strip().rstrip(" —")
    if not subject.upper().startswith("SUBJECT"):
        subject = "SUBJECT: " + subject
    c.drawString(18 * mm, y, subject[:110])
    c.setFillColor(colors.black)

    # Intro
    y -= 9 * mm
    p = Paragraph(intro_paragraph, body_style)
    w_para, h_para = p.wrap(width - 36 * mm, 40 * mm)
    p.drawOn(c, 18 * mm, y - h_para)
    y -= h_para + 6 * mm

    # Scope
    c.setFillColor(BRAND_GREEN)
    c.setFont("Roboto-Bold", 10)
    c.drawString(18 * mm, y, "SCOPE OF WORK")
    c.setStrokeColor(BRAND_GREEN)
    c.line(18 * mm, y - 1.5 * mm, width - 18 * mm, y - 1.5 * mm)
    y -= 8 * mm

    scope_text = (offer.get("scope_of_work") or "").strip()
    if not scope_text:
        scope_text = offer.get("description") or f"{offer.get('effective_type', '')} consultancy services as per industry standard practices."
        if offer.get("notes"):
            scope_text += f"\n\nInclusions / Methodology: {offer['notes']}"
    p = Paragraph(scope_text.replace("\n", "<br/>"), body_style)
    w_para, h_para = p.wrap(width - 36 * mm, 80 * mm)
    p.drawOn(c, 18 * mm, y - h_para)
    y -= h_para + 8 * mm
    c.setFillColor(colors.black)

    # Fees
    c.setFillColor(BRAND_GREEN)
    c.setFont("Roboto-Bold", 10)
    c.drawString(18 * mm, y, "PROFESSIONAL FEES")
    c.line(18 * mm, y - 1.5 * mm, width - 18 * mm, y - 1.5 * mm)
    y -= 6 * mm

    c.setFillColor(BRAND_GREEN)
    c.rect(18 * mm, y - 7 * mm, width - 36 * mm, 7 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Roboto-Bold", 9)
    c.drawString(22 * mm, y - 5 * mm, "DESCRIPTION")
    c.drawRightString(width - 22 * mm, y - 5 * mm, "AMOUNT (INR)")
    y -= 7 * mm

    def _row(label, amount, bold=False, highlight=False):
        nonlocal y
        y -= 7 * mm
        if highlight:
            c.setFillColor(BRAND_GREEN)
            c.rect(18 * mm, y - 1 * mm, width - 36 * mm, 7 * mm, fill=1, stroke=0)
            c.setFillColor(colors.white)
        else:
            c.setFillColor(colors.black)
        c.setFont("Roboto-Bold" if bold else "Roboto", 10)
        c.drawString(22 * mm, y + 1.5 * mm, label[:75])
        c.drawRightString(width - 22 * mm, y + 1.5 * mm, f"₹ {_format_inr(amount)}")
        c.setFillColor(colors.black)

    _row(f"{offer.get('effective_type', 'Consultancy')} charges as per scope above", base)
    _row(f"GST @ {gst_pct:.0f}%", gst_amt)
    _row("GRAND TOTAL (Inclusive of GST)", grand, bold=True, highlight=True)

    y -= 12 * mm

    # Payment Terms (editable list)
    c.setFillColor(BRAND_GREEN)
    c.setFont("Roboto-Bold", 10)
    c.drawString(18 * mm, y, "PAYMENT TERMS")
    c.line(18 * mm, y - 1.5 * mm, width - 18 * mm, y - 1.5 * mm)
    y -= 6 * mm
    c.setFillColor(colors.black)
    c.setFont("Roboto", 10)
    for entry in schedule:
        label = str(entry.get("label", "")).strip() or "—"
        pct = float(entry.get("percent", 0) or 0)
        amt = round(grand * pct / 100.0, 2)
        y -= 5 * mm
        c.drawString(20 * mm, y, f"• {pct:g}% {label}:  ₹ {_format_inr(amt)}")

    # T&C
    y -= 10 * mm
    c.setFillColor(BRAND_GREEN)
    c.setFont("Roboto-Bold", 10)
    c.drawString(18 * mm, y, "TERMS & CONDITIONS")
    c.line(18 * mm, y - 1.5 * mm, width - 18 * mm, y - 1.5 * mm)
    y -= 6 * mm
    c.setFillColor(colors.black)
    c.setFont("Roboto", 8.5)
    for t in tcs:
        y -= 5 * mm
        c.drawString(20 * mm, y, f"• {str(t)[:130]}")

    # Bank details
    y -= 10 * mm
    c.setFillColor(BRAND_GREEN)
    c.setFont("Roboto-Bold", 9)
    c.drawString(18 * mm, y, bank_details[:180])
    c.setFillColor(colors.black)

    # Signature
    c.setStrokeColor(BRAND_MUTED)
    c.line(width - 70 * mm, 30 * mm, width - 20 * mm, 30 * mm)
    c.setFillColor(BRAND_GREEN)
    c.setFont("Roboto-Bold", 10)
    c.drawString(width - 70 * mm, 35 * mm, "For Creator RCC Consultant LLP")
    c.setFillColor(BRAND_MUTED)
    c.setFont("Roboto", 8.5)
    c.drawString(width - 70 * mm, 26 * mm, "Authorised Signatory")
    c.setFont("Helvetica-Oblique", 8)
    c.drawString(width - 70 * mm, 22 * mm, signature_name[:60])

    _draw_footer(c)
    c.showPage()
    c.save()
    buf.seek(0)
    return buf.read()


@api_router.get("/offers/{offer_id}/pdf")
async def offer_pdf(offer_id: str):
    offer = await db.offers.find_one({"id": offer_id}, {"_id": 0})
    if not offer:
        raise HTTPException(404, "Offer not found")
    await _enrich_offer(offer)
    client_doc = None
    if offer.get("client_id"):
        client_doc = await db.clients.find_one({"id": offer["client_id"]}, {"_id": 0})
    pdf_bytes = await _build_offer_pdf(offer, client_doc)
    filename = f"offer_{offer['offer_code']}_{(offer.get('effective_type') or 'proposal').replace(' ', '_')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------- DASHBOARD ----------------------

@api_router.get("/dashboard/stats")
async def dashboard_stats():
    _deny_engineer()
    # Lightweight projection: we only need amounts + status for totals (no client/architect joins needed)
    projects = await db.projects.find(
        {},
        {"_id": 0, "quoted_amount": 1, "received_amount": 1, "status": 1},
    ).to_list(10000)
    total_quoted = 0.0
    total_received = 0.0
    outstanding_count = 0
    settled_count = 0
    for p in projects:
        q = float(p.get("quoted_amount", 0) or 0)
        r = float(p.get("received_amount", 0) or 0)
        total_quoted += q
        total_received += r
        # Derive effective status (same logic as _enrich_project)
        effective_status = "Settled" if (q > 0 and (q - r) <= 0) else (p.get("status") or "Outstanding")
        if effective_status == "Settled":
            settled_count += 1
        else:
            outstanding_count += 1
    return {
        "total_projects": len(projects),
        "total_clients": await db.clients.count_documents({}),
        "total_architects": await db.architects.count_documents({}),
        "total_offers": await db.offers.count_documents({}),
        "pending_offers": await db.offers.count_documents({"status": "Pending"}),
        "total_quoted": round(total_quoted, 2),
        "total_received": round(total_received, 2),
        "total_outstanding": round(total_quoted - total_received, 2),
        "outstanding_count": outstanding_count,
        "settled_count": settled_count,
    }


@api_router.get("/dashboard/monthly-revenue")
async def dashboard_monthly_revenue(months: int = 12):
    _deny_engineer()
    """Aggregate received payments (projects + audits) per month for the last `months` months."""
    months = max(1, min(int(months or 12), 36))
    from collections import OrderedDict
    now = datetime.now(timezone.utc)
    # Build 'YYYY-MM' buckets in chronological order
    buckets: "OrderedDict[str, dict]" = OrderedDict()
    year, month = now.year, now.month
    keys = []
    for _ in range(months):
        keys.append(f"{year:04d}-{month:02d}")
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    for k in reversed(keys):
        buckets[k] = {"month": k, "project_amount": 0.0, "audit_amount": 0.0, "total": 0.0, "count": 0}

    def add(date_str: Optional[str], amount: float, kind: str):
        if not date_str:
            return
        try:
            # date_str is ISO string with tz "Z" or "+00:00"
            d = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        except Exception:
            return
        k = f"{d.year:04d}-{d.month:02d}"
        if k not in buckets:
            return
        b = buckets[k]
        b[kind] += float(amount or 0)
        b["total"] = round(b["project_amount"] + b["audit_amount"], 2)
        b["count"] += 1

    project_pays = await db.payments.find({}, {"_id": 0, "amount": 1, "payment_date": 1}).to_list(100000)
    for p in project_pays:
        add(p.get("payment_date"), p.get("amount", 0), "project_amount")

    audit_pays = await db.audit_payments.find({}, {"_id": 0, "amount": 1, "payment_date": 1}).to_list(100000)
    for p in audit_pays:
        add(p.get("payment_date"), p.get("amount", 0), "audit_amount")

    rows = list(buckets.values())
    return {"months": rows, "total_received": round(sum(r["total"] for r in rows), 2)}


@api_router.get("/dashboard/site-visit-stats")
async def dashboard_site_visit_stats(days: int = 7):
    """Counts of site visits in the trailing `days` window, split by status, plus per-engineer breakdown.
    Used by the Projects dashboard 'Pending site visits this week' KPI card."""
    days = max(1, min(int(days) if days is not None else 7, 90))
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    q = {"created_at": {"$gte": cutoff}}
    rows = await db.site_visits.find(q, {"_id": 0, "id": 1, "visit_code": 1, "status": 1, "created_by_user_id": 1, "created_by_username": 1, "engineer_name": 1, "inspection_title": 1, "visit_date": 1, "project_code": 1}).sort("created_at", -1).to_list(2000)

    draft = sum(1 for r in rows if (r.get("status") or "").lower() == "draft")
    submitted = sum(1 for r in rows if (r.get("status") or "").lower() == "submitted")

    by_engineer: dict = {}
    for r in rows:
        name = r.get("engineer_name") or r.get("created_by_username") or "—"
        agg = by_engineer.setdefault(name, {"name": name, "draft": 0, "submitted": 0, "total": 0})
        s = (r.get("status") or "").lower()
        if s == "draft":
            agg["draft"] += 1
        elif s == "submitted":
            agg["submitted"] += 1
        agg["total"] += 1

    return {
        "days": days,
        "total": len(rows),
        "draft": draft,
        "submitted": submitted,
        "by_engineer": sorted(by_engineer.values(), key=lambda x: -x["total"])[:10],
        "recent_drafts": [r for r in rows if (r.get("status") or "").lower() == "draft"][:5],
    }


@api_router.get("/dashboard/my-sv-weekly")
async def dashboard_my_sv_weekly(month: Optional[str] = None, engineer_id: Optional[str] = None):
    """Per-user weekly site-visit counts for the given month (YYYY-MM, defaults to current month).
    Returns 5 buckets W1..W5 with draft+submitted counts. Engineers are auto-scoped to themselves;
    admins can pass ?engineer_id= to inspect anyone (omit for 'me')."""
    user = get_current_user_safe() or {}
    if not user.get("id"):
        raise HTTPException(401, "Not authenticated")

    target_id = user["id"]
    if engineer_id and user.get("role") == "admin":
        target_id = engineer_id

    now = datetime.now(timezone.utc)
    month_str = month or f"{now.year:04d}-{now.month:02d}"
    try:
        y, m = map(int, month_str.split("-"))
    except Exception:
        raise HTTPException(400, "month must be YYYY-MM")

    # Pull all visits this user created whose visit_date starts with the requested month
    rows = await db.site_visits.find(
        {"created_by_user_id": target_id},
        {"_id": 0, "visit_date": 0 if False else 1, "status": 1, "id": 1, "project_id": 1, "project_code": 1, "created_at": 1},
    ).to_list(2000)

    # Bucket by ISO-day-of-month / 7 (W1 = days 1-7, W2 = 8-14, ...)
    buckets = [{"week": f"W{i+1}", "draft": 0, "submitted": 0, "total": 0} for i in range(5)]
    by_project: dict = {}
    for r in rows:
        d_str = (r.get("visit_date") or r.get("created_at") or "")[:10]
        if not d_str.startswith(month_str):
            continue
        try:
            day = int(d_str.split("-")[2])
        except Exception:
            continue
        idx = min(4, max(0, (day - 1) // 7))
        s = (r.get("status") or "").lower()
        if s == "draft":
            buckets[idx]["draft"] += 1
        elif s == "submitted":
            buckets[idx]["submitted"] += 1
        buckets[idx]["total"] += 1
        # Per-project tally
        pcode = r.get("project_code") or "—"
        agg = by_project.setdefault(pcode, {"project_code": pcode, "count": 0})
        agg["count"] += 1

    return {
        "month": month_str,
        "target_user_id": target_id,
        "weeks": buckets,
        "by_project": sorted(by_project.values(), key=lambda x: -x["count"])[:8],
        "total": sum(b["total"] for b in buckets),
    }


@api_router.get("/users/{user_id}/activity")
async def user_activity_feed(user_id: str, limit: int = 100):
    """All activity_log events created by user_id, plus the user's own site visits.
    Used by the per-engineer activity feed in Settings."""
    limit = max(1, min(int(limit) if limit is not None else 100, 500))

    log_rows = await db.activity_log.find(
        {"user_id": user_id},
        {"_id": 0},
    ).sort("created_at", -1).to_list(limit)

    # Enrich with the visit_code / project_code / audit_code if missing
    for r in log_rows:
        if r.get("site_visit_id") and not r.get("site_visit_code"):
            sv = await db.site_visits.find_one({"id": r["site_visit_id"]}, {"_id": 0, "visit_code": 1})
            if sv:
                r["site_visit_code"] = sv.get("visit_code", "")
        if r.get("project_id") and not r.get("project_code"):
            p = await db.projects.find_one({"id": r["project_id"]}, {"_id": 0, "project_code": 1})
            if p:
                r["project_code"] = p.get("project_code", "")
        if r.get("audit_id") and not r.get("audit_code"):
            a = await db.audits.find_one({"id": r["audit_id"]}, {"_id": 0, "audit_code": 1})
            if a:
                r["audit_code"] = a.get("audit_code", "")

    # Also include their own site visits as a separate stream for quick scanning
    visits = await db.site_visits.find(
        {"created_by_user_id": user_id},
        {"_id": 0, "id": 1, "visit_code": 1, "inspection_title": 1, "status": 1, "visit_date": 1, "project_code": 1, "project_name": 1, "created_at": 1},
    ).sort("created_at", -1).to_list(limit)

    return {"activity": log_rows, "visits": visits}



# ---------------------- EXPORT / IMPORT ----------------------
@api_router.get("/export/excel")
async def export_excel():
    projects = await db.projects.find({}, {"_id": 0}).sort("project_code", 1).to_list(10000)
    await _enrich_projects_batch(projects)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"
    headers = [
        "Project ID", "Project Name", "Client", "Architect", "Site Location",
        "Quoted (INR)", "Received (INR)", "Outstanding (INR)", "Status", "Notes",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="061A11")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for p in projects:
        ws.append([
            p.get("project_code", ""),
            p.get("name", ""),
            p.get("client_name", ""),
            p.get("architect_name", ""),
            p.get("site_location", ""),
            p.get("quoted_amount", 0),
            p.get("received_amount", 0),
            p.get("outstanding_amount", 0),
            p.get("status", ""),
            p.get("notes", ""),
        ])

    # Auto width
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Clients sheet
    ws2 = wb.create_sheet("Clients")
    ws2.append(["Name", "Phone", "Email", "Company", "Address"])
    for col in range(1, 6):
        c = ws2.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
    clients = await db.clients.find({}, {"_id": 0}).to_list(5000)
    for c in clients:
        ws2.append([c.get("name", ""), c.get("phone", ""), c.get("email", ""), c.get("company", ""), c.get("address", "")])

    # Architects sheet
    ws3 = wb.create_sheet("Architects")
    ws3.append(["Name", "Phone", "Email", "Firm"])
    for col in range(1, 5):
        c = ws3.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
    architects = await db.architects.find({}, {"_id": 0}).to_list(5000)
    for a in architects:
        ws3.append([a.get("name", ""), a.get("phone", ""), a.get("email", ""), a.get("firm", "")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"creator_consultant_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.post("/import/excel")
async def import_excel(file: UploadFile = File(...)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Invalid Excel file: {e}")

    imported = {"projects": 0, "clients": 0, "architects": 0}

    # Clients
    if "Clients" in wb.sheetnames:
        ws = wb["Clients"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            existing = await db.clients.find_one({"name": name}, {"_id": 0})
            if existing:
                continue
            doc = {
                "id": _new_id(),
                "name": name,
                "phone": str(row[1]) if len(row) > 1 and row[1] else "",
                "email": str(row[2]) if len(row) > 2 and row[2] else "",
                "company": str(row[3]) if len(row) > 3 and row[3] else "",
                "address": str(row[4]) if len(row) > 4 and row[4] else "",
                "created_at": _now(),
            }
            await db.clients.insert_one(doc)
            imported["clients"] += 1

    # Architects
    if "Architects" in wb.sheetnames:
        ws = wb["Architects"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            existing = await db.architects.find_one({"name": name}, {"_id": 0})
            if existing:
                continue
            doc = {
                "id": _new_id(),
                "name": name,
                "phone": str(row[1]) if len(row) > 1 and row[1] else "",
                "email": str(row[2]) if len(row) > 2 and row[2] else "",
                "firm": str(row[3]) if len(row) > 3 and row[3] else "",
                "created_at": _now(),
            }
            await db.architects.insert_one(doc)
            imported["architects"] += 1

    # Projects — from first sheet or "Projects" sheet
    proj_sheet_name = "Projects" if "Projects" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[proj_sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if not row or not row[1]:  # needs a name at least
            continue
        project_code = str(row[0]).strip() if row[0] else await _next_project_code()
        # Skip if already exists
        if await db.projects.find_one({"project_code": project_code}):
            continue

        client_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        architect_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        # Resolve/create client and architect
        client_id = None
        if client_name and client_name.lower() != "none":
            c = await db.clients.find_one({"name": client_name}, {"_id": 0})
            if not c:
                c = {"id": _new_id(), "name": client_name, "phone": "", "email": "", "company": "", "address": "", "created_at": _now()}
                await db.clients.insert_one(c.copy())
                imported["clients"] += 1
            client_id = c["id"]

        architect_id = None
        if architect_name and architect_name.lower() != "none":
            a = await db.architects.find_one({"name": architect_name}, {"_id": 0})
            if not a:
                a = {"id": _new_id(), "name": architect_name, "phone": "", "email": "", "firm": "", "created_at": _now()}
                await db.architects.insert_one(a.copy())
                imported["architects"] += 1
            architect_id = a["id"]

        def _num(v):
            try:
                return float(v) if v is not None and v != "" else 0.0
            except Exception:
                return 0.0

        quoted = _num(row[5] if len(row) > 5 else 0)
        received = _num(row[6] if len(row) > 6 else 0)
        status = str(row[8]).strip() if len(row) > 8 and row[8] else "Outstanding"
        notes = str(row[9]) if len(row) > 9 and row[9] else ""

        doc = {
            "id": _new_id(),
            "project_code": project_code,
            "name": str(row[1]).strip(),
            "client_id": client_id,
            "client_name": client_name,
            "architect_id": architect_id,
            "architect_name": architect_name,
            "site_location": str(row[4]) if len(row) > 4 and row[4] else "",
            "quoted_amount": quoted,
            "received_amount": received,
            "outstanding_amount": round(quoted - received, 2),
            "status": "Settled" if (quoted > 0 and received >= quoted) else (status or "Outstanding"),
            "notes": notes,
            "created_at": _now(),
        }
        await db.projects.insert_one(doc)
        imported["projects"] += 1

    return {"ok": True, "imported": imported}


@api_router.post("/import/sqlite")
async def import_sqlite(file: UploadFile = File(...), replace: bool = False):
    """Import data from an uploaded SQLite .db file (creator_consultant legacy format).
    Expects tables: clients(id, name), architects(id, name), projects(id, project_id, project_name,
    client_name, architect_name, site_location, quoted_amount, created_at), payments(id, project_id, amount, note, paid_at).
    If replace=True, clears existing data first.
    """
    import sqlite3
    import tempfile

    content = await file.read()
    if not content[:16].startswith(b"SQLite"):
        raise HTTPException(400, "Uploaded file is not a valid SQLite database")

    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as tf:
        tf.write(content)
        tmp_path = tf.name
    try:
        if replace:
            await db.projects.delete_many({})
            await db.clients.delete_many({})
            await db.architects.delete_many({})
            await db.payments.delete_many({})
            await db.counters.delete_many({})

        conn = sqlite3.connect(tmp_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        def _parse_dt(s):
            if not s:
                return _now()
            try:
                return datetime.strptime(s, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc).isoformat()
            except Exception:
                return _now()

        imported = {"clients": 0, "architects": 0, "projects": 0, "payments": 0}

        # Clients (dedupe by name, keep existing)
        client_by_name = {}
        existing_clients = await db.clients.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(10000)
        for ec in existing_clients:
            client_by_name[ec["name"].strip().lower()] = ec["id"]
        try:
            cur.execute("SELECT name FROM clients")
            for r in cur.fetchall():
                nm = (r["name"] or "").strip()
                if not nm or nm.lower() == "none":
                    continue
                key = nm.lower()
                if key in client_by_name:
                    continue
                cid = _new_id()
                await db.clients.insert_one({
                    "id": cid, "name": nm, "phone": "", "email": "",
                    "company": "", "address": "", "created_at": _now(),
                })
                client_by_name[key] = cid
                imported["clients"] += 1
        except sqlite3.Error:
            pass

        # Architects
        arch_by_name = {}
        existing_arch = await db.architects.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(10000)
        for ea in existing_arch:
            arch_by_name[ea["name"].strip().lower()] = ea["id"]
        try:
            cur.execute("SELECT name FROM architects")
            for r in cur.fetchall():
                nm = (r["name"] or "").strip()
                if not nm or nm.lower() == "none":
                    continue
                key = nm.lower()
                if key in arch_by_name:
                    continue
                aid = _new_id()
                await db.architects.insert_one({
                    "id": aid, "name": nm, "phone": "", "email": "", "firm": "",
                    "created_at": _now(),
                })
                arch_by_name[key] = aid
                imported["architects"] += 1
        except sqlite3.Error:
            pass

        # Projects
        projects_by_code = {}
        existing_codes = {p["project_code"] async for p in db.projects.find({}, {"_id": 0, "project_code": 1})}
        max_seq = 0
        for code in existing_codes:
            try:
                s = int(str(code).split("-")[-1])
                if s > max_seq:
                    max_seq = s
            except Exception:
                pass
        try:
            cur.execute("SELECT * FROM projects ORDER BY id")
            for p in cur.fetchall():
                code = p["project_id"] or f"CC-{p['id']:04d}"
                if code in existing_codes:
                    projects_by_code[code] = None
                    continue
                try:
                    s = int(str(code).split("-")[-1])
                    if s > max_seq:
                        max_seq = s
                except Exception:
                    pass
                cl_name = (p["client_name"] or "").strip()
                ar_name = (p["architect_name"] or "").strip()
                pid = _new_id()
                doc = {
                    "id": pid,
                    "project_code": code,
                    "name": (p["project_name"] or "").strip() or "Untitled",
                    "client_id": client_by_name.get(cl_name.lower()),
                    "client_name": cl_name if cl_name.lower() != "none" else "",
                    "architect_id": arch_by_name.get(ar_name.lower()),
                    "architect_name": ar_name if ar_name.lower() != "none" else "",
                    "site_location": (p["site_location"] or "").strip(),
                    "quoted_amount": float(p["quoted_amount"] or 0),
                    "received_amount": 0.0,
                    "outstanding_amount": 0.0,
                    "status": "Outstanding",
                    "notes": "",
                    "archived": False,
                    "offer_id": None, "offer_code": "", "offer_type": "", "offer_file_path": "",
                    "created_at": _parse_dt(p["created_at"]),
                }
                await db.projects.insert_one(doc)
                projects_by_code[code] = pid
                imported["projects"] += 1
        except sqlite3.Error as e:
            raise HTTPException(400, f"Projects import failed: {e}")

        # Payments
        received_by_code = {}
        try:
            cur.execute("SELECT * FROM payments ORDER BY id")
            for pay in cur.fetchall():
                code = pay["project_id"]
                pid = projects_by_code.get(code)
                if not pid:
                    # project might exist from earlier — look up by code
                    existing = await db.projects.find_one({"project_code": code}, {"_id": 0, "id": 1})
                    if not existing:
                        continue
                    pid = existing["id"]
                amt = float(pay["amount"] or 0)
                if amt <= 0:
                    continue
                await db.payments.insert_one({
                    "id": _new_id(),
                    "project_id": pid,
                    "project_code": code,
                    "amount": amt,
                    "taxable_amount": amt,
                    "payment_date": _parse_dt(pay["paid_at"]),
                    "notes": (pay["note"] or "").strip(),
                    "created_at": _parse_dt(pay["paid_at"]),
                })
                received_by_code[code] = pid
                imported["payments"] += 1
        except sqlite3.Error:
            pass

        # Update totals on projects
        for code, pid in received_by_code.items():
            await _recalculate_project_received(pid)

        # Update counter
        await db.counters.update_one(
            {"_id": "project_code"}, {"$max": {"seq": max_seq}}, upsert=True
        )
        conn.close()
        return {"ok": True, "imported": imported}
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


# ---------------------- SEED ----------------------
@api_router.post("/seed")
async def seed_demo():
    """Populate demo data if DB is empty."""
    if await db.projects.count_documents({}) > 0:
        return {"ok": True, "seeded": False, "message": "Data already exists"}

    # Clients
    clients_data = [
        {"name": "M/S. Kian Dines Pvt. Ltd.", "phone": "+91 98200 11111", "email": "info@kiandines.com", "company": "Kian Dines Pvt. Ltd.", "address": "Rabale MIDC, Navi Mumbai"},
        {"name": "Mrs. Husna Ara Sayed", "phone": "+91 98765 43210", "email": "husna.sayed@example.com", "company": "", "address": "Karanja Road, Rajapal Naka"},
        {"name": "Rohan Enterprises", "phone": "+91 90000 22222", "email": "contact@rohanent.in", "company": "Rohan Enterprises", "address": "Ambernath MIDC"},
        {"name": "Sunrise Developers", "phone": "+91 91234 56789", "email": "sales@sunrisedev.in", "company": "Sunrise Developers LLP", "address": "Novo Rabale MIDC"},
    ]
    client_ids = []
    for c in clients_data:
        doc = {**c, "id": _new_id(), "created_at": _now()}
        await db.clients.insert_one(doc.copy())
        client_ids.append(doc["id"])

    # Architects
    architects_data = [
        {"name": "Tanay Mehta", "phone": "+91 99999 00001", "email": "tanay@mehtaarch.in", "firm": "Mehta & Associates"},
        {"name": "Tstg Architect", "phone": "+91 88888 00002", "email": "studio@tstg.in", "firm": "TSTG Studio"},
        {"name": "Priya Shah", "phone": "+91 77777 00003", "email": "priya@shahdesigns.in", "firm": "Shah Designs"},
    ]
    arch_ids = []
    for a in architects_data:
        doc = {**a, "id": _new_id(), "created_at": _now()}
        await db.architects.insert_one(doc.copy())
        arch_ids.append(doc["id"])

    # Projects
    projects_data = [
        {"name": "Acceptance & Supervision", "client_id": client_ids[0], "architect_id": arch_ids[1], "site_location": "Plot No - Pap R 641 Rabale Midc, Navi Mumbai", "quoted_amount": 125000, "received_amount": 50000, "status": "Outstanding", "notes": "Phase 1 supervision"},
        {"name": "Residential Plan 3269", "client_id": client_ids[2], "architect_id": arch_ids[0], "site_location": "Plot No. RTC-91, Ambernath", "quoted_amount": 85000, "received_amount": 85000, "status": "Settled", "notes": ""},
        {"name": "Industrial Shed M-85", "client_id": client_ids[2], "architect_id": arch_ids[0], "site_location": "Plot No. M-85, Ambernath MIDC", "quoted_amount": 210000, "received_amount": 210000, "status": "Settled", "notes": "Full payment received"},
        {"name": "Commercial Design 3241", "client_id": client_ids[3], "architect_id": arch_ids[0], "site_location": "Plot No - Pap - 73 Ambernath Midc", "quoted_amount": 175000, "received_amount": 75000, "status": "Outstanding", "notes": ""},
        {"name": "Acceptance & Supervision (Karanja)", "client_id": client_ids[1], "architect_id": arch_ids[1], "site_location": "Cts No - 959+959/1, House No - 92, Karanja Road, Rajapal Naka", "quoted_amount": 95000, "received_amount": 95000, "status": "Settled", "notes": ""},
        {"name": "Novo Rabale Project 3240", "client_id": client_ids[3], "architect_id": arch_ids[0], "site_location": "Plot No - 374 Novo Rabale Midc", "quoted_amount": 150000, "received_amount": 0, "status": "Outstanding", "notes": "Kickoff pending"},
        {"name": "Villa Renovation", "client_id": client_ids[1], "architect_id": arch_ids[2], "site_location": "Bandra West, Mumbai", "quoted_amount": 320000, "received_amount": 120000, "status": "Outstanding", "notes": ""},
    ]
    for p in projects_data:
        code = await _next_project_code()
        doc = {
            "id": _new_id(),
            "project_code": code,
            "name": p["name"],
            "client_id": p["client_id"],
            "architect_id": p["architect_id"],
            "site_location": p["site_location"],
            "quoted_amount": float(p["quoted_amount"]),
            "received_amount": float(p["received_amount"]),
            "status": p["status"],
            "notes": p.get("notes", ""),
            "created_at": _now(),
        }
        await _enrich_project(doc)
        await db.projects.insert_one(doc.copy())

    # Offers (sample pending + accepted)
    offers_data = [
        {
            "offer_type": "Audit", "custom_type": "",
            "client_id": client_ids[1],
            "description": "RCC-Basic-Audit of Row House",
            "site_location": "Plot 44, Sector 4, Koparkhairane, Navi Mumbai",
            "base_amount": 28000, "gst_percent": 18.0,
            "file_path": "D:\\CreatorConsultant\\Offers\\2026\\STR-AUDIT-2026-023.pdf",
            "status": "Pending",
            "reference_no": "STR/AUDIT/2026/023",
            "offer_date": _now(),
            "notes": "Half cell + Rebound Hammer + Carbonation + UPV. 50% advance.",
        },
        {
            "offer_type": "Steel", "custom_type": "",
            "client_id": client_ids[2],
            "description": "MS Structural Design & Consultancy",
            "site_location": "TTC IND. Area, Rabale MIDC, Navi Mumbai",
            "base_amount": 200000, "gst_percent": 18.0,
            "file_path": "D:\\CreatorConsultant\\Offers\\2025\\STR-QUOT-2025-160.pdf",
            "status": "Pending",
            "reference_no": "STR/QUOT/2025/160",
            "offer_date": _now(),
            "notes": "20,000 sq.ft. @ Rs 10/sq.ft.",
        },
        {
            "offer_type": "Other", "custom_type": "PMC",
            "client_id": client_ids[3],
            "description": "Project Management Consultancy",
            "site_location": "Novo Rabale MIDC",
            "base_amount": 150000, "gst_percent": 18.0,
            "file_path": "D:\\CreatorConsultant\\Offers\\2026\\PMC-offer.pdf",
            "status": "Pending",
            "reference_no": "STR/PMC/2026/005",
            "offer_date": _now(),
            "notes": "Quarterly site visits + BOQ review",
        },
    ]
    for o in offers_data:
        od = {
            "id": _new_id(),
            "offer_code": await _next_offer_code(),
            "offer_type": o["offer_type"],
            "custom_type": o["custom_type"],
            "client_id": o["client_id"],
            "description": o["description"],
            "site_location": o["site_location"],
            "base_amount": float(o["base_amount"]),
            "gst_percent": float(o["gst_percent"]),
            "file_path": o["file_path"],
            "status": o["status"],
            "offer_date": o["offer_date"],
            "reference_no": o["reference_no"],
            "notes": o["notes"],
            "linked_project_id": None,
            "linked_project_code": "",
            "created_at": _now(),
        }
        await _enrich_offer(od)
        await db.offers.insert_one(od.copy())

    return {"ok": True, "seeded": True}


# ---------------------- DOCUMENTS MODULE ----------------------

DEFAULT_DOCUMENT_TYPES = [
    ("Quotation", "QT"),
    ("PMC Quotation", "PMC-QT"),
    ("Inspection Report Letter", "INSP"),
    ("Acceptance Letter", "ACC"),
    ("Demolition Letter", "DEM"),
    ("Supervision Certificate", "SUP"),
    ("Stability Certificate", "STAB"),
    ("To Whomsoever It May Concern", "TWMC"),
    ("Earthquake Certificate", "EQ"),
    ("Commencement Certificate", "COM"),
    ("MCGM Certificate", "MCGM"),
    ("Plinth Completion Certificate", "PLINTH"),
    ("Column Location Certificate", "COL"),
    ("RERA Certificate", "RERA"),
    ("Lift Certificate", "LIFT"),
    ("General Certificate", "GEN"),
    ("Scaffolding Certificate", "SCAF"),
    ("Declaration Certificate", "DEC"),
]


async def _seed_document_types_if_missing():
    if await db.document_types.count_documents({}) > 0:
        # One-shot migration: drop legacy "Audit Report" doc type if it exists
        # and hasn't yet been used to issue a number (counter ≤ 0).
        await db.document_types.delete_many({
            "prefix": "AUD-RPT",
            "$or": [{"counter": {"$lte": 0}}, {"counter": {"$exists": False}}],
        })
        # One-shot migration: drop "Audit Offer" doc type entirely. Audit offer
        # numbers are now entered manually on the Audit form (no auto-numbering).
        await db.document_types.delete_many({"prefix": "AUD-OFR"})
        return
    now = _now()
    docs = []
    for name, prefix in DEFAULT_DOCUMENT_TYPES:
        docs.append({
            "id": _new_id(),
            "name": name,
            "prefix": prefix,
            "description": "",
            "year_reset": True,
            "counter": 0,
            "last_year": 0,
            "created_at": now,
        })
    if docs:
        await db.document_types.insert_many(docs)


async def _next_document_number(doc_type: dict) -> str:
    """Generate `STR/{prefix}/{YYYY}/{counter:03}` and atomically increment counter on the type."""
    year = datetime.now(timezone.utc).year
    if doc_type.get("year_reset", True) and (doc_type.get("last_year") or 0) != year:
        await db.document_types.update_one(
            {"id": doc_type["id"]},
            {"$set": {"counter": 1, "last_year": year}},
        )
        counter = 1
    else:
        updated = await db.document_types.find_one_and_update(
            {"id": doc_type["id"]},
            {"$inc": {"counter": 1}, "$set": {"last_year": year}},
            return_document=True,
        )
        counter = (updated or {}).get("counter", 1)
    prefix = (doc_type.get("prefix") or "DOC").strip()
    return f"STR/{prefix}/{year}/{counter:03d}"


@api_router.get("/document-types", response_model=List[DocumentType])
async def list_document_types():
    rows = await db.document_types.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return rows


@api_router.get("/document-types/by-prefix/{prefix}/preview")
async def preview_document_number_by_prefix(prefix: str):
    """Return what the next document number WILL be for the given prefix, without
    incrementing the counter. Used to pre-fill the Audit Offer Number on the
    New Audit form so the user sees the upcoming `STR/AUD-OFR/2026/007`."""
    dt = await db.document_types.find_one({"prefix": prefix.upper()}, {"_id": 0})
    if not dt:
        raise HTTPException(404, f"No document type with prefix '{prefix}'")
    year = datetime.now(timezone.utc).year
    if dt.get("year_reset", True) and (dt.get("last_year") or 0) != year:
        next_counter = 1
    else:
        next_counter = (dt.get("counter") or 0) + 1
    number = f"STR/{dt['prefix']}/{year}/{next_counter:03d}"
    return {"number": number, "year": year, "counter": next_counter, "prefix": dt["prefix"]}


@api_router.post("/document-types", response_model=DocumentType)
async def create_document_type(data: DocumentTypeIn):
    prefix = (data.prefix or "").strip().upper().replace(" ", "-")
    if not prefix:
        raise HTTPException(400, "Prefix is required")
    if await db.document_types.find_one({"prefix": prefix}):
        raise HTTPException(400, f"A document type with prefix '{prefix}' already exists")
    doc = {
        "id": _new_id(),
        "name": data.name.strip(),
        "prefix": prefix,
        "description": data.description or "",
        "year_reset": bool(data.year_reset),
        "counter": 0,
        "last_year": 0,
        "created_at": _now(),
    }
    await db.document_types.insert_one(doc.copy())
    return doc


@api_router.put("/document-types/{type_id}", response_model=DocumentType)
async def update_document_type(type_id: str, data: DocumentTypeIn):
    existing = await db.document_types.find_one({"id": type_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Document type not found")
    prefix = (data.prefix or "").strip().upper().replace(" ", "-")
    if not prefix:
        raise HTTPException(400, "Prefix is required")
    dup = await db.document_types.find_one({"prefix": prefix, "id": {"$ne": type_id}})
    if dup:
        raise HTTPException(400, f"A document type with prefix '{prefix}' already exists")
    update = {
        "name": data.name.strip(),
        "prefix": prefix,
        "description": data.description or "",
        "year_reset": bool(data.year_reset),
    }
    await db.document_types.update_one({"id": type_id}, {"$set": update})
    merged = {**existing, **update}
    return merged


@api_router.put("/document-types/{type_id}/counter")
async def reset_document_type_counter(type_id: str, counter: int = 0, last_year: Optional[int] = None):
    existing = await db.document_types.find_one({"id": type_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Document type not found")
    new_year = last_year if last_year is not None else (existing.get("last_year") or 0)
    await db.document_types.update_one(
        {"id": type_id},
        {"$set": {"counter": max(int(counter), 0), "last_year": int(new_year)}},
    )
    return {"ok": True, "counter": int(counter), "last_year": int(new_year)}


@api_router.delete("/document-types/{type_id}")
async def delete_document_type(type_id: str):
    count = await db.documents.count_documents({"doc_type_id": type_id})
    if count > 0:
        raise HTTPException(400, f"Cannot delete — {count} document(s) exist for this type. Delete them first or rename the type.")
    res = await db.document_types.delete_one({"id": type_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Document type not found")
    return {"ok": True}


async def _enrich_document(d: dict) -> dict:
    # Backfill status field for documents created before the status feature existed
    if not d.get("status"):
        d["status"] = "confirmed" if d.get("confirmed") else "pending"
    if d.get("client_id"):
        c = await db.clients.find_one({"id": d["client_id"]}, {"_id": 0, "name": 1, "phone": 1, "email": 1, "address": 1})
        if c:
            d["client_name"] = c.get("name", "")
            d["client_phone"] = c.get("phone", "")
            d["client_email"] = c.get("email", "")
            d["client_address"] = c.get("address", "")
    if d.get("architect_id"):
        a = await db.architects.find_one({"id": d["architect_id"]}, {"_id": 0, "name": 1, "phone": 1, "email": 1, "firm": 1})
        if a:
            d["architect_name"] = a.get("name", "")
            d["architect_phone"] = a.get("phone", "")
            d["architect_email"] = a.get("email", "")
            d["architect_firm"] = a.get("firm", "")
    return d


@api_router.get("/documents", response_model=List[Document])
async def list_documents(
    type_id: Optional[str] = None,
    client_id: Optional[str] = None,
    architect_id: Optional[str] = None,
    archived: Optional[bool] = False,
    search: Optional[str] = None,
):
    q: dict = {"archived": True} if archived else {"archived": {"$ne": True}}
    if type_id: q["doc_type_id"] = type_id
    if client_id: q["client_id"] = client_id
    if architect_id: q["architect_id"] = architect_id
    if search:
        rx = {"$regex": search, "$options": "i"}
        q["$or"] = [
            {"doc_number": rx}, {"doc_type_name": rx}, {"client_name": rx}, {"architect_name": rx},
            {"plot_place": rx}, {"contact_person": rx}, {"mobile": rx}, {"remark": rx},
        ]
    rows = await db.documents.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    for r in rows:
        await _enrich_document(r)
    return rows


@api_router.get("/documents/{doc_id}", response_model=Document)
async def get_document(doc_id: str):
    d = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Document not found")
    await _enrich_document(d)
    return d


@api_router.post("/documents", response_model=Document)
async def create_document(data: DocumentIn):
    dt = await db.document_types.find_one({"id": data.doc_type_id}, {"_id": 0})
    if not dt:
        raise HTTPException(400, "Invalid document type")
    doc_number = (data.doc_number or "").strip() or await _next_document_number(dt)
    doc = {
        "id": _new_id(),
        "doc_type_id": dt["id"],
        "doc_type_name": dt["name"],
        "doc_number": doc_number,
        "document_date": data.document_date or _now(),
        "client_id": data.client_id or None,
        "client_name": "",
        "architect_id": data.architect_id or None,
        "architect_name": "",
        "plot_place": data.plot_place or "",
        "phase": data.phase or "",
        "number_field": data.number_field or "",
        "remark": data.remark or "",
        "contact_person": data.contact_person or "",
        "mobile": data.mobile or "",
        "other_comments": data.other_comments or "",
        "update_date": data.update_date or None,
        "archived": False,
        "created_at": _now(),
    }
    _stamp_edit(doc)
    await _enrich_document(doc)
    await db.documents.insert_one(doc.copy())
    return doc


@api_router.put("/documents/{doc_id}", response_model=Document)
async def update_document(doc_id: str, data: DocumentIn):
    existing = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Document not found")
    dt = await db.document_types.find_one({"id": data.doc_type_id}, {"_id": 0})
    if not dt:
        raise HTTPException(400, "Invalid document type")
    update = {
        "doc_type_id": dt["id"],
        "doc_type_name": dt["name"],
        "doc_number": (data.doc_number or existing.get("doc_number") or "").strip(),
        "document_date": data.document_date or existing.get("document_date"),
        "client_id": data.client_id or None,
        "architect_id": data.architect_id or None,
        "plot_place": data.plot_place or "",
        "phase": data.phase or "",
        "number_field": data.number_field or "",
        "remark": data.remark or "",
        "contact_person": data.contact_person or "",
        "mobile": data.mobile or "",
        "other_comments": data.other_comments or "",
        "update_date": data.update_date or existing.get("update_date"),
    }
    _stamp_edit(update)
    await db.documents.update_one({"id": doc_id}, {"$set": update})
    merged = {**existing, **update}
    await _enrich_document(merged)
    return merged


@api_router.delete("/documents/{doc_id}")
async def delete_document(doc_id: str):
    res = await db.documents.delete_one({"id": doc_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Document not found")
    return {"ok": True}


@api_router.post("/documents/{doc_id}/archive")
async def archive_document(doc_id: str):
    res = await db.documents.update_one({"id": doc_id}, {"$set": {"archived": True}})
    if res.matched_count == 0:
        raise HTTPException(404, "Document not found")
    return {"ok": True}


class ConfirmDocumentIn(BaseModel):
    project_id: Optional[str] = None
    audit_id: Optional[str] = None


class DocumentStatusIn(BaseModel):
    status: str  # one of: pending | confirmed | on_hold | cancelled
    project_id: Optional[str] = None
    audit_id: Optional[str] = None


_DOC_STATUSES = {"pending", "confirmed", "on_hold", "cancelled"}


async def _apply_document_status(doc_id: str, status: str, project_id: Optional[str], audit_id: Optional[str]) -> dict:
    if status not in _DOC_STATUSES:
        raise HTTPException(400, f"Invalid status. Use one of: {sorted(_DOC_STATUSES)}")
    existing = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Document not found")
    if project_id and audit_id:
        raise HTTPException(400, "Link to either a project OR an audit, not both")

    update: dict = {"status": status}
    is_confirmed = status == "confirmed"
    update["confirmed"] = is_confirmed
    update["confirmed_at"] = _now() if is_confirmed else None

    if status in ("on_hold", "cancelled", "pending"):
        # Clear links when status is not 'confirmed'
        update.update({
            "linked_project_id": None, "linked_project_code": "", "linked_project_name": "",
            "linked_audit_id": None, "linked_audit_code": "", "linked_audit_offer": "",
        })
    elif is_confirmed:
        if project_id:
            proj = await db.projects.find_one({"id": project_id}, {"_id": 0, "project_code": 1, "name": 1})
            if not proj:
                raise HTTPException(404, "Project not found")
            update.update({
                "linked_project_id": project_id,
                "linked_project_code": proj.get("project_code", ""),
                "linked_project_name": proj.get("name", ""),
                "linked_audit_id": None, "linked_audit_code": "", "linked_audit_offer": "",
            })
        elif audit_id:
            audit = await db.audits.find_one({"id": audit_id}, {"_id": 0, "audit_code": 1, "audit_offer": 1})
            if not audit:
                raise HTTPException(404, "Audit not found")
            update.update({
                "linked_audit_id": audit_id,
                "linked_audit_code": audit.get("audit_code", ""),
                "linked_audit_offer": audit.get("audit_offer", ""),
                "linked_project_id": None, "linked_project_code": "", "linked_project_name": "",
            })
        else:
            # Confirmed without link → clear both
            update.update({
                "linked_project_id": None, "linked_project_code": "", "linked_project_name": "",
                "linked_audit_id": None, "linked_audit_code": "", "linked_audit_offer": "",
            })
    _stamp_edit(update)
    await db.documents.update_one({"id": doc_id}, {"$set": update})
    return update


@api_router.post("/documents/{doc_id}/status")
async def set_document_status(doc_id: str, data: DocumentStatusIn):
    update = await _apply_document_status(doc_id, data.status, data.project_id, data.audit_id)
    return {"ok": True, **update}


@api_router.post("/documents/{doc_id}/confirm")
async def confirm_document(doc_id: str, data: ConfirmDocumentIn):
    """Backward-compat: confirm + optional link. Equivalent to status='confirmed'."""
    update = await _apply_document_status(doc_id, "confirmed", data.project_id, data.audit_id)
    return {"ok": True, **update}


@api_router.post("/documents/{doc_id}/unconfirm")
async def unconfirm_document(doc_id: str):
    """Backward-compat: reset to pending."""
    update = await _apply_document_status(doc_id, "pending", None, None)
    return {"ok": True, **update}


@api_router.post("/documents/{doc_id}/unarchive")
async def unarchive_document(doc_id: str):
    res = await db.documents.update_one({"id": doc_id}, {"$set": {"archived": False}})
    if res.matched_count == 0:
        raise HTTPException(404, "Document not found")
    return {"ok": True}


@api_router.get("/documents/{doc_id}/pdf")
async def document_pdf(doc_id: str):
    d = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Document not found")
    await _enrich_document(d)

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 18 * mm

    # Branded header
    c.setFillColor(colors.HexColor("#0A2E1F"))
    c.rect(0, height - 28 * mm, width, 28 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Roboto-Bold", 16)
    c.drawString(margin, height - 14 * mm, "CREATOR RCC CONSULTANT LLP")
    c.setFont("Roboto", 8.5)
    c.drawString(margin, height - 20 * mm, "Structural Audits • RCC / Steel Design • PMC • Retrofitting")
    c.drawString(margin, height - 25 * mm, "Navi Mumbai • info@creatorconsultant.online")
    c.setFillColor(colors.HexColor("#10B981"))
    c.setFont("Roboto-Bold", 12)
    c.drawRightString(width - margin, height - 14 * mm, d.get("doc_type_name", "Document").upper())
    c.setFillColor(colors.white)
    c.setFont("Roboto-Bold", 10)
    c.drawRightString(width - margin, height - 20 * mm, d.get("doc_number", ""))

    y = height - 38 * mm
    c.setFillColor(colors.black)
    c.setFont("Roboto", 10)
    c.drawRightString(width - margin, y, f"Date: {(d.get('document_date') or _now())[:10]}")
    y -= 10 * mm

    # TO block
    c.setFont("Roboto-Bold", 10)
    c.drawString(margin, y, "TO,")
    y -= 5 * mm
    c.setFont("Roboto", 10)
    if d.get("client_name"):
        c.drawString(margin, y, d.get("client_name", "")); y -= 5 * mm
    if d.get("contact_person") and d.get("contact_person") != d.get("client_name"):
        c.drawString(margin, y, f"Kind Attn.: {d.get('contact_person')}"); y -= 5 * mm
    if d.get("plot_place"):
        c.drawString(margin, y, d.get("plot_place", "")); y -= 5 * mm
    if d.get("mobile"):
        c.drawString(margin, y, f"Mobile: {d.get('mobile')}"); y -= 5 * mm

    y -= 4 * mm
    c.setFont("Roboto-Bold", 12)
    c.setFillColor(colors.HexColor("#0A2E1F"))
    c.drawString(margin, y, f"Subject: {d.get('doc_type_name', 'Document')}")
    c.setFillColor(colors.black)
    y -= 8 * mm

    rows = [
        ("Location", d.get("phase")),
        ("Path of Folder", d.get("remark")),
        ("Other Comments", d.get("other_comments")),
    ]
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle("body", parent=styles["Normal"], fontName="Roboto", fontSize=10, leading=14)
    label_x = margin
    value_x = margin + 38 * mm
    for label, val in rows:
        if not val:
            continue
        c.setFont("Roboto-Bold", 10)
        c.drawString(label_x, y, f"{label}:")
        para = Paragraph(str(val).replace("\n", "<br/>"), body_style)
        avail_w = width - value_x - margin
        w, h = para.wrap(avail_w, 100 * mm)
        para.drawOn(c, value_x, y - h + 11)
        y -= max(h + 3, 6 * mm)
        if y < 40 * mm:
            c.showPage(); y = height - margin

    y = max(y - 18 * mm, 40 * mm)
    c.setFont("Roboto", 10)
    c.drawString(margin, y, "For Creator RCC Consultant LLP")
    y -= 14 * mm
    c.drawString(margin, y, "____________________________")
    y -= 5 * mm
    c.setFont("Roboto-Bold", 10)
    c.drawString(margin, y, "Mr. Rutvij Patel")
    y -= 5 * mm
    c.setFont("Roboto", 8.5)
    c.drawString(margin, y, "Consulting Structural Engineer")

    c.showPage(); c.save()
    pdf_bytes = buf.getvalue(); buf.close()
    fname = f"{d.get('doc_number','document').replace('/', '_')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )



# ---------------------- SITE VISIT (Engineer) ENDPOINTS ----------------------

DEFAULT_SV_TEMPLATES = [
    ("Column Inspection", "Column casting / reinforcement checks", [
        "Size of members as per drawing",
        "Reinforcement - Dia, No of bars",
        "Spacing between bars, LD, Alignment of bars",
        "Reinforcement - Rust free, dust free",
        "Clear cover to reinforcement",
        "Rods are tied properly with binding wire",
        "Ties spacing, alignment, hook bend 135°",
        "Concrete cube test result for last structural member",
    ]),
    ("Slab Inspection", "Slab casting / reinforcement checks", [
        "Size and thickness of slab as per drawing",
        "Reinforcement - Dia, No of bars",
        "Spacing between bars, Alignment",
        "Top & bottom bar cover",
        "Crank bar locations",
        "Chairs / spacers between layers",
        "Electrical conduit & plumbing sleeves in place",
        "Concrete cube test result for last structural member",
    ]),
    ("Beam Inspection", "Beam casting / reinforcement checks", [
        "Size of beam as per drawing",
        "Reinforcement - Dia, No of bars",
        "Stirrups spacing, alignment, hook bend 135°",
        "Clear cover to reinforcement",
        "Anchorage length (LD) at supports",
        "Lap length / position",
        "Concrete cube test result for last structural member",
    ]),
    ("Foundation Inspection", "Footing / raft / pile cap checks", [
        "Excavation level & soil bearing condition",
        "PCC laid and level checked",
        "Reinforcement - Dia, No, spacing as per drawing",
        "Clear cover all sides",
        "Dowel bars for columns properly tied",
        "Anti-termite treatment done",
        "Concrete cube test result for last structural member",
    ]),
    ("Waterproofing Inspection", "Waterproofing / membrane checks", [
        "Surface cleaned & primer applied",
        "Membrane laid as per spec",
        "Overlap and seam joints sealed",
        "Corners and outlets treated",
        "Protection screed laid",
        "Pond / water test done for 48 hours",
    ]),
]


async def _seed_sv_templates_if_missing():
    if await db.site_visit_templates.count_documents({}) > 0:
        return
    now = _now()
    docs = []
    for name, desc, items in DEFAULT_SV_TEMPLATES:
        docs.append({
            "id": _new_id(),
            "name": name,
            "description": desc,
            "checklist": items,
            "created_at": now,
        })
    if docs:
        await db.site_visit_templates.insert_many(docs)


async def _next_visit_code() -> str:
    counter = await db.counters.find_one_and_update(
        {"_id": "site_visit"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = (counter or {}).get("seq", 1)
    return f"SV-{seq:04d}"


# ----- Admin: edit the Site Visit number series (next visit code) -----
class SiteVisitSeqIn(BaseModel):
    next_seq: int = Field(..., ge=1, le=999999)


@api_router.get("/site-visits/series")
async def get_site_visit_series():
    """Return the next visit code that will be assigned. Admins only."""
    _deny_engineer()
    counter = await db.counters.find_one({"_id": "site_visit"}, {"_id": 0})
    current_seq = (counter or {}).get("seq", 0)
    next_seq = current_seq + 1
    return {
        "current_seq": current_seq,
        "next_seq": next_seq,
        "next_code": f"SV-{next_seq:04d}",
        "prefix": "SV",
    }


@api_router.put("/site-visits/series")
async def set_site_visit_series(body: SiteVisitSeqIn, user: dict = Depends(auth_module.get_current_user)):
    """Admin sets the NEXT visit code. e.g. next_seq=100 → next visit is SV-0100.
    Internally we set `counter.seq = next_seq - 1` so the next $inc returns
    `next_seq`. Refuses to overwrite if any existing visit already uses the
    proposed code (would collide)."""
    _deny_engineer()
    if user.get("role") != "admin":
        raise HTTPException(403, "Only admins can edit the visit number series")
    next_seq = int(body.next_seq)
    proposed = f"SV-{next_seq:04d}"
    if await db.site_visits.find_one({"visit_code": proposed}, {"_id": 1}):
        raise HTTPException(
            400,
            f"Cannot set next code to {proposed} — a site visit with this code already exists.",
        )
    await db.counters.update_one(
        {"_id": "site_visit"},
        {"$set": {"seq": next_seq - 1}},
        upsert=True,
    )
    return {"ok": True, "next_seq": next_seq, "next_code": proposed}


@api_router.get("/site-visit-templates", response_model=List[SiteVisitTemplate])
async def list_sv_templates():
    rows = await db.site_visit_templates.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    return rows


@api_router.post("/site-visit-templates", response_model=SiteVisitTemplate)
async def create_sv_template(data: SiteVisitTemplateIn):
    doc = {
        "id": _new_id(),
        "name": data.name.strip(),
        "description": data.description or "",
        "checklist": [s.strip() for s in (data.checklist or []) if s.strip()],
        "created_at": _now(),
    }
    await db.site_visit_templates.insert_one(doc.copy())
    return doc


@api_router.put("/site-visit-templates/{tid}", response_model=SiteVisitTemplate)
async def update_sv_template(tid: str, data: SiteVisitTemplateIn):
    existing = await db.site_visit_templates.find_one({"id": tid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Template not found")
    update = {
        "name": data.name.strip(),
        "description": data.description or "",
        "checklist": [s.strip() for s in (data.checklist or []) if s.strip()],
    }
    await db.site_visit_templates.update_one({"id": tid}, {"$set": update})
    return {**existing, **update}


@api_router.delete("/site-visit-templates/{tid}")
async def delete_sv_template(tid: str):
    res = await db.site_visit_templates.delete_one({"id": tid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Template not found")
    return {"ok": True}


async def _enrich_site_visit(v: dict) -> dict:
    if v.get("project_id"):
        p = await db.projects.find_one({"id": v["project_id"]}, {"_id": 0, "project_code": 1, "name": 1, "site_location": 1, "client_name": 1})
        if p:
            v["project_code"] = p.get("project_code", "")
            v["project_name"] = p.get("name", "")
            # Auto-fill blanks from the linked project so historical visits also benefit
            if not v.get("customer"):
                v["customer"] = p.get("client_name", "")
            if not v.get("site_location"):
                v["site_location"] = p.get("site_location", "")
    return v


async def _log_sv_activity(visit_id: str, visit_code: str, action: str, detail: str = ""):
    """Append a site-visit activity event (parallel to project/audit activity)."""
    try:
        s = _current_user_stamp()
        await db.activity_log.insert_one({
            "id": _new_id(),
            "site_visit_id": visit_id,
            "site_visit_code": visit_code,
            "action": action,
            "detail": detail,
            "user_id": s["user_id"],
            "username": s["username"],
            "created_at": _now(),
        })
    except Exception as e:
        logger.error(f"sv activity log error: {e}")


async def _notify_user(user_id: str, message: str, related_task_id: str = ""):
    """Create an in-app notification targeted at a specific user
    AND fire a Web Push to them if subscribed."""
    try:
        s = _current_user_stamp()
        await db.notifications.insert_one({
            "id": _new_id(),
            "type": "task",
            "message": message,
            "target_role": None,
            "target_user_id": user_id,
            "related_task_id": related_task_id,
            "created_by_user_id": s["user_id"],
            "created_by_username": s["username"],
            "read_by": [],
            "created_at": _now(),
        })
    except Exception as e:
        logger.error(f"notify user error: {e}")

    try:
        await _push_to_user(user_id, {
            "title": "Task Assigned",
            "body": message[:160],
            "url": f"/tasks",
            "tag": f"task-{related_task_id}",
        })
    except Exception as e:
        logger.error(f"push to user error: {e}")


async def _notify_admins(message: str, related_visit_id: str = "", related_visit_code: str = ""):
    """Create an in-app notification targeted at all users with role=admin
    AND fire a Web Push to any admins who've subscribed on this/another device."""
    try:
        s = _current_user_stamp()
        await db.notifications.insert_one({
            "id": _new_id(),
            "type": "site_visit",
            "message": message,
            "target_role": "admin",
            "target_user_id": None,
            "related_visit_id": related_visit_id,
            "related_visit_code": related_visit_code,
            "created_by_user_id": s["user_id"],
            "created_by_username": s["username"],
            "read_by": [],          # list of user_ids who have dismissed it
            "created_at": _now(),
        })
    except Exception as e:
        logger.error(f"notify error: {e}")

    # Fire-and-forget web push (subscribed admin devices)
    try:
        await _push_to_admins({
            "title": "New site visit submitted",
            "body": message[:160],
            "url": f"/site-visits/{related_visit_id}" if related_visit_id else "/site-visits",
            "tag": f"site-visit-{related_visit_id or 'general'}",
        })
    except Exception as e:
        logger.error(f"push to admins error: {e}")


@api_router.get("/site-visits/{vid}/activity")
async def list_sv_activity(vid: str):
    """Note: this declaration sits BEFORE the catch-all /site-visits/{vid} GET (further below)
    so FastAPI matches it first."""
    items = await db.activity_log.find({"site_visit_id": vid}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.post("/site-visits/{vid}/pin")
async def pin_site_visit(vid: str, body: dict = None):
    """Toggle the 'pinned' flag on a site visit. A project can have multiple pinned visits;
    the project page surfaces them as a compact 'Key Inspections' strip at the top."""
    pin_val = bool((body or {}).get("pinned", True))
    existing = await db.site_visits.find_one({"id": vid}, {"_id": 0, "visit_code": 1})
    if not existing:
        raise HTTPException(404, "Site visit not found")
    await db.site_visits.update_one({"id": vid}, {"$set": {"is_pinned": pin_val}})
    await _log_sv_activity(vid, existing.get("visit_code", ""), "PINNED" if pin_val else "UNPINNED", "")
    return {"ok": True, "is_pinned": pin_val}


@api_router.get("/site-visits/export/excel")
async def export_site_visits_excel(
    month: Optional[str] = None,        # YYYY-MM
    engineer_id: Optional[str] = None,
    project_id: Optional[str] = None,
):
    """Excel export of site visits. Optional filters: month=YYYY-MM, engineer_id, project_id.
    Two sheets: 'Visits' (one row per visit) + 'By Engineer' (count + non-compliant items).
    Must be declared BEFORE /site-visits/{vid} GET so the static path takes precedence."""
    q: dict = {}
    if project_id:
        q["project_id"] = project_id
    if engineer_id:
        q["created_by_user_id"] = engineer_id
    # Scope engineer to their own visits
    user = get_current_user_safe()
    if user and user.get("role") in ("engineer", "draftsman"):
        q["created_by_user_id"] = user["id"]
    rows = await db.site_visits.find(q, {"_id": 0, "photos": 0, "engineer_signature": 0, "site_person_signature": 0}).sort("created_at", -1).to_list(5000)
    for r in rows:
        await _enrich_site_visit(r)

    if month:
        rows = [r for r in rows if (r.get("visit_date") or r.get("created_at") or "").startswith(month)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visits"
    headers = ["Code", "Date", "Inspection", "Template", "Project", "Customer", "Site Location", "Job No", "DRG No", "Rev", "Engineer", "Status", "GPS", "Yes", "No", "N/A", "Observations"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2E1F")
        c.alignment = Alignment(horizontal="center", vertical="center")

    per_eng: dict = {}  # engineer_name -> {count, non_compliant, last_date}
    for r in rows:
        cl = r.get("checklist") or []
        y = sum(1 for c in cl if (c.get("compliance") or "").lower() == "yes")
        n = sum(1 for c in cl if (c.get("compliance") or "").lower() == "no")
        na = sum(1 for c in cl if (c.get("compliance") or "").lower() == "na")
        eng = (r.get("engineer_name") or r.get("created_by_username") or "—")
        ws.append([
            r.get("visit_code", ""),
            (r.get("visit_date") or "")[:10],
            r.get("inspection_title", ""),
            r.get("template_name", ""),
            f"{r.get('project_code','')} {r.get('project_name','')}".strip(),
            r.get("customer", ""),
            r.get("site_location", "") or r.get("plot_no", ""),
            r.get("job_no", ""),
            r.get("drg_no", ""),
            r.get("revision", ""),
            eng,
            r.get("status", "submitted"),
            f"{r.get('latitude'):.6f}, {r.get('longitude'):.6f}" if (r.get("latitude") is not None and r.get("longitude") is not None) else "",
            y, n, na,
            "\n".join(r.get("observations") or []),
        ])
        agg = per_eng.setdefault(eng, {"count": 0, "non_compliant": 0, "last": ""})
        agg["count"] += 1
        agg["non_compliant"] += n
        d = (r.get("visit_date") or "")[:10]
        if d and (not agg["last"] or d > agg["last"]):
            agg["last"] = d

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    ws2 = wb.create_sheet("By Engineer")
    ws2.append(["Engineer", "Visits", "Non-compliant items", "Last visit date"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0A2E1F")
    for eng, agg in sorted(per_eng.items(), key=lambda x: -x[1]["count"]):
        ws2.append([eng, agg["count"], agg["non_compliant"], agg["last"]])
    for col_idx in range(1, 5):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"site-visits-{month or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api_router.get("/site-visits", response_model=List[SiteVisit])
async def list_site_visits(project_id: Optional[str] = None, mine: Optional[bool] = False, search: Optional[str] = None):
    q: dict = {}
    if project_id: q["project_id"] = project_id
    if mine:
        user = get_current_user_safe()
        if user: q["created_by_user_id"] = user["id"]
    if search:
        rx = {"$regex": search, "$options": "i"}
        q["$or"] = [{"visit_code": rx}, {"inspection_title": rx}, {"job_no": rx}, {"customer": rx}, {"plot_no": rx}, {"project_code": rx}]
    rows = await db.site_visits.find(q, {"_id": 0, "photos": 0, "engineer_signature": 0, "site_person_signature": 0}).sort("created_at", -1).to_list(2000)
    for r in rows:
        await _enrich_site_visit(r)
    return rows


@api_router.get("/site-visits/{vid}", response_model=SiteVisit)
async def get_site_visit(vid: str):
    v = await db.site_visits.find_one({"id": vid}, {"_id": 0})
    if not v:
        raise HTTPException(404, "Site visit not found")
    # Lazy-mint a public_token for legacy visits created before tokens existed.
    # This token powers the WhatsApp share + mobile-friendly PDF download.
    if not v.get("public_token"):
        tok = secrets.token_urlsafe(20)
        await db.site_visits.update_one({"id": vid}, {"$set": {"public_token": tok}})
        v["public_token"] = tok
    await _enrich_site_visit(v)
    return v


@api_router.post("/site-visits", response_model=SiteVisit)
async def create_site_visit(data: SiteVisitIn):
    user = get_current_user_safe() or {}
    doc = {
        "id": _new_id(),
        "visit_code": await _next_visit_code(),
        "template_id": data.template_id or None,
        "template_name": data.template_name or "",
        "job_no": data.job_no or "",
        "project_id": data.project_id or None,
        "project_code": "",
        "project_name": "",
        "inspection_title": data.inspection_title or "",
        "visit_date": data.visit_date or _now(),
        "customer": data.customer or "",
        "plot_no": data.plot_no or "",
        "site_location": data.site_location or "",
        "drg_no": data.drg_no or "",
        "revision": data.revision or "",
        "latitude": data.latitude,
        "longitude": data.longitude,
        "geo_accuracy": data.geo_accuracy,
        "checklist": [ci.model_dump() for ci in (data.checklist or [])],
        "observations": [o for o in (data.observations or []) if o.strip()],
        "photos": [p.model_dump() for p in (data.photos or [])],
        "engineer_name": data.engineer_name or user.get("username", ""),
        "engineer_signature": data.engineer_signature or "",
        "site_person_name": data.site_person_name or "",
        "site_person_phone": data.site_person_phone or "",
        "site_person_signature": data.site_person_signature or "",
        "status": data.status or "submitted",
        "public_token": secrets.token_urlsafe(20),
        "created_by_user_id": user.get("id"),
        "created_by_username": user.get("username", ""),
        "created_at": _now(),
    }
    _stamp_edit(doc)
    await _enrich_site_visit(doc)
    await db.site_visits.insert_one(doc.copy())
    await _log_sv_activity(doc["id"], doc["visit_code"], "VISIT CREATED", doc.get("inspection_title", ""))
    if doc.get("status") == "submitted" and (user.get("role") or "") != "admin":
        eng_name = user.get("username") or doc.get("engineer_name") or "engineer"
        proj = f" — {doc.get('project_code')}" if doc.get("project_code") else ""
        await _notify_admins(
            f"{eng_name} submitted site visit {doc['visit_code']}{proj}: {doc.get('inspection_title','')}",
            related_visit_id=doc["id"],
            related_visit_code=doc["visit_code"],
        )
    return doc


@api_router.put("/site-visits/{vid}", response_model=SiteVisit)
async def update_site_visit(vid: str, data: SiteVisitIn):
    existing = await db.site_visits.find_one({"id": vid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Site visit not found")
    update = {
        "template_id": data.template_id or existing.get("template_id"),
        "template_name": data.template_name or existing.get("template_name", ""),
        "job_no": data.job_no or "",
        "project_id": data.project_id or None,
        "inspection_title": data.inspection_title or existing.get("inspection_title", ""),
        "visit_date": data.visit_date or existing.get("visit_date"),
        "customer": data.customer or "",
        "plot_no": data.plot_no or "",
        "site_location": data.site_location or existing.get("site_location", ""),
        "drg_no": data.drg_no or "",
        "revision": data.revision or "",
        "latitude": data.latitude if data.latitude is not None else existing.get("latitude"),
        "longitude": data.longitude if data.longitude is not None else existing.get("longitude"),
        "geo_accuracy": data.geo_accuracy if data.geo_accuracy is not None else existing.get("geo_accuracy"),
        "checklist": [ci.model_dump() for ci in (data.checklist or [])],
        "observations": [o for o in (data.observations or []) if o.strip()],
        "photos": [p.model_dump() for p in (data.photos or [])],
        "engineer_name": data.engineer_name or existing.get("engineer_name", ""),
        "engineer_signature": data.engineer_signature or existing.get("engineer_signature", ""),
        "site_person_name": data.site_person_name or existing.get("site_person_name", ""),
        "site_person_phone": data.site_person_phone or existing.get("site_person_phone", ""),
        "site_person_signature": data.site_person_signature or existing.get("site_person_signature", ""),
        "status": data.status or existing.get("status", "submitted"),
    }
    _stamp_edit(update)
    await db.site_visits.update_one({"id": vid}, {"$set": update})
    merged = {**existing, **update}
    await _enrich_site_visit(merged)
    # Activity: distinguish a "status change" from a plain edit
    user = get_current_user_safe() or {}
    prev_status = existing.get("status", "submitted")
    new_status = update.get("status", prev_status)
    if prev_status != new_status:
        await _log_sv_activity(vid, merged.get("visit_code", ""), "STATUS CHANGED", f"{prev_status} → {new_status}")
        if new_status == "submitted" and (user.get("role") or "") != "admin":
            eng_name = user.get("username") or merged.get("engineer_name") or "engineer"
            proj = f" — {merged.get('project_code')}" if merged.get("project_code") else ""
            await _notify_admins(
                f"{eng_name} submitted site visit {merged.get('visit_code','')}{proj}: {merged.get('inspection_title','')}",
                related_visit_id=vid,
                related_visit_code=merged.get("visit_code", ""),
            )
    else:
        await _log_sv_activity(vid, merged.get("visit_code", ""), "VISIT UPDATED", "")
    return merged


@api_router.delete("/site-visits/{vid}")
async def delete_site_visit(vid: str):
    existing = await db.site_visits.find_one({"id": vid}, {"_id": 0, "visit_code": 1})
    if not existing:
        raise HTTPException(404, "Site visit not found")
    # Log activity BEFORE the delete so the audit trail mirrors what happened
    await _log_sv_activity(vid, existing.get("visit_code", ""), "VISIT DELETED", "")
    # Cascade: remove any notifications referencing this visit so admins
    # don't get dead links in their feed
    try:
        await db.notifications.delete_many({"related_visit_id": vid})
    except Exception:
        pass
    await db.site_visits.delete_one({"id": vid})
    return {"ok": True}


def _base64_image_from_data_url(data_url: str) -> Optional[io.BytesIO]:
    if not data_url or "," not in data_url:
        return None
    try:
        b64 = data_url.split(",", 1)[1]
        return io.BytesIO(base64.b64decode(b64))
    except Exception:
        return None


# Site-visit upload storage (mounted at /api/uploads)
UPLOAD_ROOT = Path(os.environ.get("UPLOAD_DIR", str(ROOT_DIR / "uploads")))
SITE_VISIT_UPLOAD_DIR = UPLOAD_ROOT / "site-visits"
SITE_VISIT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

_photo_bucket = None

# ---------- Letterhead overlay (Creator Consultant branded PDF background) ----------
# The PDF at /app/backend/assets/letterhead.pdf is rendered as a background on
# every site-visit report page. Drawn body content sits on top.
LETTERHEAD_PATH = ROOT_DIR / "assets" / "letterhead.pdf"
_letterhead_reader: Optional[pypdf.PdfReader] = None
try:
    if LETTERHEAD_PATH.exists():
        _letterhead_reader = pypdf.PdfReader(str(LETTERHEAD_PATH))
        print(f"[startup] Letterhead loaded ({len(_letterhead_reader.pages)} page) from {LETTERHEAD_PATH}")
except Exception as _e:
    print(f"[startup] Could not load letterhead PDF: {_e}")
    _letterhead_reader = None


def _apply_letterhead(pdf_bytes: bytes) -> bytes:
    """Stamp the Creator Consultant letterhead UNDER every page of `pdf_bytes`.
    If the letterhead is missing or merging fails, returns the input unchanged.
    The letterhead page is scaled to match each content page's media box, so
    Letter-size letterhead works under A4 content (and vice versa).

    NOTE: We re-open the letterhead PDF fresh on every call and use
    `add_blank_page` + `merge_transformed_page` + `merge_page` instead of
    `deepcopy(template)`. PyPDF PageObjects share IndirectObject references
    with their parent reader; `copy.deepcopy` does NOT fully decouple them,
    so reusing a deepcopy across iterations causes content streams to bleed
    together into a single page in the output (every page's content stacked
    on top of each other under one letterhead).
    """
    if not LETTERHEAD_PATH.exists():
        return pdf_bytes
    try:
        from pypdf import PdfReader, PdfWriter, Transformation
        src = PdfReader(io.BytesIO(pdf_bytes))
        lh_reader = PdfReader(str(LETTERHEAD_PATH))
        if not lh_reader.pages:
            return pdf_bytes
        lh_template = lh_reader.pages[0]
        lh_w = float(lh_template.mediabox.width)
        lh_h = float(lh_template.mediabox.height)
        writer = PdfWriter()
        for page in src.pages:
            pw = float(page.mediabox.width)
            ph = float(page.mediabox.height)
            # Fresh, independent page for this iteration
            new_page = writer.add_blank_page(width=pw, height=ph)
            # Layer 1 — letterhead at the bottom (scaled to fit)
            sx = pw / lh_w
            sy = ph / lh_h
            new_page.merge_transformed_page(
                lh_template, Transformation().scale(sx=sx, sy=sy),
            )
            # Layer 2 — site visit content on top
            new_page.merge_page(page)
        out = io.BytesIO()
        writer.write(out)
        return out.getvalue()
    except Exception as e:
        print(f"[pdf] Letterhead overlay failed, returning unstamped PDF: {e}")
        return pdf_bytes


def _photo_to_image_reader(p: dict) -> Optional[ImageReader]:
    """Resolve a SiteVisitPhoto dict to a reportlab ImageReader.
    Supports base64 data_url, on-disk legacy file, or GridFS-stored bytes."""
    if p.get("data_url"):
        bio = _base64_image_from_data_url(p["data_url"])
        if bio:
            try:
                return ImageReader(bio)
            except Exception:
                return None
    url = p.get("url") or ""
    if url:
        # url is /api/uploads/site-visits/<fname>. Try disk first (legacy),
        # then fall back to GridFS using the filename as the GridFS filename.
        fname = url.rsplit("/", 1)[-1]
        fpath = SITE_VISIT_UPLOAD_DIR / fname
        if fpath.exists():
            try:
                return ImageReader(str(fpath))
            except Exception:
                pass
        # GridFS fallback — we have to read synchronously for ReportLab so we
        # spawn a small async helper from the surrounding (already-async) PDF
        # builder. This branch is exercised by `_render_visit_pdf` which calls
        # `await _load_photo_bytes(url)` upstream and stuffs the bytes back
        # into the photo dict before calling us. So here we just bail.
    return None


async def _load_photo_bytes(url: str) -> Optional[bytes]:
    """Download a site-visit photo's raw bytes for the PDF builder. Tries disk
    first (legacy files), then GridFS by filename."""
    if not url:
        return None
    fname = url.rsplit("/", 1)[-1]
    fpath = SITE_VISIT_UPLOAD_DIR / fname
    if fpath.exists():
        try:
            return fpath.read_bytes()
        except Exception:
            pass
    try:
        stream = await _photo_bucket.open_download_stream_by_name(fname)
        try:
            return await stream.read()
        finally:
            try:
                stream.close()
            except Exception:
                pass
    except Exception:
        return None


@api_router.post("/site-visits/uploads")
async def upload_site_visit_photo(file: UploadFile = File(...)):
    """Accept a single image file from the engineer's phone, store the bytes in
    MongoDB GridFS (so they survive container redeploys on Kubernetes), and
    return the canonical URL `/api/uploads/site-visits/<filename>` that the
    frontend can later <img src=...>."""
    if not file.filename:
        raise HTTPException(400, "Missing filename")
    ext = (file.filename.rsplit(".", 1)[-1] or "jpg").lower()
    if ext not in {"jpg", "jpeg", "png", "webp", "heic", "heif", "gif"}:
        raise HTTPException(400, "Only image files are allowed")
    fname = f"{secrets.token_urlsafe(16)}.{ext}"
    content_type = file.content_type or f"image/{ 'jpeg' if ext == 'jpg' else ext }"
    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty upload")
    # Persist in GridFS keyed by filename. We always write a fresh doc per
    # upload so retries don't clobber each other.
    await _photo_bucket.upload_from_stream(
        fname, io.BytesIO(data), metadata={"content_type": content_type},
    )
    return {"url": f"/api/uploads/site-visits/{fname}", "filename": fname}


@auth_public_router.get("/uploads/site-visits/{filename}")
async def serve_site_visit_photo(filename: str):
    """Stream a site-visit photo back to the browser. Looks in GridFS first
    (where new uploads live), then falls back to the legacy on-disk path."""
    # GridFS lookup by filename — `open_download_stream_by_name` always picks
    # the most recently uploaded revision if duplicates exist.
    try:
        stream = await _photo_bucket.open_download_stream_by_name(filename)
        ct = (getattr(stream, "metadata", None) or {}).get("content_type") or "image/jpeg"

        async def gen():
            try:
                while True:
                    chunk = await stream.readchunk()
                    if not chunk:
                        break
                    yield chunk
            finally:
                try:
                    stream.close()
                except Exception:
                    pass

        return StreamingResponse(
            gen(),
            media_type=ct,
            headers={"Cache-Control": "public, max-age=31536000, immutable"},
        )
    except Exception:
        pass
    # Legacy: file may still exist on disk (older preview uploads)
    fpath = SITE_VISIT_UPLOAD_DIR / filename
    if fpath.exists() and fpath.is_file():
        ext = filename.rsplit(".", 1)[-1].lower()
        ct = f"image/{ 'jpeg' if ext == 'jpg' else ext }"
        return StreamingResponse(
            iter([fpath.read_bytes()]),
            media_type=ct,
            headers={"Cache-Control": "public, max-age=31536000, immutable"},
        )
    raise HTTPException(404, "Photo not found")


@api_router.delete("/site-visits/uploads/{filename}")
async def delete_site_visit_upload(filename: str):
    # Disk (legacy)
    fpath = SITE_VISIT_UPLOAD_DIR / filename
    if fpath.exists() and fpath.is_file():
        try:
            fpath.unlink()
        except Exception:
            pass
    # GridFS — delete all matching docs
    try:
        async for grid_doc in _photo_bucket.find({"filename": filename}):
            try:
                await _photo_bucket.delete(grid_doc._id)
            except Exception:
                pass
    except Exception:
        pass
    return {"ok": True}


# ---------------------- NOTIFICATIONS (in-app feed) ----------------------

def _user_can_see_notification(user: dict, n: dict) -> bool:
    if not user:
        return False
    if n.get("target_user_id") and n["target_user_id"] == user.get("id"):
        return True
    if n.get("target_role") and n["target_role"] == user.get("role"):
        return True
    return False


@api_router.get("/notifications")
async def list_notifications(limit: int = 25, unread_only: bool = False):
    user = get_current_user_safe() or {}
    # Pull anything targeted at this user OR their role; sort newest first
    q = {"$or": [
        {"target_user_id": user.get("id")},
        {"target_role": user.get("role")},
    ]}
    rows = await db.notifications.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit * 4)
    out = []
    for n in rows:
        if not _user_can_see_notification(user, n):
            continue
        n["is_read"] = user.get("id") in (n.get("read_by") or [])
        if unread_only and n["is_read"]:
            continue
        out.append(n)
        if len(out) >= limit:
            break
    unread = sum(1 for n in out if not n["is_read"])
    return {"items": out, "unread": unread}


@api_router.post("/notifications/{nid}/read")
async def mark_notification_read(nid: str):
    user = get_current_user_safe() or {}
    if not user.get("id"):
        raise HTTPException(401, "Not authenticated")
    await db.notifications.update_one(
        {"id": nid},
        {"$addToSet": {"read_by": user["id"]}},
    )
    return {"ok": True}


@api_router.post("/notifications/read-all")
async def mark_all_notifications_read():
    user = get_current_user_safe() or {}
    if not user.get("id"):
        raise HTTPException(401, "Not authenticated")
    q = {"$or": [
        {"target_user_id": user.get("id")},
        {"target_role": user.get("role")},
    ]}
    await db.notifications.update_many(q, {"$addToSet": {"read_by": user["id"]}})
    return {"ok": True}


# ---------------------- WEB PUSH (VAPID + service worker) ----------------------

_VAPID_CACHE: dict = {"public": "", "private_pem": "", "private_raw_b64": "", "claims_email": "mailto:admin@creatorconsultant.online"}


def _vapid_generate_keypair() -> tuple[str, str, str]:
    """Generate a fresh VAPID P-256 keypair. Returns (public_b64url, private_pem, private_raw_b64url).
    pywebpush 2.x wants the *raw 32-byte private* in base64url (not the PEM), so we cache both."""
    from cryptography.hazmat.primitives.asymmetric import ec
    from cryptography.hazmat.primitives import serialization
    key = ec.generate_private_key(ec.SECP256R1())
    private_pem = key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.TraditionalOpenSSL,
        encryption_algorithm=serialization.NoEncryption(),
    ).decode()
    nums = key.public_key().public_numbers()
    raw_pub = b"\x04" + nums.x.to_bytes(32, "big") + nums.y.to_bytes(32, "big")
    public_b64 = base64.urlsafe_b64encode(raw_pub).decode().rstrip("=")
    raw_priv = key.private_numbers().private_value.to_bytes(32, "big")
    private_raw_b64 = base64.urlsafe_b64encode(raw_priv).decode().rstrip("=")
    return public_b64, private_pem, private_raw_b64


async def _ensure_vapid_keys():
    """Persist VAPID keys in app_settings; auto-generate once on first run."""
    doc = await db.app_settings.find_one({"id": "vapid"}, {"_id": 0})
    if doc and doc.get("public") and doc.get("private_pem") and doc.get("private_raw_b64"):
        _VAPID_CACHE.update(doc)
        logger.info("VAPID keys loaded from db")
        return
    public, private_pem, private_raw_b64 = _vapid_generate_keypair()
    doc = {
        "id": "vapid",
        "public": public,
        "private_pem": private_pem,
        "private_raw_b64": private_raw_b64,
        "claims_email": "mailto:admin@creatorconsultant.online",
        "created_at": _now(),
    }
    await db.app_settings.update_one({"id": "vapid"}, {"$set": doc}, upsert=True)
    _VAPID_CACHE.update(doc)
    logger.info("VAPID keypair generated and stored")


async def _send_web_push(subscription: dict, payload: dict) -> bool:
    """Send a single web push. Returns True on success, False otherwise.
    Deletes the subscription row if the endpoint returns 404/410 (subscription gone)."""
    priv = _VAPID_CACHE.get("private_raw_b64")
    if not priv:
        return False
    from pywebpush import webpush, WebPushException
    import json as _json
    try:
        webpush(
            subscription_info={
                "endpoint": subscription["endpoint"],
                "keys": subscription["keys"],
            },
            data=_json.dumps(payload),
            vapid_private_key=priv,
            vapid_claims={"sub": _VAPID_CACHE.get("claims_email", "mailto:admin@example.com")},
            ttl=86400,
        )
        return True
    except WebPushException as e:
        code = getattr(getattr(e, "response", None), "status_code", None)
        if code in (404, 410):
            try:
                await db.push_subscriptions.delete_one({"endpoint": subscription["endpoint"]})
                logger.info(f"Pruned dead push subscription (HTTP {code}): {subscription['endpoint'][:60]}…")
            except Exception as ce:
                logger.error(f"failed to prune dead sub: {ce}")
            return False
        logger.error(f"web push send failed (HTTP {code}): {e}")
        return False
    except Exception as e:
        logger.error(f"web push send failed: {e}")
        return False


async def _push_to_admins(payload: dict):
    """Broadcast a Web Push payload to every admin's active subscription(s)."""
    admin_ids = [u["id"] async for u in db.users.find({"role": "admin"}, {"_id": 0, "id": 1})]
    if not admin_ids:
        return 0
    subs = await db.push_subscriptions.find({"user_id": {"$in": admin_ids}}, {"_id": 0}).to_list(500)
    sent = 0
    for sub in subs:
        ok = await _send_web_push(sub, payload)
        if ok:
            sent += 1
    return sent

async def _push_to_user(user_id: str, payload: dict):
    """Send a Web Push payload to a specific user's active subscription(s)."""
    subs = await db.push_subscriptions.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    sent = 0
    for sub in subs:
        if await _send_web_push(sub, payload):
            sent += 1
    return sent


@api_router.get("/push/vapid-public")
async def get_vapid_public():
    if not _VAPID_CACHE.get("public"):
        await _ensure_vapid_keys()
    return {"public_key": _VAPID_CACHE.get("public", "")}


class PushSubscriptionIn(BaseModel):
    endpoint: str
    keys: dict  # {p256dh, auth}
    expirationTime: Optional[int] = None


@api_router.post("/push/subscribe")
async def push_subscribe(body: PushSubscriptionIn):
    user = get_current_user_safe() or {}
    if not user.get("id"):
        raise HTTPException(401, "Not authenticated")
    doc = {
        "id": _new_id(),
        "user_id": user["id"],
        "username": user.get("username", ""),
        "role": user.get("role", ""),
        "endpoint": body.endpoint,
        "keys": body.keys,
        "expiration_time": body.expirationTime,
        "user_agent": "",
        "created_at": _now(),
    }
    # Upsert by endpoint so the same browser doesn't pile up duplicates
    await db.push_subscriptions.update_one(
        {"endpoint": body.endpoint},
        {"$set": doc},
        upsert=True,
    )
    return {"ok": True}


@api_router.post("/push/unsubscribe")
async def push_unsubscribe(body: dict):
    endpoint = body.get("endpoint")
    if not endpoint:
        raise HTTPException(400, "Missing endpoint")
    await db.push_subscriptions.delete_one({"endpoint": endpoint})
    return {"ok": True}


@api_router.get("/push/status")
async def push_status():
    """Returns whether the current user has any active subscription on this account."""
    user = get_current_user_safe() or {}
    if not user.get("id"):
        raise HTTPException(401, "Not authenticated")
    count = await db.push_subscriptions.count_documents({"user_id": user["id"]})
    return {"subscribed": count > 0, "count": count}


@api_router.post("/push/test")
async def push_test():
    user = get_current_user_safe() or {}
    if not user.get("id"):
        raise HTTPException(401, "Not authenticated")
    subs = await db.push_subscriptions.find({"user_id": user["id"]}, {"_id": 0}).to_list(50)
    if not subs:
        raise HTTPException(400, "You have no active push subscriptions on this device")
    sent = 0
    for sub in subs:
        ok = await _send_web_push(sub, {
            "title": "Creator Consultant",
            "body": "Test push notification — you're all set!",
            "url": "/site-visits",
        })
        if ok:
            sent += 1
    return {"ok": True, "sent": sent, "total": len(subs)}


# ---------------------- HOUSEKEEPING (daily) ----------------------

NOTIFICATION_TTL_DAYS = 30
_housekeeping_scheduler = None


async def _cleanup_old_read_notifications() -> int:
    """Daily job: drop in-app notifications that are older than 30 days AND have been read
    by at least one user. We keep brand-new unread items even if old so nothing slips through."""
    cutoff = (datetime.now(timezone.utc) - timedelta(days=NOTIFICATION_TTL_DAYS)).isoformat()
    res = await db.notifications.delete_many({
        "created_at": {"$lt": cutoff},
        "read_by.0": {"$exists": True},   # has at least one reader
    })
    if res.deleted_count:
        logger.info(f"Housekeeping: pruned {res.deleted_count} read notifications older than {NOTIFICATION_TTL_DAYS}d")
    return res.deleted_count


@api_router.post("/notifications/cleanup")
async def notifications_cleanup_now():
    """Manual trigger for the housekeeping job (admin convenience)."""
    user = get_current_user_safe() or {}
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    deleted = await _cleanup_old_read_notifications()
    return {"ok": True, "deleted": deleted}


async def _send_task_due_reminders() -> int:
    """Daily job: Send reminders for upcoming task due dates."""
    today = datetime.now(timezone.utc).date()
    sent_count = 0
    
    # Only look at incomplete tasks that have a due_date and an assigned user
    q = {
        "status": "pending",
        "due_date": {"$exists": True, "$ne": None, "$ne": ""},
        "assigned_to_user_id": {"$exists": True, "$ne": None, "$ne": ""}
    }
    
    async for task in db.tasks.find(q):
        try:
            due_date = datetime.fromisoformat(task["due_date"].split("T")[0]).date()
        except Exception:
            continue
            
        start_date = None
        if task.get("start_date"):
            try:
                start_date = datetime.fromisoformat(task["start_date"].split("T")[0]).date()
            except Exception:
                pass
                
        # Calculate timeline
        if start_date:
            timeline_days = (due_date - start_date).days
        else:
            timeline_days = 3 # treat as > 2 days
            
        days_until_due = (due_date - today).days
        
        should_send = False
        message = ""
        
        if timeline_days <= 2:
            # Short timeline: only send on due date
            if days_until_due == 0:
                should_send = True
                message = f"Reminder: Task '{task.get('work', 'Untitled')}' is due today!"
        else:
            # Long timeline: send on due_date - 1 and due_date
            if days_until_due == 1:
                should_send = True
                message = f"Reminder: Task '{task.get('work', 'Untitled')}' is due tomorrow!"
            elif days_until_due == 0:
                should_send = True
                message = f"Reminder: Task '{task.get('work', 'Untitled')}' is due today!"
                
        if should_send:
            doc = {
                "id": _new_id(),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "title": "Task Due Reminder",
                "message": message,
                "target_user_id": task["assigned_to_user_id"],
                "target_role": None,
                "related_entity_id": task["id"],
                "read_by": []
            }
            await db.notifications.insert_one(doc)
            
            # Send push notification to the assigned user
            await _push_to_user(task["assigned_to_user_id"], {
                "title": "Task Due Reminder",
                "body": message,
                "url": "/tasks"
            })
            
            sent_count += 1
            
    if sent_count > 0:
        logger.info(f"Housekeeping: sent {sent_count} task due reminders")
    return sent_count


@api_router.post("/notifications/trigger-task-reminders")
async def trigger_task_reminders_now():
    """Manual trigger for task reminders."""
    user = get_current_user_safe() or {}
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    sent = await _send_task_due_reminders()
    return {"ok": True, "sent": sent}


def _start_housekeeping_scheduler():
    """Run the cleanup once a day at 03:15 UTC."""
    global _housekeeping_scheduler
    if _housekeeping_scheduler is not None:
        return
    from apscheduler.schedulers.asyncio import AsyncIOScheduler
    from apscheduler.triggers.cron import CronTrigger
    sched = AsyncIOScheduler(timezone="UTC")
    sched.add_job(_cleanup_old_read_notifications, CronTrigger(hour=3, minute=15), id="cleanup_old_notifications", replace_existing=True)
    sched.add_job(_send_task_due_reminders, CronTrigger(hour=8, minute=0), id="task_due_reminders", replace_existing=True)
    sched.start()
    _housekeeping_scheduler = sched
    logger.info("Housekeeping scheduler started — daily cleanup at 03:15 UTC, reminders at 08:00 UTC")


def _stop_housekeeping_scheduler():
    global _housekeeping_scheduler
    if _housekeeping_scheduler is not None:
        _housekeeping_scheduler.shutdown(wait=False)
        _housekeeping_scheduler = None


@api_router.get("/site-visits/{vid}/pdf")
async def site_visit_pdf(vid: str):
    v = await db.site_visits.find_one({"id": vid}, {"_id": 0})
    if not v:
        raise HTTPException(404, "Site visit not found")
    await _enrich_site_visit(v)
    await _preload_visit_photo_bytes(v)
    return _render_site_visit_pdf_response(v)


# Public (no-auth) PDF endpoint — shared via WhatsApp link
@auth_public_router.get("/site-visits/public/{token}/pdf")
async def site_visit_public_pdf(token: str):
    v = await db.site_visits.find_one({"public_token": token}, {"_id": 0})
    if not v:
        raise HTTPException(404, "Site visit not found or link expired")
    await _enrich_site_visit(v)
    await _preload_visit_photo_bytes(v)
    return _render_site_visit_pdf_response(v)


async def _preload_visit_photo_bytes(v: dict) -> None:
    """Mutate v.photos[i] to add `_raw_bytes` for photos that only have a URL.
    This is needed because the PDF renderer is sync but photos now live in
    GridFS (async-only access)."""
    for p in (v.get("photos") or []):
        if p.get("data_url") or p.get("_raw_bytes"):
            continue
        if p.get("url"):
            try:
                p["_raw_bytes"] = await _load_photo_bytes(p["url"])
            except Exception:
                p["_raw_bytes"] = None


def _render_site_visit_pdf_response(v: dict) -> StreamingResponse:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 18 * mm
    # The Creator Consultant letterhead reserves the top ~5mm (clean white) and
    # the bottom ~32mm (green band + contact details). We keep our body well
    # inside those zones so nothing visually clashes.
    LH_TOP_RESERVE = 22 * mm   # space below top of page where our title sits
    LH_BOTTOM_RESERVE = 34 * mm  # space above the letterhead footer band
    page_bottom_limit = LH_BOTTOM_RESERVE

    def header():
        # Slim title row — no full-width green band (the letterhead already
        # carries the brand). We just print the report title in dark green
        # and the visit code in accent green on the right.
        c.setFillColor(colors.HexColor("#0A2E1F"))
        c.setFont("Roboto-Bold", 14)
        c.drawString(margin, height - LH_TOP_RESERVE, "SITE VISIT REPORT")
        c.setFillColor(colors.HexColor("#10B981"))
        c.setFont("Roboto-Bold", 12)
        c.drawRightString(width - margin, height - LH_TOP_RESERVE, v.get("visit_code", ""))
        # Thin underline
        c.setStrokeColor(colors.HexColor("#10B981"))
        c.setLineWidth(0.8)
        c.line(margin, height - LH_TOP_RESERVE - 2 * mm, width - margin, height - LH_TOP_RESERVE - 2 * mm)
        c.setFillColor(colors.black)
        c.setStrokeColor(colors.black)

    header()
    y = height - LH_TOP_RESERVE - 10 * mm
    c.setFont("Roboto", 8.5)
    meta_rows = [
        ("Job No", v.get("job_no") or "—", "Date", (v.get("visit_date") or "")[:10]),
        ("Customer", v.get("customer") or "—", "Site Location", v.get("site_location") or v.get("plot_no") or "—"),
        ("DRG No", v.get("drg_no") or "—", "Revision", v.get("revision") or "—"),
    ]
    # Add GPS row if captured
    if v.get("latitude") is not None and v.get("longitude") is not None:
        meta_rows.append((
            "GPS",
            f"{v.get('latitude'):.6f}, {v.get('longitude'):.6f}",
            "Accuracy",
            f"±{v.get('geo_accuracy'):.0f} m" if v.get('geo_accuracy') is not None else "—",
        ))
    for r in meta_rows:
        c.setFont("Roboto-Bold", 9); c.drawString(margin, y, r[0] + ":")
        c.setFont("Roboto", 8.5); c.drawString(margin + 22 * mm, y, str(r[1]))
        c.setFont("Roboto-Bold", 9); c.drawString(width / 2 + 5 * mm, y, r[2] + ":")
        c.setFont("Roboto", 8.5); c.drawString(width / 2 + 25 * mm, y, str(r[3]))
        y -= 6 * mm

    y -= 2 * mm
    c.setFillColor(colors.HexColor("#0A2E1F"))
    c.setFont("Roboto-Bold", 12)
    c.drawString(margin, y, f"Inspection of {v.get('inspection_title', '')}")
    c.setFillColor(colors.black); y -= 8 * mm

    styles = getSampleStyleSheet()
    body = ParagraphStyle("body", parent=styles["Normal"], fontName="Roboto", fontSize=9, leading=12)
    data = [["Description", "Compliance", "Remark"]]
    for ci in (v.get("checklist") or []):
        comp = (ci.get("compliance") or "").upper()
        data.append([Paragraph(ci.get("label", ""), body), comp, Paragraph(ci.get("remark") or "—", body)])
    if len(data) > 1:
        tbl = Table(data, colWidths=[width - margin * 2 - 60 * mm, 28 * mm, 32 * mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A2E1F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Roboto-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        tw, th = tbl.wrapOn(c, width - margin * 2, height)
        tbl.drawOn(c, margin, y - th)
        y -= th + 6 * mm

    if v.get("observations"):
        if y < page_bottom_limit + 30 * mm:
            c.showPage(); header(); y = height - LH_TOP_RESERVE - 10 * mm
        c.setFont("Roboto-Bold", 12); c.setFillColor(colors.HexColor("#0A2E1F"))
        c.drawString(margin, y, "Observations:"); c.setFillColor(colors.black)
        y -= 6 * mm
        for i, obs in enumerate(v.get("observations") or [], 1):
            para = Paragraph(f"{i}. {obs}", body)
            w, h = para.wrap(width - margin * 2 - 6 * mm, 100 * mm)
            para.drawOn(c, margin + 4 * mm, y - h + 9)
            y -= h + 2 * mm
            if y < page_bottom_limit + 10 * mm:
                c.showPage(); header(); y = height - LH_TOP_RESERVE - 10 * mm

    if y < page_bottom_limit + 40 * mm:
        c.showPage(); header(); y = height - LH_TOP_RESERVE - 10 * mm
    y -= 8 * mm
    sig_w = 70 * mm; sig_h = 18 * mm
    for label, name_key, sig_key, x_off, phone_key in [
        ("Structural Engineer", "engineer_name", "engineer_signature", margin, None),
        ("Site Person", "site_person_name", "site_person_signature", margin + (width - margin * 2) / 2 + 5 * mm, "site_person_phone"),
    ]:
        c.setFont("Roboto-Bold", 9); c.setFillColor(colors.black)
        c.drawString(x_off, y, label + ":")
        sig_io = _base64_image_from_data_url(v.get(sig_key, ""))
        if sig_io:
            try:
                c.drawImage(ImageReader(sig_io), x_off, y - sig_h - 2 * mm, width=sig_w, height=sig_h, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        c.setFont("Roboto", 8.5)
        c.drawString(x_off, y - sig_h - 6 * mm, f"Name: {v.get(name_key) or '—'}")
        if phone_key and v.get(phone_key):
            c.drawString(x_off, y - sig_h - 10 * mm, f"Phone: {v.get(phone_key)}")
    y -= sig_h + 12 * mm

    photos = v.get("photos") or []
    if photos:
        c.showPage(); header(); y = height - LH_TOP_RESERVE - 10 * mm
        c.setFont("Roboto-Bold", 12); c.setFillColor(colors.HexColor("#0A2E1F"))
        c.drawString(margin, y, "Site Visit Images & Remarks:"); c.setFillColor(colors.black)
        y -= 8 * mm
        img_w = (width - margin * 2 - 6 * mm) / 2; img_h = 55 * mm
        col = 0
        for p in photos:
            img_reader = _photo_to_image_reader(p)
            if not img_reader and p.get("_raw_bytes"):
                try:
                    img_reader = ImageReader(io.BytesIO(p["_raw_bytes"]))
                except Exception:
                    img_reader = None
            if not img_reader:
                continue
            x = margin + col * (img_w + 6 * mm)
            try:
                c.drawImage(img_reader, x, y - img_h, width=img_w, height=img_h, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
            cap = p.get("caption") or ""
            if cap:
                c.setFont("Roboto", 8)
                c.drawString(x, y - img_h - 4 * mm, cap[:80])
            col += 1
            if col >= 2:
                col = 0
                y -= img_h + 12 * mm
                if y < page_bottom_limit + img_h + 8 * mm:
                    c.showPage(); header(); y = height - LH_TOP_RESERVE - 10 * mm

    c.showPage(); c.save()
    pdf_bytes = buf.getvalue(); buf.close()
    # Stamp the Creator Consultant letterhead under every page
    pdf_bytes = _apply_letterhead(pdf_bytes)
    fname = f"{v.get('visit_code', 'site_visit')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------- TASKS ----------------------

class TaskIn(BaseModel):
    category: str                        # "engineering" or "accounting"
    project_id: Optional[str] = None     # Reference to project for engineering/accounting
    audit_id: Optional[str] = None       # Reference to audit for accounting
    site_location: Optional[str] = ""    # Pre-filled from project, editable
    work: str                            # Required short work summary
    description: Optional[str] = ""      # extra notes / description
    start_date: Optional[str] = None
    due_date: Optional[str] = None
    assigned_to_user_id: Optional[str] = None


class TaskOut(BaseModel):
    id: str
    sr_no: int
    category: str
    project_id: Optional[str] = None
    project_code: str = ""
    audit_id: Optional[str] = None
    audit_code: str = ""
    site_location: Optional[str] = ""
    work: str
    description: str = ""
    start_date: Optional[str] = None
    due_date: Optional[str] = None
    status: str = "pending"
    assigned_to_user_id: Optional[str] = None
    assigned_to_name: str = ""
    assigned_to_username: str = ""
    assigned_to_color: str = ""
    created_by_user_id: Optional[str] = None
    created_by_username: str = ""
    created_at: str


class PaginatedTasks(BaseModel):
    data: List[TaskOut]
    total: int
    total_pending: int = 0
    total_in_progress: int = 0
    total_done: int = 0
    total_cancelled: int = 0

async def _enrich_task(t: dict):
    if t.get("assigned_to_user_id"):
        u = await db.users.find_one({"id": t["assigned_to_user_id"]}, {"_id": 0, "name": 1, "username": 1, "color": 1})
        if u:
            t["assigned_to_name"] = u.get("name") or ""
            t["assigned_to_username"] = u.get("username") or ""
            t["assigned_to_color"] = u.get("color") or ""
            
    # Enrich audit_code if missing
    if t.get("audit_id") and not t.get("audit_code"):
        a = await db.audits.find_one({"id": t["audit_id"]})
        if a:
            t["audit_code"] = a.get("audit_code") or ""

    # Enrich project_code if missing
    if t.get("project_id") and not t.get("project_code"):
        p = await db.projects.find_one({"id": t["project_id"]}, {"_id": 0, "project_code": 1, "job_no": 1})
        if p:
            t["project_code"] = p.get("job_no") or p.get("project_code") or ""


@api_router.get("/tasks")
async def list_tasks(category: Optional[str] = None, project_id: Optional[str] = None):
    user = get_current_user_safe()
    query: dict = {}
    if user and user.get("role") != "admin":
        query["assigned_to_user_id"] = user["id"]
    elif category == "engineering":
        eng_users = await db.users.find({"role": {"$in": ["admin", "engineer", "draftsman"]}}, {"id": 1}).to_list(None)
        query["assigned_to_user_id"] = {"$in": [u["id"] for u in eng_users]}
    elif category == "accounting":
        acc_users = await db.users.find({"role": {"$in": ["admin", "accountant"]}}, {"id": 1}).to_list(None)
        query["assigned_to_user_id"] = {"$in": [u["id"] for u in acc_users]}
        
    if category:
        query["category"] = category
    if project_id:
        query["project_id"] = project_id
    items = await db.tasks.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    for idx, t in enumerate(items):
        t["sr_no"] = idx + 1
        await _enrich_task(t)
    return items


@api_router.get("/tasks/paginated", response_model=PaginatedTasks)
async def list_tasks_paginated(
    page: int = 1, 
    limit: int = 25, 
    q: Optional[str] = None,
    category: Optional[str] = None
):
    skip = (page - 1) * limit
    user = get_current_user_safe()
    query: dict = {}
    if user and user.get("role") != "admin":
        query["assigned_to_user_id"] = user["id"]
    elif category == "engineering":
        eng_users = await db.users.find({"role": {"$in": ["admin", "engineer", "draftsman"]}}, {"id": 1}).to_list(None)
        query["assigned_to_user_id"] = {"$in": [u["id"] for u in eng_users]}
    elif category == "accounting":
        acc_users = await db.users.find({"role": {"$in": ["admin", "accountant"]}}, {"id": 1}).to_list(None)
        query["assigned_to_user_id"] = {"$in": [u["id"] for u in acc_users]}
        
    if category:
        query["category"] = category
        
    if q:
        s = q.strip()
        query["$or"] = [
            {"work": {"$regex": s, "$options": "i"}},
            {"description": {"$regex": s, "$options": "i"}},
            {"project_code": {"$regex": s, "$options": "i"}},
            {"audit_code": {"$regex": s, "$options": "i"}},
            {"site_location": {"$regex": s, "$options": "i"}},
            {"project_name": {"$regex": s, "$options": "i"}},
            {"assigned_to_username": {"$regex": s, "$options": "i"}},
            {"client_name": {"$regex": s, "$options": "i"}}
        ]
        
    total = await db.tasks.count_documents(query)
    
    pending_query = query.copy()
    pending_query["status"] = {"$in": ["pending", None, ""]}
    total_pending = await db.tasks.count_documents(pending_query)

    in_progress_query = query.copy()
    in_progress_query["status"] = "in progress"
    total_in_progress = await db.tasks.count_documents(in_progress_query)
    
    done_query = query.copy()
    done_query["status"] = "done"
    total_done = await db.tasks.count_documents(done_query)

    cancelled_query = query.copy()
    cancelled_query["status"] = "cancelled"
    total_cancelled = await db.tasks.count_documents(cancelled_query)

    items = await db.tasks.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Needs to match sr_no for current page
    for idx, t in enumerate(items):
        t["sr_no"] = skip + idx + 1
        await _enrich_task(t)
        
    return {"data": items, "total": total, "total_pending": total_pending, "total_in_progress": total_in_progress, "total_done": total_done, "total_cancelled": total_cancelled}

@api_router.post("/tasks")
async def create_task(data: TaskIn):
    user = get_current_user_safe()
    if data.category not in ("engineering", "accounting"):
        raise HTTPException(400, "category must be 'engineering' or 'accounting'")
    if data.assigned_to_user_id:
        if not await db.users.find_one({"id": data.assigned_to_user_id}):
            raise HTTPException(404, "Assigned user not found")

    doc = data.model_dump()
    doc["id"] = _new_id()
    doc["status"] = "pending"
    doc["project_code"] = ""
    doc["audit_code"] = ""
    doc["created_by_user_id"] = user["id"] if user else None
    doc["created_by_username"] = user.get("username", "") if user else ""
    doc["created_at"] = _now()

    if data.project_id:
        p = await db.projects.find_one({"id": data.project_id}, {"_id": 0, "site_location": 1, "project_code": 1, "job_no": 1})
        if p:
            doc["project_code"] = p.get("job_no") or p.get("project_code") or ""
            if not (data.site_location or "").strip():
                doc["site_location"] = p.get("site_location") or ""

    await db.tasks.insert_one(doc.copy())
    doc["sr_no"] = 1
    await _enrich_task(doc)
    doc.pop("_id", None)
    
    if doc.get("assigned_to_user_id"):
        await _notify_user(
            user_id=doc["assigned_to_user_id"],
            message=f"You have been assigned a task: {doc.get('work')}",
            related_task_id=doc["id"]
        )
        
    return doc


@api_router.put("/tasks/{task_id}")
async def update_task(task_id: str, data: TaskIn):
    old_task = await db.tasks.find_one({"id": task_id})
    if not old_task:
        raise HTTPException(404, "Task not found")
    if data.assigned_to_user_id:
        if not await db.users.find_one({"id": data.assigned_to_user_id}):
            raise HTTPException(404, "Assigned user not found")

    update = data.model_dump()
    if data.project_id:
        p = await db.projects.find_one({"id": data.project_id}, {"_id": 0, "site_location": 1, "project_code": 1, "job_no": 1})
        if p:
            update["project_code"] = p.get("job_no") or p.get("project_code") or ""
            if not (data.site_location or "").strip():
                update["site_location"] = p.get("site_location") or ""

    result = await db.tasks.find_one_and_update(
        {"id": task_id}, {"$set": update}, return_document=True, projection={"_id": 0}
    )
    
    if result.get("assigned_to_user_id") and result.get("assigned_to_user_id") != old_task.get("assigned_to_user_id"):
        await _notify_user(
            user_id=result["assigned_to_user_id"],
            message=f"You have been assigned a task: {result.get('work')}",
            related_task_id=result["id"]
        )

    result["sr_no"] = 0
    await _enrich_task(result)
    return result


@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, user: dict = Depends(get_current_user_safe)):
    if user and user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins are permitted to delete tasks")
    result = await db.tasks.delete_one({"id": task_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Task not found")
    return {"ok": True}


class TaskStatusUpdate(BaseModel):
    status: str

@api_router.put("/tasks/{task_id}/status")
async def update_task_status(task_id: str, data: TaskStatusUpdate):
    if data.status not in ["pending", "in progress", "done", "cancelled"]:
        raise HTTPException(400, "Invalid status")
    
    t = await db.tasks.find_one({"id": task_id})
    if not t:
        raise HTTPException(404, "Task not found")
    
    result = await db.tasks.find_one_and_update(
        {"id": task_id},
        {"$set": {"status": data.status}},
        return_document=True,
        projection={"_id": 0}
    )
    result["sr_no"] = 0
    await _enrich_task(result)
    return result


@api_router.get("/")
async def root():
    return {"message": "Creator Consultant API", "status": "ok"}


# Backup module — Google Drive auto-backup
import backup as backup_module
backup_module.init(
    db,
    collections_to_backup=[
        "projects", "clients", "architects", "payments",
        "audits", "audit_payments", "audit_quote_revisions",
        "offers", "activity_log", "quote_revisions", "counters",
        "documents", "document_types",
        "site_visits", "site_visit_templates", "users", "notifications",
        "push_subscriptions", "app_settings", "company_details",
    ],
)
api_router.include_router(backup_module.router)


# ---------------------- COMPANY DETAILS ----------------------

async def seed_company_details():
    existing = await db.company_details.find_one({"id": "singleton"})
    if not existing:
        doc = {
            "id": "singleton",
            "name": "CREATOR RCC CONSULTANT LLP",
            "address": "A-001, sidhhivinayak park, Sector No 8A , Plot No-21, Airoli Nr. D Mart,Navi Mumbai 400708, Thane, Maharashtra, 400708",
            "gstin": "27AASFC7539E1Z2",
            "mobile": "9892683460",
            "pan": "AASFC7539E",
            "email": "project@creatorconsultant.net",
            "bank_name": "Kotak Mahindra Bank",
            "bank_account_name": "CREATOR RCC CONSULTANT LLP",
            "bank_ifsc": "KKBK0001360",
            "bank_account_no": "9987076241",
            "bank_branch": "Kotak Mahindra Bank ,SHIVSHANKAR PLAZA I SECTOR EIGHT AIROLI BRANCH",
            "upi_id": "creatorconsultantLLP@kotak",
            "qr_code_url": "",
            "company_logo_url": "",
        }
        await db.company_details.insert_one(doc.copy())
        logger.info("Seeded default company details.")

@api_router.get("/company-details")
async def get_company_details():
    doc = await db.company_details.find_one({"id": "singleton"}, {"_id": 0})
    if not doc:
        doc = {
            "id": "singleton",
            "name": "CREATOR RCC CONSULTANT LLP",
            "address": "A-001, sidhhivinayak park, Sector No 8A , Plot No-21, Airoli Nr. D Mart,Navi Mumbai 400708, Thane, Maharashtra, 400708",
            "gstin": "27AASFC7539E1Z2",
            "mobile": "9892683460",
            "pan": "AASFC7539E",
            "email": "project@creatorconsultant.net",
            "bank_name": "Kotak Mahindra Bank",
            "bank_account_name": "CREATOR RCC CONSULTANT LLP",
            "bank_ifsc": "KKBK0001360",
            "bank_account_no": "9987076241",
            "bank_branch": "Kotak Mahindra Bank ,SHIVSHANKAR PLAZA I SECTOR EIGHT AIROLI BRANCH",
            "upi_id": "creatorconsultantLLP@kotak",
            "qr_code_url": "",
            "company_logo_url": "",
        }
        await db.company_details.insert_one(doc.copy())
        doc.pop("_id", None)
    return doc

@api_router.put("/company-details")
async def update_company_details(data: CompanyDetailsIn):
    update_data = data.model_dump()
    result = await db.company_details.find_one_and_update(
        {"id": "singleton"},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Company details not found")
    return result

@api_router.post("/company-details/uploads")
async def upload_company_asset(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(400, "Missing filename")
    ext = (file.filename.rsplit(".", 1)[-1] or "jpg").lower()
    if ext not in {"jpg", "jpeg", "png", "webp", "pdf"}:
        raise HTTPException(400, "Only images or PDF files are allowed")
    fname = f"company_asset_{secrets.token_urlsafe(8)}.{ext}"
    content_type = file.content_type or ("application/pdf" if ext == "pdf" else f"image/{ 'jpeg' if ext == 'jpg' else ext }")
    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty upload")
    
    await _photo_bucket.upload_from_stream(
        fname, io.BytesIO(data), metadata={"content_type": content_type},
    )
    return {"url": f"/api/uploads/company/{fname}", "filename": fname}

@auth_public_router.get("/uploads/company/{filename}")
async def serve_company_asset(filename: str):
    try:
        stream = await _photo_bucket.open_download_stream_by_name(filename)
        ct = (getattr(stream, "metadata", None) or {}).get("content_type") or "image/jpeg"

        async def gen():
            try:
                while True:
                    chunk = await stream.readchunk()
                    if not chunk:
                        break
                    yield chunk
            finally:
                try:
                    stream.close()
                except Exception:
                    pass

        return StreamingResponse(
            gen(),
            media_type=ct,
        )
    except Exception:
        raise HTTPException(404, "File not found")


# ---------------------- INVOICES ----------------------

def num_to_words_indian(number: float) -> str:
    whole = int(number)
    frac = int(round((number - whole) * 100))
    
    if whole == 0 and frac == 0:
        return "Zero Rupees Only"
    
    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
             "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    
    def _convert_below_thousand(n):
        if n == 0:
            return ""
        elif n < 20:
            return units[n]
        elif n < 100:
            return tens[n // 10] + (" " + units[n % 10] if n % 10 != 0 else "")
        else:
            return units[n // 100] + " Hundred" + (" and " + _convert_below_thousand(n % 100) if n % 100 != 0 else "")

    def _helper(n):
        words = []
        if n >= 10000000: # Crore
            crore = n // 10000000
            words.append(_convert_below_thousand(crore) + " Crore")
            n %= 10000000
        if n >= 100000: # Lakh
            lakh = n // 100000
            words.append(_convert_below_thousand(lakh) + " Lakh")
            n %= 100000
        if n >= 1000: # Thousand
            thousand = n // 1000
            words.append(_convert_below_thousand(thousand) + " Thousand")
            n %= 1000
        if n > 0:
            words.append(_convert_below_thousand(n))
        return " ".join(words)

    whole_words = _helper(whole).strip() if whole > 0 else "Zero"
    result = whole_words + " Rupees"
    
    if frac > 0:
        frac_words = ""
        if frac < 20:
            frac_words = units[frac]
        else:
            frac_words = tens[frac // 10] + (" " + units[frac % 10] if frac % 10 != 0 else "")
        result += f" and {frac_words.strip()} Paise"
        
    return result + " Only"


async def _next_proforma_no() -> str:
    counter = await db.counters.find_one_and_update(
        {"_id": "proforma_invoice"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = (counter or {}).get("seq", 1)
    return f"CC > PIC > {seq:03d}"


async def _next_tax_invoice_no() -> str:
    counter = await db.counters.find_one_and_update(
        {"_id": "tax_invoice"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = (counter or {}).get("seq", 1)
    return f"CC > ARL > {seq:03d}"


@api_router.get("/invoices/paginated", response_model=PaginatedInvoices)
async def list_invoices_paginated(
    page: int = 1, 
    limit: int = 25, 
    q: Optional[str] = None,
    type: Optional[str] = None,
    client_id: Optional[str] = None
):
    skip = (page - 1) * limit
    query = {}
    if type:
        query["type"] = type
    if client_id:
        query["client_id"] = client_id
        
    if q:
        s = q.strip()
        query["$or"] = [
            {"invoice_no": {"$regex": s, "$options": "i"}},
            {"client_name": {"$regex": s, "$options": "i"}},
            {"client_gstin": {"$regex": s, "$options": "i"}},
            {"client_pan": {"$regex": s, "$options": "i"}},
            {"place_of_supply": {"$regex": s, "$options": "i"}}
        ]
        
    total = await db.invoices.count_documents(query)
    items = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"data": items, "total": total}


@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices():
    docs = await db.invoices.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return docs
@api_router.post("/invoices/bulk-import-b2b")
async def bulk_import_invoices_b2b(type: str = "tax", file: UploadFile = File(...), current_user: dict = Depends(auth_module.get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only Excel files (.xlsx, .xls) are supported.")
    
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheets_to_process = [name for name in wb.sheetnames if "b2b" in name.lower() or "b2cl" in name.lower()]
        if not sheets_to_process:
            sheets_to_process = [wb.active.title]
    except Exception as e:
        raise HTTPException(400, f"Failed to read Excel file: {str(e)}")

    imported = 0
    skipped = 0
    
    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        is_b2cl = "b2cl" in sheet_name.lower()
        
        header_row_idx = None
        gstin_col = name_col = pos_col = inv_no_col = inv_date_col = taxable_val_col = rate_col = None
        
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for c_idx, cell in enumerate(row):
                val = str(cell).strip().lower() if cell else ""
                if "gstin" in val and "e-commerce" not in val:
                    gstin_col = c_idx
                elif "customer name" in val or "receiver name" in val:
                    name_col = c_idx
                elif "place of supply" in val:
                    pos_col = c_idx
                elif "invoice number" in val:
                    inv_no_col = c_idx
                elif "invoice date" in val:
                    inv_date_col = c_idx
                elif "taxable value" in val:
                    taxable_val_col = c_idx
                elif val == "rate":
                    rate_col = c_idx
            
            if inv_date_col is not None and taxable_val_col is not None:
                if is_b2cl or name_col is not None:
                    header_row_idx = r_idx
                    break
                
        if header_row_idx is None:
            continue
            
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            try:
                if is_b2cl:
                    pos = str(row[pos_col]).strip() if pos_col is not None and len(row) > pos_col and row[pos_col] else ""
                    if not pos or pos.lower() == "none" or "total" in pos.lower():
                        continue
                    name = f"Unregistered Customer - {pos}"
                    gstin = ""
                else:
                    name = str(row[name_col]).strip() if name_col is not None and len(row) > name_col and row[name_col] else ""
                    if not name or name.lower() == "none" or "total" in name.lower():
                        continue
                    gstin = str(row[gstin_col]).strip() if gstin_col is not None and len(row) > gstin_col and row[gstin_col] else ""
                    if gstin.lower() == "none": gstin = ""
                    
                inv_date = str(row[inv_date_col]).strip() if inv_date_col is not None and len(row) > inv_date_col and row[inv_date_col] else ""
                if not inv_date or inv_date.lower() == "none" or "total" in inv_date.lower():
                    continue
                    
                try:
                    if isinstance(row[inv_date_col], datetime):
                        parsed_date = row[inv_date_col].strftime("%Y-%m-%d")
                    else:
                        parsed_date = datetime.strptime(inv_date.split()[0], "%d-%b-%Y").strftime("%Y-%m-%d")
                except Exception:
                    parsed_date = _now().split("T")[0]
                    
                rate_val = 18.0
                if rate_col is not None and len(row) > rate_col and row[rate_col]:
                    try:
                        rate_val = float(str(row[rate_col]).strip())
                    except ValueError:
                        pass
                        
                taxable_val = 0.0
                if taxable_val_col is not None and len(row) > taxable_val_col and row[taxable_val_col]:
                    try:
                        taxable_val = float(str(row[taxable_val_col]).strip())
                    except ValueError:
                        pass
                        
                if taxable_val <= 0:
                    continue
                    
                inv_no = str(row[inv_no_col]).strip() if inv_no_col is not None and len(row) > inv_no_col and row[inv_no_col] else ""
                if not inv_no or inv_no.lower() == "none":
                    prefix = "B2CL" if is_b2cl else "B2B"
                    if type == "proforma":
                        prefix = "PRO"
                    inv_no = f"{prefix}-{_new_id()[:6].upper()}"
                    
                dup_inv = await db.invoices.find_one({"invoice_no": inv_no})
                if dup_inv:
                    skipped += 1
                    continue
                
                import re
                safe_name = re.escape(name)
                client = await db.clients.find_one({"name": {"$regex": f"^{safe_name}$", "$options": "i"}})
                if not client:
                    client_id = _new_id() if type == "tax" else ""
                    client = {
                        "name": name,
                        "gstin": gstin,
                        "place_of_supply": pos if is_b2cl else "",
                        "pan": gstin[2:12] if len(gstin) >= 12 else "",
                        "phone": "",
                        "email": "",
                        "company": name,
                        "address": "",
                        "id": client_id,
                        "created_at": _now()
                    }
                    if type == "tax":
                        _stamp_edit(client)
                        await db.clients.insert_one(client)
                else:
                    client_id = client["id"]
                    
                invoice_id = _new_id()
                
                effective_type = type
                if "CC / ARL" in inv_no.upper() or "CC/ARL" in inv_no.upper():
                    effective_type = "tax"
                elif "CC / PIC" in inv_no.upper() or "CC/PIC" in inv_no.upper():
                    effective_type = "proforma"
    
                doc = {
                    "type": effective_type,
                    "invoice_date": parsed_date,
                    "expiry_date": "",
                    "hsn_code": "998332",
                    "client_id": client_id,
                    "client_name": client["name"],
                    "client_address": client.get("address", ""),
                    "client_gstin": client.get("gstin", ""),
                    "client_mobile": client.get("phone", ""),
                    "client_pan": client.get("pan", ""),
                    "place_of_supply": client.get("place_of_supply", ""),
                    "service_description": "Professional Services",
                    "qty": 1.0,
                    "rate": taxable_val,
                    "gst_percent": rate_val,
                    "tds_percent": 10.0,
                    "tds_section": "194J",
                    "received_amount": 0.0,
                    "id": invoice_id,
                    "invoice_no": inv_no,
                    "created_by_user_id": current_user.get("id"),
                    "created_by_username": current_user.get("username"),
                    "created_at": _now()
                }
                await db.invoices.insert_one(doc)
                imported += 1
            except Exception as e:
                print(f"Row skipped due to error: {e}")
                skipped += 1
            
    if imported == 0 and skipped == 0:
        raise HTTPException(400, "Could not find valid B2B or B2CL invoice data to import.")
        
    return {"imported": imported, "skipped": skipped, "message": f"Successfully imported {imported} invoices. Skipped {skipped} duplicates."}


@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(data: InvoiceIn, current_user: dict = Depends(auth_module.get_current_user)):
    invoice_data = data.model_dump()
    invoice_data["place_of_supply"] = format_place_of_supply(invoice_data.get("place_of_supply"), invoice_data.get("client_gstin"))
    
    if invoice_data["type"] == "proforma":
        invoice_no = await _next_proforma_no()
    else:
        invoice_no = await _next_tax_invoice_no()
        
    invoice_id = _new_id()
    doc = {
        **invoice_data,
        "id": invoice_id,
        "invoice_no": invoice_no,
        "created_by_user_id": current_user.get("id"),
        "created_by_username": current_user.get("username"),
        "created_at": _now()
    }
    
    await db.invoices.insert_one(doc.copy())
    
    # Auto-update client in db
    if invoice_data["client_id"]:
        update_fields = {}
        if invoice_data.get("client_gstin"):
            update_fields["gstin"] = invoice_data["client_gstin"]
        if invoice_data.get("client_pan"):
            update_fields["pan"] = invoice_data["client_pan"]
        if invoice_data.get("place_of_supply"):
            update_fields["place_of_supply"] = invoice_data["place_of_supply"]
            
        if update_fields:
            await db.clients.update_one(
                {"id": invoice_data["client_id"]},
                {"$set": update_fields}
            )
            
    # Auto-record payment if this is a tax invoice linked to a project
    if invoice_data["type"] == "tax" and invoice_data.get("project_id"):
        project_id = invoice_data["project_id"]
        project = await db.projects.find_one({"id": project_id}, {"_id": 0})
        if project:
            items = invoice_data.get("items", [])
            if not items:
                items = [{
                    "qty": float(invoice_data.get("qty", 1.0)),
                    "rate": float(invoice_data.get("rate", 0.0))
                }]
            base_value = sum(float(it.get("qty", 1.0)) * float(it.get("rate", 0.0)) for it in items)
            gst_percent = float(invoice_data.get("gst_percent", 18))
            gst_amount = base_value * (gst_percent / 100)
            total_amount_with_gst = base_value + gst_amount
            
            tds_percent = float(invoice_data.get("tds_percent", 0))
            tds_amount = base_value * (tds_percent / 100) if tds_percent > 0 else 0
            
            payable_amount = round(total_amount_with_gst - tds_amount, 2)
            
            # Create payment
            payment_doc = {
                "id": _new_id(),
                "project_id": project_id,
                "project_code": project.get("project_code", ""),
                "amount": float(payable_amount),
                "taxable_amount": float(base_value),
                "payment_date": invoice_data.get("invoice_date", _now()[:10]),
                "notes": f"Auto-recorded from Tax Invoice",
                "invoice_no": invoice_no,
                "created_at": _now(),
            }
            _stamp_edit(payment_doc)
            await db.payments.insert_one(payment_doc.copy())
            
            # Recalculate project totals
            await _recalculate_project_received(project_id)
            
            await _log_activity(
                project_id, project.get("project_code", ""),
                "PAYMENT ADDED",
                f"Amount: ₹ {float(base_value):,.2f} | Note: Auto-recorded from Tax Invoice {invoice_no}",
            )

    doc.pop("_id", None)
    return doc


@api_router.get("/invoices/export/excel")
async def export_invoices_excel(start_date: Optional[str] = None, end_date: Optional[str] = None):
    query = {"type": "tax"}
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        query["invoice_date"] = date_query

    invoices = await db.invoices.find(query, {"_id": 0}).sort("invoice_date", -1).to_list(10000)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tax Invoices"
    
    headers = [
        "Invoice No", "Date", "Client Name", "Client GSTIN", "Place of Supply",
        "Project Linked", "Total Base Value (INR)", "GST %", "GST Amount (INR)", 
        "TDS %", "TDS Amount (INR)", "Payable Amount (INR)", "Received Amount (INR)"
    ]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="061A11")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for inv in invoices:
        items = inv.get("items", [])
        base_value = sum(float(it.get("qty", 1.0)) * float(it.get("rate", 0.0)) for it in items)
        gst_percent = float(inv.get("gst_percent", 18))
        gst_amount = base_value * (gst_percent / 100)
        total_amount_with_gst = base_value + gst_amount
        
        tds_percent = float(inv.get("tds_percent", 0))
        tds_amount = base_value * (tds_percent / 100) if tds_percent > 0 else 0
        
        payable_amount = round(total_amount_with_gst - tds_amount, 2)
        
        ws.append([
            inv.get("invoice_no", ""),
            inv.get("invoice_date", ""),
            inv.get("client_name", ""),
            inv.get("client_gstin", ""),
            inv.get("place_of_supply", ""),
            "Yes" if inv.get("project_id") else "No",
            round(base_value, 2),
            gst_percent,
            round(gst_amount, 2),
            tds_percent,
            round(tds_amount, 2),
            payable_amount,
            float(inv.get("received_amount", 0) or 0)
        ])

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"Tax_Invoices_Export_{start_date or 'ALL'}_to_{end_date or 'ALL'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@api_router.put("/invoices/{invoice_id}", response_model=Invoice)
async def update_invoice(invoice_id: str, data: InvoiceIn, current_user: dict = Depends(auth_module.get_current_user)):
    invoice_data = data.model_dump()
    
    existing = await db.invoices.find_one({"id": invoice_id})
    if not existing:
        raise HTTPException(404, "Invoice not found")
        
    doc = {
        **existing,
        **invoice_data,
        "place_of_supply": format_place_of_supply(invoice_data.get("place_of_supply"), invoice_data.get("client_gstin")),
        "last_edited_by_user_id": current_user.get("id"),
        "last_edited_by_username": current_user.get("username"),
        "last_edited_at": _now()
    }
    
    inv_no = existing.get("invoice_no", "")
    t = existing.get("type", invoice_data.get("type", "tax"))
    if "CC / ARL" in inv_no.upper() or "CC/ARL" in inv_no.upper():
        t = "tax"
    elif "CC / PIC" in inv_no.upper() or "CC/PIC" in inv_no.upper():
        t = "proforma"
        
    doc["type"] = t
    doc["invoice_no"] = inv_no
    
    await db.invoices.replace_one({"id": invoice_id}, doc)
    
    if invoice_data["client_id"]:
        update_fields = {}
        if invoice_data.get("client_gstin"):
            update_fields["gstin"] = invoice_data["client_gstin"]
        if invoice_data.get("client_pan"):
            update_fields["pan"] = invoice_data["client_pan"]
        if invoice_data.get("place_of_supply"):
            update_fields["place_of_supply"] = format_place_of_supply(invoice_data["place_of_supply"], invoice_data.get("client_gstin"))
            
        if update_fields:
            await db.clients.update_one(
                {"id": invoice_data["client_id"]},
                {"$set": update_fields}
            )
            
    # Sync payment record if needed
    old_project_id = existing.get("project_id")
    new_project_id = invoice_data.get("project_id")
    
    # Calculate payable amount
    items = invoice_data.get("items", [])
    if not items:
        items = [{
            "qty": float(invoice_data.get("qty", 1.0)),
            "rate": float(invoice_data.get("rate", 0.0))
        }]
    base_value = sum(float(it.get("qty", 1.0)) * float(it.get("rate", 0.0)) for it in items)
    gst_percent = float(invoice_data.get("gst_percent", 18))
    gst_amount = base_value * (gst_percent / 100)
    total_amount_with_gst = base_value + gst_amount
    tds_percent = float(invoice_data.get("tds_percent", 0))
    tds_amount = base_value * (tds_percent / 100) if tds_percent > 0 else 0
    payable_amount = round(total_amount_with_gst - tds_amount, 2)

    existing_payment = await db.payments.find_one({"invoice_no": inv_no}) if inv_no else None
    
    if t == "tax" and new_project_id:
        project = await db.projects.find_one({"id": new_project_id}, {"_id": 0})
        if project:
            if existing_payment:
                # Update existing payment
                await db.payments.update_one(
                    {"id": existing_payment["id"]},
                    {"$set": {
                        "project_id": new_project_id,
                        "project_code": project.get("project_code", ""),
                        "amount": float(payable_amount),
                        "taxable_amount": float(base_value),
                        "payment_date": invoice_data.get("invoice_date", _now()[:10]),
                        "last_edited_by_user_id": current_user.get("id"),
                        "last_edited_by_username": current_user.get("username"),
                        "last_edited_at": _now()
                    }}
                )
            else:
                # Create new payment
                payment_doc = {
                    "id": _new_id(),
                    "project_id": new_project_id,
                    "project_code": project.get("project_code", ""),
                    "amount": float(payable_amount),
                    "taxable_amount": float(base_value),
                    "payment_date": invoice_data.get("invoice_date", _now()[:10]),
                    "notes": "Auto-recorded from Tax Invoice",
                    "invoice_no": inv_no,
                    "created_at": _now(),
                }
                _stamp_edit(payment_doc)
                await db.payments.insert_one(payment_doc)
            
            # Recalculate new project totals
            await _recalculate_project_received(new_project_id)
            
            # If project changed, recalculate old project too
            if old_project_id and old_project_id != new_project_id:
                await _recalculate_project_received(old_project_id)
    else:
        # If not tax, or no project, remove payment if it exists
        if existing_payment:
            await db.payments.delete_one({"id": existing_payment["id"]})
            if existing_payment.get("project_id"):
                await _recalculate_project_received(existing_payment["project_id"])            
    doc.pop("_id", None)
    return doc


@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str):
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(404, "Invoice not found")
        
    invoice_no = invoice.get("invoice_no", "")
    invoice_type = invoice.get("type", "tax_invoice")
    
    res = await db.invoices.delete_one({"id": invoice_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Invoice not found")
        
    # Rollback the counter if this was the very last invoice generated
    import re
    match = re.search(r'(\d+)$', invoice_no)
    if match:
        seq = int(match.group(1))
        counter_id = "proforma" if invoice_type == "proforma" else "tax_invoice"
        counter = await db.counters.find_one({"_id": counter_id})
        if counter and counter.get("seq") == seq:
            await db.counters.update_one({"_id": counter_id}, {"$inc": {"seq": -1}})
    # Also delete the corresponding payment record if it exists
    if invoice_no:
        payment = await db.payments.find_one({"invoice_no": invoice_no})
        if payment:
            await db.payments.delete_one({"id": payment["id"]})
            if payment.get("project_id"):
                await _recalculate_project_received(payment["project_id"])
                project = await db.projects.find_one({"id": payment["project_id"]}, {"_id": 0})
                if project:
                    await _log_activity(
                        payment["project_id"], project.get("project_code", ""),
                        "PAYMENT DELETED",
                        f"Auto-deleted due to invoice deletion: {invoice_no}"
                    )
            
    return {"ok": True}


async def _build_invoice_document_pdf(invoice: dict) -> bytes:
    cd = await db.company_details.find_one({"id": "singleton"})
    if not cd:
        cd = {
            "name": "CREATOR RCC CONSULTANT LLP",
            "address": "A-001, sidhhivinayak park, Airoli, Navi Mumbai",
            "gstin": "27AASFC7539E1Z2",
            "mobile": "9892683460",
            "pan": "AASFC7539E",
            "email": "project@creatorconsultant.net",
            "bank_name": "Kotak Mahindra Bank",
            "bank_account_name": "CREATOR RCC CONSULTANT LLP",
            "bank_ifsc": "KKBK0001360",
            "bank_account_no": "9987076241",
            "bank_branch": "Airoli Branch",
            "upi_id": "creatorconsultantLLP@kotak",
            "qr_code_url": "",
            "company_logo_url": ""
        }

    is_proforma = invoice.get("type") == "proforma"

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    margin = 10 * mm
    printable_width = width - 2 * margin

    # ── Title line ────────────────────────────────────────────────────────
    title_str = "PROFORMA INVOICE" if is_proforma else "TAX INVOICE"
    c.setFont("Roboto-Medium", 10)
    c.drawString(margin, height - 6.5 * mm, title_str)

    if not is_proforma:
        title_w = pdfmetrics.stringWidth(title_str, "Roboto-Medium", 10)
        badge_text = "ORIGINAL FOR RECIPIENT"
        badge_text_w = pdfmetrics.stringWidth(badge_text, "Roboto-Medium", 8.5)
        
        badge_x = margin + title_w + 3 * mm
        badge_y = height - 8 * mm
        badge_w = badge_text_w + 4 * mm
        badge_h = 5 * mm
        
        c.setStrokeColor(colors.HexColor("#84849A"))
        c.setLineWidth(1.0)
        c.roundRect(badge_x, badge_y, badge_w, badge_h, 2, fill=0, stroke=1)
        c.setFillColor(colors.HexColor("#84849A"))
        c.setFont("Roboto-Medium", 8.5)
        c.drawCentredString(badge_x + badge_w / 2, badge_y + 1.5 * mm, badge_text)
        c.setFillColor(colors.black)

    # ── Header box ────────────────────────────────────────────────────────
    y_top = height - 10 * mm
    
    # Pre-calculate to find dynamic header height
    v_split = margin + 98 * mm
    info_x = margin + 24 * mm
    info_max_w = v_split - info_x - 3 * mm
    styles = getSampleStyleSheet()
    addr_style = ParagraphStyle(
        'AddrStyleInv', parent=styles['Normal'], fontName='Roboto',
        fontSize=8.5, leading=10, textColor=colors.black,
    )
    addr_para_tmp = Paragraph(cd.get("address", ""), addr_style)
    _, addr_para_h = addr_para_tmp.wrap(info_max_w, 25 * mm)
    
    _gstin_y = y_top - 8 * mm - addr_para_h - 3 * mm
    _gstin_val_y = _gstin_y - 3.5 * mm
    _pan_y = _gstin_val_y - 3.8 * mm
    _email_y = _pan_y - 4 * mm
    
    header_h = y_top - (_email_y - 5 * mm)
    y_header_bottom = y_top - header_h

    c.setStrokeColor(colors.black)
    c.setLineWidth(1.0)
    c.rect(margin, y_header_bottom, printable_width, header_h, fill=0, stroke=1)

    # Left panel width = 90mm, right panel = 90mm
    v_split = margin + 98 * mm
    c.line(v_split, y_header_bottom, v_split, y_top)

    # ── Logo ──────────────────────────────────────────────────────────────
    logo_path = Path(__file__).parent.parent / "frontend" / "public" / "logo.jpg"
    logo_bytes = None
    logo_is_pdf = False
    logo_url = cd.get("company_logo_url")
    if logo_url:
        logo_bytes = await _load_photo_bytes(logo_url)
        if logo_bytes and logo_bytes.startswith(b'%PDF'):
            logo_is_pdf = True

    logo_size = 18 * mm
    logo_x = margin + 4 * mm
    # Center logo vertically in the 44mm header box
    logo_y = y_header_bottom + (header_h - logo_size) / 2

    if logo_bytes and not logo_is_pdf:
        try:
            logo_img = ImageReader(io.BytesIO(logo_bytes))
            c.drawImage(logo_img, logo_x, logo_y, width=logo_size, height=logo_size,
                        preserveAspectRatio=True, mask='auto')
        except Exception:
            if logo_path.exists():
                try:
                    c.drawImage(str(logo_path), logo_x, logo_y, width=logo_size, height=logo_size,
                                preserveAspectRatio=True)
                except Exception:
                    pass
    elif not logo_is_pdf:
        if logo_path.exists():
            try:
                c.drawImage(str(logo_path), logo_x, logo_y, width=logo_size, height=logo_size,
                            preserveAspectRatio=True)
            except Exception:
                pass

    # ── Company info (left panel, right of logo) ──────────────────────────
    info_x = margin + 24 * mm
    info_max_w = v_split - info_x - 3 * mm   # available width for text

    c.setFillColor(colors.HexColor("#0A6B9F"))
    c.setFont("Roboto-Medium", 10.5)
    c.drawString(info_x, y_top - 5 * mm, cd.get("name", ""))
    c.setFillColor(colors.black)

    # Wrapped address
    styles = getSampleStyleSheet()
    addr_style = ParagraphStyle(
        'AddrStyleInv',
        parent=styles['Normal'],
        fontName='Roboto',
        fontSize=8.5,
        leading=10,
        textColor=colors.black,
    )
    addr_para = Paragraph(cd.get("address", ""), addr_style)
    _, addr_para_h = addr_para.wrap(info_max_w, 25 * mm)
    addr_para.drawOn(c, info_x, y_top - 8 * mm - addr_para_h)

    # GSTIN + Mobile labels and values
    gstin_y = y_top - 8 * mm - addr_para_h - 3 * mm
    c.setFont("Roboto", 8.5)
    c.drawString(info_x, gstin_y, "GSTIN:")
    c.drawString(info_x + 35 * mm, gstin_y, "Mobile:")
    
    gstin_val_y = gstin_y - 3.5 * mm
    c.setFont("Roboto", 8.5)
    c.drawString(info_x, gstin_val_y, cd.get("gstin", ""))
    c.drawString(info_x + 35 * mm, gstin_val_y, cd.get("mobile", ""))

    pan_y = gstin_val_y - 3.8 * mm
    c.setFont("Roboto", 8.5)
    c.drawString(info_x, pan_y, "PAN Number:")
    pan_w = pdfmetrics.stringWidth("PAN Number: ", "Roboto", 8.5)
    c.setFont("Roboto", 8.5)
    c.drawString(info_x + pan_w, pan_y, cd.get("pan", ""))

    email_y = pan_y - 4 * mm
    c.setFont("Roboto", 8.5)
    c.drawString(info_x, email_y, "Email:")
    email_w = pdfmetrics.stringWidth("Email: ", "Roboto", 8.5)
    c.setFont("Roboto", 8.5)
    c.drawString(info_x + email_w, email_y, cd.get("email", ""))

    # ── Right panel: invoice meta ─────────────────────────────────────────
    inv_date_str = ""
    if invoice.get("invoice_date"):
        try:
            inv_date_str = datetime.strptime(invoice["invoice_date"], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            inv_date_str = invoice["invoice_date"]

    exp_date_str = ""
    if invoice.get("expiry_date"):
        try:
            exp_date_str = datetime.strptime(invoice["expiry_date"], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            exp_date_str = invoice["expiry_date"]

    rp_w = (width - margin) - v_split
    # Push the text down to vertically center it in the 22mm box
    y_meta_top = y_top - 10 * mm

    if is_proforma:
        # Three columns, centered in thirds
        c1 = v_split + (rp_w * 0.23)
        c2 = v_split + (rp_w * 0.60)
        c3 = v_split + (rp_w * 0.86)
        
        c.setFont("Roboto-Bold", 8.5)
        c.drawCentredString(c1, y_meta_top, "Proforma Invoice No.")
        c.drawCentredString(c2, y_meta_top, "Proforma Date")
        c.drawCentredString(c3, y_meta_top, "Expiry Date")
        
        c.setFont("Roboto", 8.5)
        c.drawCentredString(c1, y_meta_top - 4 * mm, invoice.get("invoice_no", ""))
        c.drawCentredString(c2, y_meta_top - 4 * mm, inv_date_str)
        c.drawCentredString(c3, y_meta_top - 4 * mm, exp_date_str)
    else:
        # Two columns, centered in halves
        c1 = v_split + (rp_w / 4)
        c2 = v_split + (rp_w * 3 / 4)
        
        c.setFont("Roboto-Bold", 8.5)
        c.drawCentredString(c1, y_meta_top, "Invoice No.")
        c.drawCentredString(c2, y_meta_top, "Invoice Date")
        
        c.setFont("Roboto", 8.5)
        c.drawCentredString(c1, y_meta_top - 4 * mm, invoice.get("invoice_no", ""))
        c.drawCentredString(c2, y_meta_top - 4 * mm, inv_date_str)

    # Horizontal divider in right panel below invoice no / date
    y_split_line = y_top - 22 * mm
    c.setLineWidth(1.0)
    c.line(v_split, y_split_line, width - margin, y_split_line)

    # HSN CODE | PAN NO — two sub-columns in lower part of right panel
    v_pan_split = v_split + (rp_w / 2)
    # Line removed based on user feedback

    c1_hsn = v_split + 12.5 * mm
    c2_pan = (width - margin) - 12.5 * mm

    h2 = header_h - 22 * mm
    y_meta_bottom_top = y_split_line - (h2 / 2 - 2 * mm)

    c.setFont("Roboto", 8.5)
    c.drawCentredString(c1_hsn, y_meta_bottom_top, "HSN CODE")
    c.drawCentredString(c2_pan, y_meta_bottom_top, "PAN NO")

    c.setFont("Roboto", 8.5)
    c.drawCentredString(c1_hsn, y_meta_bottom_top - 4 * mm, invoice.get("hsn_code", "998332"))
    c.drawCentredString(c2_pan, y_meta_bottom_top - 4 * mm, cd.get("pan", ""))

    # ── Bill To block ─────────────────────────────────────────────────────
    y_billto = y_header_bottom
    billto_h = 28 * mm
    y_billto_bottom = y_billto - billto_h

    c.setLineWidth(1.0)
    c.rect(margin, y_billto_bottom, printable_width, billto_h, fill=0, stroke=1)

    bx = margin + 2 * mm   # left text offset inside bill-to box

    c.setFillColor(colors.black)
    c.setFont("Roboto", 9)
    c.drawString(bx, y_billto - 3.5 * mm, "BILL TO")

    c.setFont("Roboto-Bold", 9) # client name full bold
    c.drawString(bx, y_billto - 8.5 * mm, invoice.get("client_name", "").upper())

    # Full address — wrapped
    addr_client_style = ParagraphStyle(
        'ClientAddrStyleInv',
        parent=styles['Normal'],
        fontName='Roboto',
        fontSize=8.5,
        leading=10,
        textColor=colors.black,
    )
    client_addr = invoice.get("client_address", "")
    addr_client_para = Paragraph(f"Address:  {client_addr}", addr_client_style)
    _, addr_client_h = addr_client_para.wrap(printable_width - 4 * mm, 14 * mm)
    addr_client_para.drawOn(c, bx, y_billto - 10 * mm - addr_client_h)

    # GSTIN row
    gstin_row_y = y_billto - 10 * mm - addr_client_h - 4.5 * mm # space adjusted to match GSTIN/Mobile row
    c.setFont("Roboto", 8.5)
    c.drawString(bx, gstin_row_y, "GSTIN:")
    gstin_w = pdfmetrics.stringWidth("GSTIN: ", "Roboto", 8.5)
    c.setFont("Roboto", 8.5)
    c.drawString(bx + gstin_w, gstin_row_y, invoice.get("client_gstin") or "\u2014")
    
    c.setFont("Roboto", 8.5)
    c.drawString(bx + 40 * mm, gstin_row_y, "Place of Supply:")
    pos_w = pdfmetrics.stringWidth("Place of Supply: ", "Roboto", 8.5)
    c.setFont("Roboto", 8.5)
    pos = invoice.get("place_of_supply") or "\u2014"
    c.drawString(bx + 40 * mm + pos_w, gstin_row_y, format_place_of_supply(pos, invoice.get("client_gstin")))

    # Mobile / PAN row
    mob_row_y = gstin_row_y - 4.5 * mm
    c.setFont("Roboto", 8.5)
    c.drawString(bx, mob_row_y, "Mobile:")
    mob_w = pdfmetrics.stringWidth("Mobile: ", "Roboto", 8.5)
    c.setFont("Roboto", 8.5)
    c.drawString(bx + mob_w, mob_row_y, invoice.get("client_mobile") or "\u2014")
    
    c.setFont("Roboto", 8.5)
    c.drawString(bx + 32 * mm, mob_row_y, "PAN Number:")
    bill_pan_w = pdfmetrics.stringWidth("PAN Number: ", "Roboto", 8.5)
    c.setFont("Roboto", 8.5)
    c.drawString(bx + 32 * mm + bill_pan_w, mob_row_y, invoice.get("client_pan") or "\u2014")

    # ── Services table ────────────────────────────────────────────────────
    # Column widths (all in pts): matching user requested layout
    sno_w    = 14.0 * mm
    svc_w    = 102.1 * mm
    qty_w    = 14.0 * mm
    rate_w   = 17.6 * mm
    tax_w    = 18.7 * mm
    # col_x[i] = left edge of column i+1 (= right edge of column i)
    col_x = [
        margin + sno_w,
        margin + sno_w + svc_w,
        margin + sno_w + svc_w + qty_w,
        margin + sno_w + svc_w + qty_w + rate_w,
        margin + sno_w + svc_w + qty_w + rate_w + tax_w,
    ]
    # amount col right edge = width - margin

    y_table = y_billto_bottom
    
    # Dynamically calculate table bottom so footer hits the bottom margin
    required_bottom_space = 85 + 30 + 42 + (2 * mm) # HSN Summary table height changed to 42
    if not is_proforma:
        required_bottom_space += 21 * 4
        
    y_table_bottom = margin + required_bottom_space
    table_h = y_table - y_table_bottom

    c.setLineWidth(1.0)
    c.rect(margin, y_table_bottom, printable_width, table_h, fill=0, stroke=1)

    # Header row
    hdr_h = 22
    y_hdr_bottom = y_table - hdr_h
    c.setFillColor(colors.HexColor("#CFE9FA"))
    c.rect(margin, y_hdr_bottom, printable_width, hdr_h, fill=1, stroke=0)
    c.setStrokeColor(colors.black)
    c.setLineWidth(1.0)
    c.line(margin, y_hdr_bottom, width - margin, y_hdr_bottom)

    for x in col_x:
        c.line(x, y_table_bottom, x, y_table)

    c.setFillColor(colors.black)
    c.setFont("Roboto", 9)
    mid_sno   = margin + sno_w / 2
    mid_svc   = margin + sno_w + svc_w / 2
    mid_qty   = col_x[1] + qty_w / 2
    mid_rate  = col_x[2] + rate_w / 2
    mid_tax   = col_x[3] + tax_w / 2
    mid_amt   = col_x[4] + (width - margin - col_x[4]) / 2
    y_hdr_text = y_table - 14
    c.drawCentredString(mid_sno,  y_hdr_text, "S.NO.")
    c.drawCentredString(mid_svc,  y_hdr_text, "SERVICES")
    c.drawCentredString(mid_qty,  y_hdr_text, "QTY.")
    c.drawCentredString(mid_rate, y_hdr_text, "RATE")
    c.drawCentredString(mid_tax,  y_hdr_text, "TAX")
    c.drawCentredString(mid_amt,  y_hdr_text, "AMOUNT")

    # ── Compute amounts ───────────────────────────────────────────────────
    items = invoice.get("items", [])
    if not items:
        # Fallback for old invoices
        items = [{
            "service_description": invoice.get("service_description", ""),
            "qty": float(invoice.get("qty", 1.0)),
            "rate": float(invoice.get("rate", 0.0))
        }]

    base_taxable     = sum(float(it.get("qty", 1.0)) * float(it.get("rate", 0.0)) for it in items)
    gst_percent      = float(invoice.get("gst_percent", 18.0))
    tax_amount       = base_taxable * (gst_percent / 100.0)
    cgst_amount      = tax_amount / 2.0
    sgst_amount      = tax_amount / 2.0
    total_amount_with_gst = base_taxable + tax_amount

    tds_percent      = float(invoice.get("tds_percent", 10.0))
    tds_amount       = base_taxable * (tds_percent / 100.0)
    payable_amount   = total_amount_with_gst - tds_amount
    received_amount  = float(invoice.get("received_amount", 0.0))
    balance_amount   = payable_amount - received_amount

    # ── Data rows ─────────────────────────────────────────────────────────
    c.setFont("Roboto", 9)
    desc_style = ParagraphStyle(
        'InvDescStyle',
        parent=styles['Normal'],
        fontName='Roboto',
        fontSize=9,
        leading=12,
        textColor=colors.black,
    )
    current_y = y_table - 45
    
    amt_w_avail = (width - margin) - col_x[4]

    def _draw_scaled_right(cv, x, y, text, max_w, font="Roboto", base_size=9):
        sz = base_size
        while pdfmetrics.stringWidth(text, font, sz) > max_w and sz > 4.0:
            sz -= 0.5
        cv.setFont(font, sz)
        cv.drawRightString(x, y, text)
        cv.setFont(font, base_size)

    def _draw_scaled_center(cv, x, y, text, max_w, font="Roboto", base_size=9):
        sz = base_size
        while pdfmetrics.stringWidth(text, font, sz) > max_w and sz > 4.0:
            sz -= 0.5
        cv.setFont(font, sz)
        cv.drawCentredString(x, y, text)
        cv.setFont(font, base_size)
    for idx, item in enumerate(items):
        item_qty = float(item.get("qty", 1.0))
        item_rate = float(item.get("rate", 0.0))
        item_tax = (item_qty * item_rate) * (gst_percent / 100.0)
        item_total = (item_qty * item_rate) + item_tax

        c.setFillColor(colors.black)
        c.setFont("Roboto", 9)
        text_y = current_y + 6
        
        # S.NO.
        _draw_scaled_center(c, mid_sno, text_y, str(idx + 1), max_w=sno_w - 2 * mm)

        # Services
        p = Paragraph(item.get("service_description", ""), desc_style)
        p.wrap(svc_w - 4 * mm, 260)
        p.drawOn(c, margin + sno_w + 2 * mm, current_y + 13 - p.height)

        # Quantity
        _draw_scaled_center(c, mid_qty, text_y, f"{int(item_qty)} PCS", max_w=qty_w - 2 * mm)
        
        # Rate
        _draw_scaled_right(c, col_x[3] - 2 * mm, text_y, _format_inr(item_rate), max_w=rate_w - 2 * mm)
        
        # Tax
        _draw_scaled_right(c, col_x[4] - 2 * mm, text_y, _format_inr(item_tax), max_w=tax_w - 2 * mm)
        c.setFillColor(BRAND_MUTED)
        _draw_scaled_right(c, col_x[4] - 2 * mm, text_y - 12, f"({int(gst_percent)}%)", max_w=tax_w - 2 * mm, base_size=9)
        c.setFillColor(colors.black)
        
        # Amount
        _draw_scaled_right(c, width - margin - 2 * mm, text_y, _format_inr(item_total), max_w=amt_w_avail - 2 * mm)

        row_h = max(p.height, 12)
        current_y -= max(row_h + 15, 30)

    # ── TOTAL row (blue background, at bottom of table) ───────────────────
    y_total_row = y_table_bottom + hdr_h
    c.setFillColor(colors.HexColor("#CFE9FA"))
    c.rect(margin, y_table_bottom, printable_width, hdr_h, fill=1, stroke=0)
    c.setStrokeColor(colors.black)
    c.setLineWidth(1.0)
    c.line(margin, y_total_row, width - margin, y_total_row)
    for x in col_x:
        c.line(x, y_table_bottom, x, y_total_row)

    c.setFillColor(colors.black)
    c.setFont("Roboto-Bold", 9)
    # The Y position for text is relative to y_table_bottom (the bottom of this row)
    c.drawRightString(col_x[1] - 3 * mm, y_table_bottom + 6, "TOTAL")
    
    _draw_scaled_center(c, mid_qty, y_table_bottom + 6, f"{int(sum(float(it.get('qty', 1.0)) for it in items))}", max_w=qty_w - 2 * mm, font="Roboto-Bold", base_size=8.5)
    _draw_scaled_right(c, col_x[4] - 2 * mm, y_table_bottom + 6, f"₹ {_format_inr(tax_amount)}", max_w=tax_w - 2 * mm, font="Roboto-Bold", base_size=8.5)
    _draw_scaled_right(c, width - margin - 2 * mm, y_table_bottom + 6, f"₹ {_format_inr(total_amount_with_gst)}", max_w=amt_w_avail - 2 * mm, font="Roboto-Bold", base_size=8.5)

    # ── Math summary (Tax only — full-width rows below table) ─────────────
    # Attach directly to the bottom of the main table
    y_cursor = y_table_bottom

    if not is_proforma:
        tds_section = invoice.get("tds_section", "194J")
        math_row_h  = 21

        math_rows = []
        if tds_amount > 0:
            math_rows.append((f"TDS @{tds_percent:g}% {tds_section}", f"- ₹ {_format_inr(tds_amount)}",  False, False))
            
        math_rows.extend([
            ("AMOUNT PAYABLE",                         f"₹ {_format_inr(payable_amount, show_decimals=False)}", False, False),
            ("RECEIVED AMOUNT",                        f"₹ {_format_inr(received_amount)}", False, False),
            ("BALANCE AMOUNT",                         f"₹ {_format_inr(balance_amount)}",  False, False),
        ])

        n_rows        = len(math_rows)
        math_block_h  = math_row_h * n_rows
        y_math_top    = y_cursor
        y_math_bottom = y_math_top - math_block_h

        c.setLineWidth(1.0)
        c.rect(margin, y_math_bottom, printable_width, math_block_h, fill=0, stroke=1)

        y_r = y_math_top
        for label, value, highlighted, bold in math_rows:
            row_bottom = y_r - math_row_h
            if highlighted:
                c.setFillColor(colors.HexColor("#CFE9FA"))
                c.rect(margin, row_bottom, printable_width, math_row_h, fill=1, stroke=0)
                c.setStrokeColor(colors.black)
            
            c.setLineWidth(1.0)
            c.line(margin, row_bottom, width - margin, row_bottom)
            # Continue vertical lines from main table
            for x in col_x:
                c.line(x, row_bottom, x, y_r)

            c.setFillColor(colors.black)
            fn = "Roboto-Medium" if bold else "Roboto"
            c.setFont(fn, 9)
            # Label: right-aligned in SERVICES column (col_x[1])
            lbl_x = col_x[1] - 3 * mm
            # Value: right-aligned in AMOUNT column
            text_y = row_bottom + 6.5
            c.drawRightString(lbl_x, text_y, label)
            
            _draw_scaled_right(c, width - margin - 2 * mm, text_y, value, max_w=amt_w_avail - 2 * mm, font=fn, base_size=9)
            
            y_r -= math_row_h

        y_cursor = y_math_bottom

    # ── HSN Summary Table ─────────────────────────────────────────────────
    # Layout: HSN/SAC(30mm) | Taxable Value(40mm) | CGST(47mm) | SGST(47mm) | Total Tax(rest)
    hsn_h       = 42 # increased height for y-padding
    y_hsn_top   = y_cursor
    y_hsn_bot   = y_hsn_top - hsn_h
    hsn_mid     = y_hsn_bot + hsn_h / 2   # divides header row from data row

    # Blue background for header half
    c.setFillColor(colors.HexColor("#CFE9FA"))
    c.rect(margin, hsn_mid, printable_width, hsn_h / 2, fill=1, stroke=0)

    c.setStrokeColor(colors.black)
    c.setFillColor(colors.black)
    c.setLineWidth(1.0)
    c.rect(margin, y_hsn_bot, printable_width, hsn_h, fill=0, stroke=1)
    c.line(margin, hsn_mid, width - margin, hsn_mid)   # header/data divider

    # Column dividers
    hx = [
        margin + 26.5 * mm,
        margin + 26.5 * mm + 38 * mm,
        margin + 26.5 * mm + 38 * mm + 34.5 * mm,
        margin + 26.5 * mm + 38 * mm + 34.5 * mm + 34.5 * mm,
    ]
    for x in hx:
        c.line(x, y_hsn_bot, x, y_hsn_top)

    # CGST/SGST sub-dividers and horizontal split
    cgst_mid = hx[1] + 10.5 * mm
    sgst_mid = hx[2] + 10.5 * mm
    
    hdr_mid = hsn_mid + (hsn_h / 4)
    # Horizontal line under CGST and SGST
    c.line(hx[1], hdr_mid, hx[3], hdr_mid)
    
    # Vertical lines for Rate|Amount from bottom up to hdr_mid
    c.line(cgst_mid, y_hsn_bot, cgst_mid, hdr_mid)
    c.line(sgst_mid, y_hsn_bot, sgst_mid, hdr_mid)

    # Header labels (top half)
    y_hdr1 = y_hsn_top - (hsn_h / 4)   # centre of entire header
    y_hdr_cgst = y_hsn_top - (hsn_h / 8) # centre of top quarter
    
    c.setFont("Roboto", 9)
    # Full-height headers
    c.drawCentredString(margin + (hx[0] - margin) / 2,     y_hdr1 - 3.0, "HSN/SAC")
    c.drawCentredString(hx[0] + (hx[1] - hx[0]) / 2,       y_hdr1 - 3.0, "Taxable Value")
    c.drawCentredString(hx[3] + (width - margin - hx[3]) / 2, y_hdr1 - 3.0, "Total Tax Amount")
    
    # Half-height headers
    c.drawCentredString(hx[1] + (hx[2] - hx[1]) / 2,       y_hdr_cgst - 3.0, "CGST")
    c.drawCentredString(hx[2] + (hx[3] - hx[2]) / 2,       y_hdr_cgst - 3.0, "SGST")

    # CGST / SGST Rate / Amount sub-headers — in the second quarter
    c.setFont("Roboto", 9)
    rate_label_y  = hdr_mid - (hsn_h / 8) - 2.5
    # Rate is centered in the left half of the CGST/SGST column
    c.drawCentredString(hx[1] + (cgst_mid - hx[1]) / 2, rate_label_y, "Rate")
    c.drawCentredString(hx[2] + (sgst_mid - hx[2]) / 2, rate_label_y, "Rate")
    # Amount is centered in the right half of the CGST/SGST column
    c.drawCentredString(cgst_mid + (hx[2] - cgst_mid) / 2, rate_label_y, "Amount")
    c.drawCentredString(sgst_mid + (hx[3] - sgst_mid) / 2, rate_label_y, "Amount")

    # Data row (bottom half)
    y_data = y_hsn_bot + (hsn_h / 2) / 2 - 3.8   # centre of bottom half
    cgst_rate_str = f"{gst_percent / 2:.1f}%".rstrip("0").rstrip(".")  + "%"
    if not cgst_rate_str[0].isdigit():
        cgst_rate_str = f"{gst_percent / 2:g}%"
    hsn_display = invoice.get("hsn_code") or "-"
    c.setFont("Roboto", 9)
    # HSN data centered
    c.drawCentredString(margin + (hx[0] - margin) / 2,       y_data, hsn_display)
    
    # Taxable value right-aligned
    _draw_scaled_right(c, hx[1] - 2 * mm, y_data, _format_inr(base_taxable), max_w=(hx[1]-hx[0]) - 2 * mm)
    
    # Rate centered
    c.drawCentredString(hx[1] + (cgst_mid - hx[1]) / 2,     y_data, f"{gst_percent/2:g}%")
    c.drawCentredString(hx[2] + (sgst_mid - hx[2]) / 2,     y_data, f"{gst_percent/2:g}%")
    
    # Amount right-aligned
    _draw_scaled_right(c, hx[2] - 2 * mm, y_data, _format_inr(cgst_amount), max_w=(hx[2]-hx[1]) - 4 * mm)
    _draw_scaled_right(c, hx[3] - 2 * mm, y_data, _format_inr(sgst_amount), max_w=(hx[3]-hx[2]) - 4 * mm)
    
    # Total tax right-aligned
    _draw_scaled_right(c, width - margin - 2 * mm, y_data, f"₹ {_format_inr(tax_amount)}", max_w=(width - margin - hx[3]) - 2 * mm)

    # ── Total Amount in Words ─────────────────────────────────────────────
    # Attach directly to the bottom of the HSN table
    y_words = y_hsn_bot
    words_h = 30
    y_words_bot = y_words - words_h

    c.setLineWidth(1.0)
    c.rect(margin, y_words_bot, printable_width, words_h, fill=0, stroke=1)
    c.setFont("Roboto-Bold", 9)
    c.drawString(margin + 1 * mm, y_words_bot + words_h - 13, "Total Amount (in words)")
    c.setFont("Roboto", 8)
    c.drawString(margin + 1 * mm, y_words_bot + words_h - 25, num_to_words_indian(total_amount_with_gst))

    # ── Footer ────────────────────────────────────────────────────────────
    # Attach directly to the bottom of the Total Amount block
    y_footer     = y_words_bot
    footer_h     = 85
    y_footer_bot = y_footer - footer_h

    c.setLineWidth(1.0)
    c.rect(margin, y_footer_bot, printable_width, footer_h, fill=0, stroke=1)

    fc1 = margin + 74 * mm    # divides Bank | QR | Signatory
    fc2 = margin + 132 * mm
    c.line(fc1, y_footer_bot, fc1, y_footer)
    c.line(fc2, y_footer_bot, fc2, y_footer)

    fhdr_h = 13   # height of "section title" row at top of footer
    # Remove the horizontal line that separated the headers
    # c.line(margin, y_footer - fhdr_h, width - margin, y_footer - fhdr_h)

    c.setFont("Roboto-Bold", 9)
    c.drawString(margin + 1 * mm, y_footer - 9, "Bank Details")
    # "Payment QR Code" will be drawn separately with a larger font
    # c.drawString(fc2 + 1 * mm,    y_footer - 9, "Signatory Block")

    # Bank detail lines
    bl = 9.5  # bank line height (gives more y-padding)
    c.setFont("Roboto", 8.5)
    lbl_x = margin + 1 * mm
    val_x = margin + 22 * mm # offset to accommodate larger text size of labels
    c.drawString(lbl_x, y_footer - fhdr_h - 1*bl, "Name:")
    c.drawString(val_x, y_footer - fhdr_h - 1*bl, cd.get('bank_account_name',''))
    c.drawString(lbl_x, y_footer - fhdr_h - 2*bl, "IFSC Code:")
    c.drawString(val_x, y_footer - fhdr_h - 2*bl, cd.get('bank_ifsc',''))
    c.drawString(lbl_x, y_footer - fhdr_h - 3*bl, "Account No:")
    c.drawString(val_x, y_footer - fhdr_h - 3*bl, cd.get('bank_account_no',''))
    c.drawString(lbl_x, y_footer - fhdr_h - 4*bl, "Bank:")
    c.drawString(val_x, y_footer - fhdr_h - 4*bl, cd.get('bank_name',''))
    bank_branch = cd.get("bank_branch", "")
    if bank_branch:
        parts = [bank_branch[i:i+24] for i in range(0, len(bank_branch), 24)]
        for bi, bp in enumerate(parts[:4]):
            c.drawString(val_x, y_footer - fhdr_h - (5 + bi) * bl, bp)

    # ── QR section ────────────────────────────────────────────────────────
    qr_sec_x = fc1 + 1.5 * mm
    c.setFont("Roboto-Bold", 9)
    c.drawString(qr_sec_x, y_footer - 9, "Payment QR Code")

    qr_bytes = await _load_photo_bytes(cd.get("qr_code_url"))
    if qr_bytes:
        try:
            qr_image = ImageReader(io.BytesIO(qr_bytes))
            
            qr_size = 18 * mm
            # draw at right edge of the QR section
            qr_x = fc2 - qr_size - 0.5 * mm
            qr_y = y_footer_bot + (footer_h - qr_size) / 2
            c.drawImage(qr_image, qr_x, qr_y, width=qr_size, height=qr_size)
        except Exception:
            pass

    upi_id = cd.get("upi_id", "")
    c.setFont("Roboto", 8)
    c.drawString(qr_sec_x, y_footer - 30, "UPI ID:")
    c.setFont("Roboto", 7.5)
    c.drawString(qr_sec_x, y_footer - 42, upi_id)

    try:
        payment_icons_path = Path(__file__).parent / "assets" / "payment_icons_white.jpg"
        if payment_icons_path.exists():
            c.drawImage(str(payment_icons_path), qr_sec_x, y_footer - 60, width=80, height=5.2, preserveAspectRatio=True)
    except Exception:
        pass

    # Signatory (Only computer generated message)
    c.setFont("Roboto", 8)
    c.setFillColor(colors.gray)
    c.drawRightString(width - margin, y_footer_bot - 4 * mm, "This is computer generated bill,")
    c.drawRightString(width - margin, y_footer_bot - 7 * mm, "signature is not required.")
    c.setFillColor(colors.black)

    c.showPage()
    c.save()
    buf.seek(0)
    pdf_bytes = buf.read()

    # ── Merge PDF logo if logo is a PDF ───────────────────────────────────
    if logo_is_pdf and logo_bytes:
        try:
            from pypdf import PdfReader, PdfWriter, Transformation
            src      = PdfReader(io.BytesIO(pdf_bytes))
            logo_pdf = PdfReader(io.BytesIO(logo_bytes))
            if logo_pdf.pages:
                lp    = logo_pdf.pages[0]
                lw    = float(lp.mediabox.width)
                lh    = float(lp.mediabox.height)
                t     = Transformation().scale(sx=logo_size/lw, sy=logo_size/lh).translate(tx=logo_x, ty=logo_y)
                src.pages[0].merge_transformed_page(lp, t)
                writer = PdfWriter()
                for pg in src.pages:
                    writer.add_page(pg)
                out = io.BytesIO()
                writer.write(out)
                pdf_bytes = out.getvalue()
        except Exception as e:
            logger.error(f"Failed to merge PDF logo: {e}")

    return pdf_bytes



@api_router.get("/invoices/{invoice_id}/pdf")
async def serve_invoice_pdf(invoice_id: str):
    doc = await db.invoices.find_one({"id": invoice_id})
    if not doc:
        raise HTTPException(404, "Invoice not found")
        
    pdf_bytes = await _build_invoice_document_pdf(doc)
    
    fn = f"invoice_{doc.get('invoice_no', 'doc')}.pdf".replace(" ", "_").replace(">", "")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )


# Include the routers
app.include_router(auth_public_router)
app.include_router(api_router)

# Static file mount for uploaded site-visit photos. Served at /api/uploads/site-visits/<fname>
# (mounted AFTER the api_router so explicit endpoints take precedence)
app.mount("/api/uploads", StaticFiles(directory=str(UPLOAD_ROOT)), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def on_startup():
    global _photo_bucket
    _photo_bucket = AsyncIOMotorGridFSBucket(db, bucket_name="site_visit_photos")
    # Auto-seed demo data on first run
    try:
        if await db.projects.count_documents({}) == 0:
            logger.info("Seeding demo data on startup...")
            await seed_demo()
        # Separately seed offers if missing (back-compat for existing DBs)
        if await db.offers.count_documents({}) == 0:
            logger.info("Seeding demo offers...")
            clients = await db.clients.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(10)
            if clients:
                sample_offers = [
                    {"offer_type": "Audit", "custom_type": "", "client_id": clients[min(1, len(clients)-1)]["id"],
                     "description": "RCC-Basic-Audit of Row House",
                     "site_location": "Plot 44, Sector 4, Koparkhairane, Navi Mumbai",
                     "base_amount": 28000.0, "gst_percent": 18.0,
                     "file_path": "D:\\CreatorConsultant\\Offers\\2026\\STR-AUDIT-2026-023.pdf",
                     "status": "Pending", "reference_no": "STR/AUDIT/2026/023",
                     "notes": "Half cell + Rebound Hammer + Carbonation + UPV. 50% advance."},
                    {"offer_type": "Steel", "custom_type": "", "client_id": clients[min(2, len(clients)-1)]["id"],
                     "description": "MS Structural Design & Consultancy",
                     "site_location": "TTC IND. Area, Rabale MIDC, Navi Mumbai",
                     "base_amount": 200000.0, "gst_percent": 18.0,
                     "file_path": "D:\\CreatorConsultant\\Offers\\2025\\STR-QUOT-2025-160.pdf",
                     "status": "Pending", "reference_no": "STR/QUOT/2025/160",
                     "notes": "20,000 sq.ft. @ Rs 10/sq.ft."},
                    {"offer_type": "Other", "custom_type": "PMC", "client_id": clients[min(3, len(clients)-1)]["id"],
                     "description": "Project Management Consultancy",
                     "site_location": "Novo Rabale MIDC",
                     "base_amount": 150000.0, "gst_percent": 18.0,
                     "file_path": "D:\\CreatorConsultant\\Offers\\2026\\PMC-offer.pdf",
                     "status": "Pending", "reference_no": "STR/PMC/2026/005",
                     "notes": "Quarterly site visits + BOQ review"},
                ]
                for o in sample_offers:
                    od = {**o, "id": _new_id(), "offer_code": await _next_offer_code(),
                          "offer_date": _now(), "linked_project_id": None, "linked_project_code": "",
                          "created_at": _now()}
                    await _enrich_offer(od)
                    await db.offers.insert_one(od.copy())
    except Exception as e:
        logger.error(f"Seed error: {e}")

    # Seed default document types if missing
    try:
        await _seed_document_types_if_missing()
    except Exception as e:
        logger.error(f"Document types seed error: {e}")

    # Seed default site-visit inspection templates if missing
    try:
        await _seed_sv_templates_if_missing()
    except Exception as e:
        logger.error(f"Site visit templates seed error: {e}")

    # Ensure VAPID keypair exists for Web Push (auto-generates on first run)
    try:
        await _ensure_vapid_keys()
    except Exception as e:
        logger.error(f"VAPID init error: {e}")

    # Start Google Drive auto-backup scheduler
    try:
        await backup_module.start_scheduler()
    except Exception as e:
        logger.error(f"Failed to start backup scheduler: {e}")

    # Start daily housekeeping (cleanup of old read notifications)
    try:
        _start_housekeeping_scheduler()
    except Exception as e:
        logger.error(f"Housekeeping scheduler init failed: {e}")

    # Seed/refresh the admin account from .env
    try:
        await auth_module.seed_admin()
        await auth_module.backfill_user_colors()
    except Exception as e:
        logger.error(f"Admin seed failed: {e}")

    # Seed company details if missing
    try:
        await seed_company_details()
    except Exception as e:
        logger.error(f"Company details seed failed: {e}")

    # Seed invoice counters if missing
    try:
        await db.counters.update_one(
            {"_id": "proforma_invoice"},
            {"$setOnInsert": {"seq": 46}},
            upsert=True
        )
        await db.counters.update_one(
            {"_id": "tax_invoice"},
            {"$setOnInsert": {"seq": 57}},
            upsert=True
        )
    except Exception as e:
        logger.error(f"Invoice counters seed failed: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        await backup_module.stop_scheduler()
    except Exception:
        pass
    try:
        _stop_housekeeping_scheduler()
    except Exception:
        pass
    client.close()
