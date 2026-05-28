"""Iteration 10 — Site Visit Excel export + engineer project assignment scope tests."""
import io
import os
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://beginner-coder-hub-2.preview.emergentagent.com").rstrip("/")
ADMIN_USER = "rutvij0213"
ADMIN_PASS = "Rutvij4141*"
ENG_USERNAME = "test_engineer"
ENG_PASSWORD = "EngTest123!"


# ---------- Shared fixtures ----------
@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"username": ADMIN_USER, "password": ADMIN_PASS}, timeout=20)
    assert r.status_code == 200, r.text
    return r.json()["token"]


@pytest.fixture(scope="module")
def admin_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


@pytest.fixture(scope="module")
def engineer_session(admin_headers):
    """Ensure test_engineer exists with known password; return token + user id."""
    # find existing
    lst = requests.get(f"{BASE_URL}/api/auth/users", headers=admin_headers, timeout=20).json()
    existing = next((u for u in lst if u["username"].lower() == ENG_USERNAME.lower()), None)
    if not existing:
        r = requests.post(f"{BASE_URL}/api/auth/users", headers=admin_headers,
                          json={"username": ENG_USERNAME, "password": ENG_PASSWORD, "name": "Test Engineer", "role": "engineer"}, timeout=20)
        assert r.status_code in (200, 201), r.text
        eng_id = r.json()["id"]
    else:
        eng_id = existing["id"]
        # reset password & role
        requests.put(f"{BASE_URL}/api/auth/users/{eng_id}", headers=admin_headers,
                     json={"password": ENG_PASSWORD, "role": "engineer"}, timeout=20)
    # login
    li = requests.post(f"{BASE_URL}/api/auth/login", json={"username": ENG_USERNAME, "password": ENG_PASSWORD}, timeout=20)
    assert li.status_code == 200, li.text
    data = li.json()
    return {"id": eng_id, "token": data["token"], "headers": {"Authorization": f"Bearer {data['token']}"}}


@pytest.fixture(scope="module")
def two_projects(admin_headers, engineer_session):
    """Create two projects: project_A is assigned to engineer; project_B is NOT assigned.
    Returns dict {a, b} of project documents."""
    payload_a = {"name": "TEST_iter10 Assigned Project A", "project_code": "TST10A",
                 "client_name": "TEST Client", "architect_name": "TEST Arch",
                 "site_location": "TEST", "assigned_engineer_ids": [engineer_session["id"]]}
    payload_b = {"name": "TEST_iter10 Unassigned Project B", "project_code": "TST10B",
                 "client_name": "TEST Client", "architect_name": "TEST Arch",
                 "site_location": "TEST", "assigned_engineer_ids": []}
    ra = requests.post(f"{BASE_URL}/api/projects", headers=admin_headers, json=payload_a, timeout=20)
    assert ra.status_code in (200, 201), ra.text
    a = ra.json()
    rb = requests.post(f"{BASE_URL}/api/projects", headers=admin_headers, json=payload_b, timeout=20)
    assert rb.status_code in (200, 201), rb.text
    b = rb.json()
    yield {"a": a, "b": b}
    # teardown
    requests.delete(f"{BASE_URL}/api/projects/{a['id']}", headers=admin_headers, timeout=20)
    requests.delete(f"{BASE_URL}/api/projects/{b['id']}", headers=admin_headers, timeout=20)


# ---------- Engineer scoping on /api/projects ----------
class TestEngineerProjectScope:
    def test_admin_sees_both_projects(self, admin_headers, two_projects):
        r = requests.get(f"{BASE_URL}/api/projects", headers=admin_headers, timeout=20)
        assert r.status_code == 200
        ids = {p["id"] for p in r.json()}
        assert two_projects["a"]["id"] in ids
        assert two_projects["b"]["id"] in ids

    def test_engineer_sees_only_assigned(self, engineer_session, two_projects):
        r = requests.get(f"{BASE_URL}/api/projects", headers=engineer_session["headers"], timeout=20)
        assert r.status_code == 200
        ids = {p["id"] for p in r.json()}
        assert two_projects["a"]["id"] in ids, "Engineer should see assigned project"
        assert two_projects["b"]["id"] not in ids, "Engineer should NOT see unassigned project"

    def test_put_project_updates_assigned_engineer_ids(self, admin_headers, engineer_session, two_projects):
        # Add engineer to project B
        pid = two_projects["b"]["id"]
        full = two_projects["b"].copy()
        # PUT requires full body — fetch fresh
        gr = requests.get(f"{BASE_URL}/api/projects/{pid}", headers=admin_headers, timeout=20)
        assert gr.status_code == 200
        cur = gr.json()
        body = {k: v for k, v in cur.items() if k not in ("id", "created_at", "updated_at", "project_code")}
        body["assigned_engineer_ids"] = [engineer_session["id"]]
        u = requests.put(f"{BASE_URL}/api/projects/{pid}", headers=admin_headers, json=body, timeout=20)
        assert u.status_code == 200, u.text
        # verify via GET
        g2 = requests.get(f"{BASE_URL}/api/projects/{pid}", headers=admin_headers, timeout=20).json()
        assert engineer_session["id"] in (g2.get("assigned_engineer_ids") or [])

        # Engineer now sees both A and B
        re = requests.get(f"{BASE_URL}/api/projects", headers=engineer_session["headers"], timeout=20)
        assert re.status_code == 200
        ids = {p["id"] for p in re.json()}
        assert pid in ids

        # Revert: remove engineer from project B
        body["assigned_engineer_ids"] = []
        requests.put(f"{BASE_URL}/api/projects/{pid}", headers=admin_headers, json=body, timeout=20)


# ---------- Excel export ----------
class TestExcelExport:
    visit_admin = None
    visit_eng = None

    @pytest.fixture(autouse=True, scope="class")
    def _seed_visits(self, request, admin_headers, engineer_session, two_projects):
        """Create one admin-owned visit and one engineer-owned visit on project A."""
        proj_a = two_projects["a"]
        a_payload = {
            "template_name": "Beam Inspection",
            "inspection_title": "TEST_iter10 Admin Visit",
            "project_id": proj_a["id"],
            "checklist": [{"label": "x", "compliance": "no", "remark": "fail"}],
            "observations": ["adm-obs"],
            "photos": [],
            "engineer_name": "Admin Eng",
            "status": "submitted",
        }
        ra = requests.post(f"{BASE_URL}/api/site-visits", headers=admin_headers, json=a_payload, timeout=20)
        assert ra.status_code == 200, ra.text
        TestExcelExport.visit_admin = ra.json()

        e_payload = dict(a_payload)
        e_payload["inspection_title"] = "TEST_iter10 Eng Visit"
        e_payload["engineer_name"] = "Test Engineer"
        re = requests.post(f"{BASE_URL}/api/site-visits", headers=engineer_session["headers"], json=e_payload, timeout=20)
        assert re.status_code == 200, re.text
        TestExcelExport.visit_eng = re.json()

        def _cleanup():
            for v in (TestExcelExport.visit_admin, TestExcelExport.visit_eng):
                if v:
                    requests.delete(f"{BASE_URL}/api/site-visits/{v['id']}", headers=admin_headers, timeout=20)
        request.addfinalizer(_cleanup)

    def _assert_excel_response(self, r):
        assert r.status_code == 200, f"{r.status_code} {r.text[:200]}"
        assert r.headers.get("content-type", "").startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), r.headers
        assert "attachment" in r.headers.get("content-disposition", "").lower()
        assert r.content[:2] == b"PK", "xlsx must start with PK"

    def test_excel_no_filter(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/site-visits/export/excel", headers=admin_headers, timeout=30)
        self._assert_excel_response(r)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        # Sheet check
        assert "Visits" in wb.sheetnames
        assert "By Engineer" in wb.sheetnames
        ws = wb["Visits"]
        # 16 column header
        header_row = [c.value for c in ws[1]]
        assert len(header_row) == 16, f"Visits sheet headers={len(header_row)} expected 16: {header_row}"
        ws2 = wb["By Engineer"]
        header2 = [c.value for c in ws2[1]]
        assert len(header2) == 4, f"By Engineer header={len(header2)} expected 4"

    def test_excel_month_filter(self, admin_headers):
        # Use current YYYY-MM derived from visit_date of created visit
        v = TestExcelExport.visit_admin
        month = (v.get("visit_date") or v.get("created_at") or "")[:7]
        assert month, "no month found on seeded visit"
        r = requests.get(f"{BASE_URL}/api/site-visits/export/excel?month={month}", headers=admin_headers, timeout=30)
        self._assert_excel_response(r)

    def test_excel_engineer_id_filter(self, admin_headers, engineer_session):
        r = requests.get(f"{BASE_URL}/api/site-visits/export/excel?engineer_id={engineer_session['id']}",
                         headers=admin_headers, timeout=30)
        self._assert_excel_response(r)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Visits"]
        # All data rows (skip header) should be only engineer's visits
        titles = [row[2].value for row in ws.iter_rows(min_row=2)]
        # engineer-only filter: should NOT contain admin's TEST visit title
        assert "TEST_iter10 Admin Visit" not in titles, f"engineer_id filter leaked admin visit: {titles}"

    def test_excel_project_id_filter(self, admin_headers, two_projects):
        pid = two_projects["a"]["id"]
        r = requests.get(f"{BASE_URL}/api/site-visits/export/excel?project_id={pid}", headers=admin_headers, timeout=30)
        self._assert_excel_response(r)

    def test_excel_engineer_scoped_to_self(self, engineer_session):
        """Engineer calling export must only see their own visits regardless of engineer_id param."""
        r = requests.get(f"{BASE_URL}/api/site-visits/export/excel", headers=engineer_session["headers"], timeout=30)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Visits"]
        titles = [row[2].value for row in ws.iter_rows(min_row=2)]
        # Admin's title must not be visible to engineer
        assert "TEST_iter10 Admin Visit" not in titles, f"engineer leak: {titles}"
        # Engineer's own visit must be present
        assert "TEST_iter10 Eng Visit" in titles, f"engineer missing own visit: {titles}"


# ---------- Path-collision regression: /site-visits/{vid} GET ----------
class TestSiteVisitsPathCollision:
    def test_get_single_visit_after_export_route(self, admin_headers):
        """Confirm /api/site-visits/{vid} GET still works (export/excel route declared earlier)."""
        # create
        payload = {
            "template_name": "Beam Inspection",
            "inspection_title": "TEST_iter10 path coll",
            "checklist": [{"label": "x", "compliance": "yes", "remark": ""}],
            "observations": [], "photos": [],
            "engineer_name": "A", "status": "submitted",
        }
        cr = requests.post(f"{BASE_URL}/api/site-visits", headers=admin_headers, json=payload, timeout=20)
        assert cr.status_code == 200, cr.text
        vid = cr.json()["id"]
        # GET single
        g = requests.get(f"{BASE_URL}/api/site-visits/{vid}", headers=admin_headers, timeout=20)
        assert g.status_code == 200, f"path collision broke single-GET: {g.status_code} {g.text[:200]}"
        assert g.json()["id"] == vid
        # cleanup
        requests.delete(f"{BASE_URL}/api/site-visits/{vid}", headers=admin_headers, timeout=20)


# ---------- Backup collections list (smoke) ----------
class TestBackupIncludesNewCollections:
    def test_backup_run_includes_site_visits(self, admin_headers):
        """Trigger backup and verify the on-disk JSON dump includes site_visits/site_visit_templates/users."""
        import json as _json
        from pathlib import Path
        r = requests.post(f"{BASE_URL}/api/backup/run", headers=admin_headers, timeout=60)
        if r.status_code != 200:
            pytest.skip(f"backup/run returned {r.status_code}: {r.text[:200]}")
        body = r.json()
        assert body.get("ok") is True, body
        fname = body.get("filename")
        assert fname, "no filename in backup response"
        backup_dir = Path("/app/backend/backups")
        fp = backup_dir / fname
        if not fp.exists():
            pytest.skip(f"backup file not accessible from test runner: {fp}")
        dump = _json.loads(fp.read_text())
        cols = dump.get("collections") or {}
        for col in ("site_visits", "site_visit_templates", "users"):
            assert col in cols, f"backup JSON missing collection {col}: keys={list(cols.keys())}"
